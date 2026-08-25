"""从中金周报算酒店月度的算法测试。

这套算法是本次沉淀的核心 knowhow，必须钉住三件事：
1. 跨月的周按**天**归属，不整周归一个月；
2. ADR 是加权算的，**不能**按天平均——按天平均会给出不同的数；
3. 基期用表里的「上年」列（可比周），不是去年同月日历。

固定件用 2026 年 6–8 月的真实周数据（中金表 Mainland China (STR) tab，r293–r303）。
它们算出的结果已被券商报告独立印证：6 月 -4%/+2%/-2%、7 月 -3%/-1%/-4%。
"""

from __future__ import annotations

import datetime as dt
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

import openpyxl

from modules.industry_data import str_source

# (周区间, OccTY, OccLY, ADRTY, ADRLY, RevTY, RevLY, yoyOcc, yoyADR, yoyRev)
REAL_WEEKS = [
    ("2026/05/31-2026/06/06", 58.889, 58.149, 372.51, 366.26, 219.36, 212.97, 0.012725, 0.017058, 0.03),
    ("2026/06/07-2026/06/13", 63.437, 65.114, 371.63, 364.02, 235.75, 237.03, -0.025751, 0.020898, -0.0053913),
    ("2026/06/14-2026/06/20", 61.674, 67.040, 387.52, 367.13, 239.00, 246.12, -0.080032, 0.055560, -0.028919),
    ("2026/06/21-2026/06/27", 63.690, 67.607, 381.21, 381.70, 242.79, 258.06, -0.057941, -0.0012693, -0.059137),
    ("2026/06/28-2026/07/04", 62.170, 64.785, 382.72, 387.97, 237.94, 251.35, -0.040358, -0.013553, -0.053364),
    ("2026/07/05-2026/07/11", 66.673, 70.056, 389.65, 396.75, 259.79, 277.95, -0.048291, -0.017881, -0.065309),
    ("2026/07/12-2026/07/18", 69.331, 72.364, 392.58, 397.78, 272.18, 287.85, -0.041920, -0.013064, -0.054436),
    ("2026/07/19-2026/07/25", 71.400, 72.300, 397.24, 400.20, 283.69, 289.50, -0.012448, -0.0073963, -0.020069),
    ("2026/07/26-2026/08/01", 71.197, 71.847, 407.53, 407.51, 290.15, 292.78, -0.0090431, 6.5757e-05, -0.0089779),
    ("2026/08/02-2026/08/08", 74.672, 75.851, 420.15, 419.76, 313.74, 318.39, -0.015548, 0.0009293, -0.014633),
    ("2026/08/09-2026/08/15", 74.594, 77.191, 416.77, 419.87, 310.88, 324.10, -0.033643, -0.0073863, -0.040781),
]


def build_source(path: Path, rows=REAL_WEEKS, *, sheet_name=str_source.SHEET) -> Path:
    """按中金表的列位造一份最小输入。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 表头四行（内容不参与解析，只占位）
    ws.cell(3, 5, "Occ %")
    ws.cell(4, 5, "当前年度")
    ws.cell(4, 6, "上年")

    for index, row in enumerate(rows, start=5):
        ws.cell(index, str_source.COL_WEEK, row[0])
        for offset, col in enumerate((
            str_source.COL_OCC_TY, str_source.COL_OCC_LY,
            str_source.COL_ADR_TY, str_source.COL_ADR_LY,
            str_source.COL_REV_TY, str_source.COL_REV_LY,
            str_source.COL_YOY_OCC, str_source.COL_YOY_ADR, str_source.COL_YOY_REV,
        )):
            ws.cell(index, col, row[offset + 1])
    wb.save(path)
    return path


class StrSourceCase(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = TemporaryDirectory()
        self.src = build_source(Path(self._tmp.name) / "cicc.xlsx")
        self.weeks = str_source.load(self.src)

    def tearDown(self) -> None:
        self._tmp.cleanup()


class TestLoad(StrSourceCase):
    def test_reads_all_weeks(self):
        self.assertEqual(len(self.weeks), len(REAL_WEEKS))

    def test_parses_dates(self):
        first = self.weeks[0]
        self.assertEqual(first.start, dt.date(2026, 5, 31))
        self.assertEqual(first.end, dt.date(2026, 6, 6))

    def test_short_label_matches_workbook_axis(self):
        """底稿周轴写法：无前导零、不含年份。"""
        self.assertEqual(self.weeks[-1].short_label, "8/9-8/15")
        self.assertEqual(self.weeks[4].short_label, "6/28-7/4")

    def test_missing_sheet_raises(self):
        with TemporaryDirectory() as tmp:
            path = build_source(Path(tmp) / "wrong.xlsx", sheet_name="Some Other Tab")
            with self.assertRaises(str_source.StrSourceError) as ctx:
                str_source.load(path)
            self.assertIn(str_source.SHEET, str(ctx.exception))

    def test_no_week_rows_raises(self):
        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "empty.xlsx"
            wb = openpyxl.Workbook()
            wb.active.title = str_source.SHEET
            wb.save(path)
            with self.assertRaises(str_source.StrSourceError):
                str_source.load(path)


class TestDayAttribution(StrSourceCase):
    """跨月的周必须按天拆开，不能整周归给一个月。"""

    def test_week_spanning_two_months(self):
        week = next(w for w in self.weeks if w.short_label == "6/28-7/4")
        self.assertEqual(week.days_in(2026, 6), 3)   # 6/28、29、30
        self.assertEqual(week.days_in(2026, 7), 4)   # 7/1–7/4

    def test_week_fully_inside(self):
        week = next(w for w in self.weeks if w.short_label == "7/5-7/11")
        self.assertEqual(week.days_in(2026, 7), 7)
        self.assertEqual(week.days_in(2026, 6), 0)

    def test_july_days_add_up(self):
        total = sum(w.days_in(2026, 7) for w in self.weeks)
        self.assertEqual(total, 31)

    def test_june_days_add_up(self):
        total = sum(w.days_in(2026, 6) for w in self.weeks)
        self.assertEqual(total, 30)


class TestMonthlyAggregate(StrSourceCase):
    def test_july_matches_broker_published(self):
        """券商发布 7 月 -3%/-1%/-4%，独立印证本算法。"""
        agg = str_source.aggregate_month(self.weeks, 2026, 7)
        self.assertTrue(agg.complete)
        yoy = agg.yoy()
        self.assertAlmostEqual(yoy["hotelOccupancy"], -0.0299, places=3)
        self.assertAlmostEqual(yoy["hotelADR"], -0.0100, places=3)
        self.assertAlmostEqual(yoy["hotelRevPAR"], -0.0396, places=3)

    def test_june_matches_broker_published(self):
        """券商发布 6 月 -4%/+2%/-2%。"""
        agg = str_source.aggregate_month(self.weeks, 2026, 6)
        yoy = agg.yoy()
        self.assertAlmostEqual(yoy["hotelOccupancy"], -0.0413, places=3)
        self.assertAlmostEqual(yoy["hotelADR"], 0.0188, places=3)
        self.assertAlmostEqual(yoy["hotelRevPAR"], -0.0232, places=3)

    def test_adr_is_not_plain_daily_average(self):
        """ADR 必须按「总收入/总售出房晚」算。

        若错误地按天平均 ADR，7 月会得到另一个数——这个测试就是防止有人"简化"成平均。
        """
        agg = str_source.aggregate_month(self.weeks, 2026, 7)
        naive_ty = sum(w.days_in(2026, 7) * w.adr_ty for w in self.weeks) / 31
        self.assertNotAlmostEqual(agg.adr_ty, naive_ty, places=2)

    def test_revpar_identity_holds(self):
        """RevPAR ≈ Occ% × ADR，聚合后仍应成立。"""
        agg = str_source.aggregate_month(self.weeks, 2026, 7)
        self.assertAlmostEqual(agg.rev_ty, agg.occ_ty / 100 * agg.adr_ty, places=2)
        self.assertAlmostEqual(agg.rev_ly, agg.occ_ly / 100 * agg.adr_ly, places=2)

    def test_incomplete_month_is_flagged(self):
        """8 月只到 8/15，不完整——不能当成月度值入库。"""
        agg = str_source.aggregate_month(self.weeks, 2026, 8)
        self.assertIsNotNone(agg)
        self.assertFalse(agg.complete)
        self.assertEqual(agg.expected_days, 31)

    def test_complete_months_excludes_partial(self):
        months = [a.month for a in str_source.complete_months(self.weeks, 2026)]
        self.assertIn(6, months)
        self.assertIn(7, months)
        self.assertNotIn(8, months, "8 月未走完，不该出现在完整月列表里")

    def test_month_without_data_returns_none(self):
        self.assertIsNone(str_source.aggregate_month(self.weeks, 2026, 1))


class TestWeeklyPassthrough(StrSourceCase):
    """周度直接取表里的 K/L/M——已验证与底稿手抄值逐位一致。"""

    def test_weekly_values_are_taken_as_is(self):
        rows = dict(str_source.weekly_yoy(self.weeks, 2026))
        self.assertAlmostEqual(rows["8/9-8/15"]["hotelOccupancy"], -0.033643, places=6)
        self.assertAlmostEqual(rows["8/9-8/15"]["hotelADR"], -0.0073863, places=6)
        self.assertAlmostEqual(rows["8/9-8/15"]["hotelRevPAR"], -0.040781, places=6)

    def test_week_assigned_by_end_year(self):
        """跨年周按结束日归年，与底稿周轴一致。"""
        with TemporaryDirectory() as tmp:
            rows = [("2025/12/28-2026/01/03", 50.0, 52.0, 300.0, 310.0, 150.0, 161.2,
                     -0.038, -0.032, -0.069)]
            path = build_source(Path(tmp) / "wrap.xlsx", rows=rows)
            weeks = str_source.load(path)
            self.assertEqual([lab for lab, _ in str_source.weekly_yoy(weeks, 2026)], ["12/28-1/3"])
            self.assertEqual(str_source.weekly_yoy(weeks, 2025), [])


class TestPlanTolerance(unittest.TestCase):
    """手填整数与重算值的末位差异不该报成冲突。"""

    def test_integer_vs_computed_is_not_a_conflict(self):
        from modules.industry_data.str_plan import Cell

        cell = Cell("D174", "2026年7月", "ADR", old=-0.01, new=-0.010046390766123015)
        self.assertEqual(cell.kind, "一致")

    def test_real_difference_is_a_conflict(self):
        from modules.industry_data.str_plan import Cell

        cell = Cell("E173", "2026年6月", "RevPAR", old=-0.004, new=-0.0232)
        self.assertEqual(cell.kind, "冲突")

    def test_empty_cell_is_an_addition(self):
        from modules.industry_data.str_plan import Cell

        cell = Cell("C174", "2026年7月", "入住率", old=None, new=-0.0299)
        self.assertEqual(cell.kind, "新增")

    def test_borderline_half_basis_point(self):
        """0.005 个百分点是阈值本身：略小算一致，略大算冲突。"""
        from modules.industry_data.str_plan import Cell

        self.assertEqual(Cell("X1", "w", "m", old=-0.0300, new=-0.030049).kind, "一致")
        self.assertEqual(Cell("X1", "w", "m", old=-0.0300, new=-0.030051).kind, "冲突")


if __name__ == "__main__":
    unittest.main()


class TestWriteRules(unittest.TestCase):
    """写入规则：仅填空、不覆盖、写前归档、底稿被占用时拒写。

    「人手填的数字最高权威」是这条规则的理由——周度聚合只是一种取数方式，
    自动化没有资格改 IR 经理填过的值（那些值背后可能有券商口径、预测、临时判断）。
    """

    def setUp(self) -> None:
        self._tmp = TemporaryDirectory()
        root = Path(self._tmp.name)
        (root / "workbench").mkdir()
        (root / "docs").mkdir()
        (root / "docs" / "GLOSSARY.md").write_text("x", encoding="utf-8")

        from workbench.paths import Paths

        from modules.industry_data.paths import DomainPaths

        self.base = Paths(root)
        self.paths = DomainPaths(self.base)
        self.source = build_source(root / "cicc.xlsx")
        self.workbook = self._build_workbook(root / "book.xlsx")

    def tearDown(self) -> None:
        self._tmp.cleanup()

    @staticmethod
    def _build_workbook(path: Path, *, july=(None, None, -0.05), weeks=()) -> Path:
        """造一份最小底稿：2026 块 + 12 个月份行 + 右侧周轴（可预置周次行）。"""
        from modules.industry_data import layout

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = layout.SHEET
        year_row = 5
        ws.cell(year_row, 2, "2026年")
        header = year_row + layout.OFF_HEADER
        ws.cell(header, 18, "周")
        start = year_row + layout.OFF_MONTH_START
        for index in range(12):
            ws.cell(start + index, 2, "7月 (preliminary)" if index == 6 else f"{index + 1}月")
        for col, value in zip((3, 4, 5), july):
            if value is not None:
                ws.cell(start + 6, col, value)
        # 6 月填满，用于验证「已有值不被覆盖」
        for col, value in zip((3, 4, 5), (-0.025, 0.022, -0.004)):
            ws.cell(start + 5, col, value)
        # 预置周次行：R 列周标签 + S/T/U 值
        for offset, (label, values) in enumerate(weeks):
            row = header + 1 + offset
            ws.cell(row, 18, label)
            for col, value in zip((19, 20, 21), values):
                if value is not None:
                    ws.cell(row, col, value)
        wb.save(path)
        return path

    @property
    def _all_week_rows(self):
        """REAL_WEEKS 里归属 2026 年的周次及其 K/L/M 值。"""
        from modules.industry_data import str_source

        return [
            (label, (v["hotelOccupancy"], v["hotelADR"], v["hotelRevPAR"]))
            for label, v in str_source.weekly_yoy(str_source.load(self.source), 2026)
        ]

    def test_dry_run_lists_only_empty_cells(self):
        from modules.industry_data import str_write

        result = str_write.run(self.paths, self.workbook, self.source, 2026)
        self.assertEqual(result.status, "partial")
        self.assertEqual(result.data["additions"], 2)  # 7 月入住率与 ADR
        self.assertIn("未写入", result.summary)
        # 合成底稿：2026 块在 r5，月份行自 r8 起，7 月为第 7 个 → r14
        details = " ".join(c["detail"] for c in result.checks if c["name"] == "待写入")
        self.assertIn("C14", details)
        self.assertIn("D14", details)
        self.assertNotIn("E14", details, "E 列已有 -5%，不该出现在待写入里")

    def test_dry_run_does_not_touch_workbook(self):
        from modules.industry_data import str_write

        before = self.workbook.read_bytes()
        str_write.run(self.paths, self.workbook, self.source, 2026)
        self.assertEqual(self.workbook.read_bytes(), before)

    def test_nothing_to_do_when_all_present(self):
        """月度填满、周次齐全时应当无事可做。"""
        from modules.industry_data import str_write

        book = self._build_workbook(
            self.workbook.parent / "full.xlsx",
            july=(-0.03, -0.01, -0.04),
            weeks=self._all_week_rows,
        )
        result = str_write.run(self.paths, book, self.source, 2026)
        self.assertEqual(result.status, "success")
        self.assertIn("没有要加的周次", result.summary)

    def test_missing_weeks_are_detected_in_order(self):
        """底稿缺后面几周时，须按时间先后列出——加行顺序错了表就乱了。"""
        from modules.industry_data import str_plan

        rows = self._all_week_rows
        book = self._build_workbook(
            self.workbook.parent / "partial.xlsx", weeks=rows[:-3]
        )
        _cells, _notes, new_weeks = str_plan.build(book, self.source, 2026)
        self.assertEqual([w.label for w in new_weeks], [label for label, _ in rows[-3:]])

    def test_new_week_carries_its_values(self):
        from modules.industry_data import str_plan

        rows = self._all_week_rows
        book = self._build_workbook(self.workbook.parent / "p2.xlsx", weeks=rows[:-1])
        _cells, _notes, new_weeks = str_plan.build(book, self.source, 2026)
        self.assertEqual(len(new_weeks), 1)
        week = new_weeks[0]
        self.assertEqual(week.label, "8/9-8/15")
        self.assertAlmostEqual(week.values["hotelOccupancy"], -0.033643, places=6)
        self.assertAlmostEqual(week.values["hotelRevPAR"], -0.040781, places=6)

    def test_dry_run_reports_new_weeks_without_touching_file(self):
        from modules.industry_data import str_write

        rows = self._all_week_rows
        book = self._build_workbook(self.workbook.parent / "p3.xlsx", weeks=rows[:-2])
        before = book.read_bytes()
        result = str_write.run(self.paths, book, self.source, 2026)
        self.assertEqual(result.status, "partial")
        self.assertEqual(result.data["new_weeks"], [label for label, _ in rows[-2:]])
        self.assertIn("新建 2 个周次", result.summary)
        self.assertEqual(book.read_bytes(), before)
        detail = " ".join(c["detail"] for c in result.checks if c["name"] == "将新建周次")
        self.assertIn("粗底框下移", detail)

    def test_locked_workbook_is_blocked(self):
        """底稿在 Excel 里打开时必须拒写，否则会与人的编辑打架。"""
        from modules.industry_data import str_write

        before = self.workbook.read_bytes()
        lock = self.workbook.parent / f"~${self.workbook.name}"
        lock.write_bytes(b"lock")
        result = str_write.run(self.paths, self.workbook, self.source, 2026, yes=True)
        self.assertEqual(result.status, "blocked")
        self.assertIn("Excel", result.summary)
        self.assertEqual(self.workbook.read_bytes(), before, "被拒写时底稿不该有任何改动")
        self.assertFalse(
            list(self.paths.workbook_archive_dir.glob("*.xlsx")),
            "拒写时不该留下归档文件",
        )

    def test_archive_keeps_original(self):
        from modules.industry_data import str_write

        original = self.workbook.read_bytes()
        backup = str_write.archive(self.paths, self.workbook, "str")
        self.assertTrue(backup.is_file())
        self.assertEqual(backup.read_bytes(), original)
        self.assertIn("pre-str-", backup.name)
        self.assertEqual(backup.parent, self.paths.workbook_archive_dir)

    def test_archive_does_not_overwrite_previous(self):
        """归档只增不改——往期版本要能一直留着。"""
        from modules.industry_data import str_write

        first = str_write.archive(self.paths, self.workbook, "str")
        second = str_write.archive(self.paths, self.workbook, "manual")
        self.assertNotEqual(first.name, second.name)
        self.assertTrue(first.is_file())
        self.assertEqual(len(list(self.paths.workbook_archive_dir.glob("*.xlsx"))), 2)


class TestNewWeekOrdering(unittest.TestCase):
    """加行只能加在末尾，所以只有晚于当前末周的周次才可新建。

    真实案例：跨年周 `12/28-1/3` 在中金表里归 2026 年，但底稿周轴有意从 `1/4-1/10` 起。
    把它追加到末尾会让整张表的时间顺序错乱——只提醒，不动手。
    """

    def setUp(self) -> None:
        self._tmp = TemporaryDirectory()
        root = Path(self._tmp.name)
        self.source = build_source(root / "cicc.xlsx", rows=[
            ("2025/12/28-2026/01/03", 50.0, 52.0, 300.0, 310.0, 150.0, 161.2, -0.038, -0.032, -0.069),
            ("2026/01/04-2026/01/10", 51.0, 53.0, 305.0, 312.0, 155.6, 165.4, -0.037, -0.022, -0.059),
            ("2026/01/11-2026/01/17", 52.0, 54.0, 308.0, 314.0, 160.2, 169.6, -0.037, -0.019, -0.055),
        ])
        self.root = root

    def tearDown(self) -> None:
        self._tmp.cleanup()

    def _book(self, weeks) -> Path:
        return TestWriteRules._build_workbook(self.root / "book.xlsx", weeks=weeks)

    def test_earlier_gap_is_reported_not_added(self):
        from modules.industry_data import str_plan

        # 底稿只有 1/4-1/10 与 1/11-1/17，缺的是更早的 12/28-1/3
        book = self._book([("1/4-1/10", (-0.037, -0.022, -0.059)),
                           ("1/11-1/17", (-0.037, -0.019, -0.055))])
        _cells, notes, new_weeks = str_plan.build(book, self.source, 2026)
        self.assertEqual(new_weeks, [], "早于末周的缺口不该被自动加行")
        self.assertTrue(any("12/28-1/3" in n and "不自动插入" in n for n in notes))

    def test_later_week_is_added(self):
        from modules.industry_data import str_plan

        book = self._book([("12/28-1/3", (-0.038, -0.032, -0.069)),
                           ("1/4-1/10", (-0.037, -0.022, -0.059))])
        _cells, _notes, new_weeks = str_plan.build(book, self.source, 2026)
        self.assertEqual([w.label for w in new_weeks], ["1/11-1/17"])

    def test_empty_axis_adds_everything_in_order(self):
        """周轴完全空时（新年度块）按时间顺序全部新建。"""
        from modules.industry_data import str_plan

        book = self._book([])
        _cells, _notes, new_weeks = str_plan.build(book, self.source, 2026)
        self.assertEqual(
            [w.label for w in new_weeks], ["12/28-1/3", "1/4-1/10", "1/11-1/17"]
        )
