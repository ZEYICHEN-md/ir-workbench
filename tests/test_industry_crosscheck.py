"""周度填写核对的测试。

这道检查存在的理由是一个静默失败面：右侧 W/X/Y 优先于左侧 QTD G/H/I，右侧填错时
左侧的正确值不会生效，也不报错。所以检查本身必须被验证真的能抓到不一致。
"""

from __future__ import annotations

import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

import openpyxl

from modules.industry_data import crosscheck, excel, layout

YEAR_ROW = 5

# 右侧周度区列号
COL_WEEK = 18
COL_OCC, COL_ADR, COL_REVPAR = 19, 20, 21
COL_PAX, COL_TICKET, COL_FLIGHT = 23, 24, 25

WEEKS = ["8/2-8/8", "8/9-8/15", "8/16-8/22"]


def build(
    path: Path,
    *,
    right: list[tuple | None] | None = None,
    left: list[tuple | None] | None = None,
    left_labels: list[str] | None = None,
    hotel: list[tuple | None] | None = None,
) -> Path:
    """造一份最小底稿。`right` / `left` 为每周的 (客运量, 票价, 航班量)。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = layout.SHEET

    ws.cell(YEAR_ROW, 2, "2026年")
    ws.cell(YEAR_ROW, 18, "2026年")

    group = YEAR_ROW + layout.OFF_GROUP
    for col, text in {**layout.LEFT_GROUPS, **layout.RIGHT_GROUPS}.items():
        ws.cell(group, col, text)

    header = YEAR_ROW + layout.OFF_HEADER
    ws.cell(header, 2, "月度")
    for col, (text, _field) in {**layout.LEFT_COLUMNS, **layout.RIGHT_COLUMNS}.items():
        ws.cell(header, col, text)

    # 左侧月份 / 季度骨架（读表要求存在，值无关）
    start = YEAR_ROW + layout.OFF_MONTH_START
    for index in range(12):
        ws.cell(start + index, 2, f"{index + 1}月")
    for index, quarter in enumerate(("Q1", "Q2", "Q3", "Q4")):
        ws.cell(start + 12 + index, 2, quarter)

    # 左侧「QTD周度」块
    qtd = start + 16
    ws.cell(qtd, 2, "QTD周度")
    labels = left_labels or WEEKS
    for index, label in enumerate(labels):
        ws.cell(qtd + 1 + index, 2, label)
        values = (left or [None] * len(labels))[index]
        if values:
            for offset, value in enumerate(values):
                if value is not None:
                    ws.cell(qtd + 1 + index, 7 + offset, value)

    # 右侧周度区
    for index, label in enumerate(WEEKS):
        row = header + 1 + index
        ws.cell(row, COL_WEEK, label)
        occ = (hotel or [(0.01, 0.02, 0.03)] * len(WEEKS))[index]
        if occ:
            for col, value in zip((COL_OCC, COL_ADR, COL_REVPAR), occ):
                if value is not None:
                    ws.cell(row, col, value)
        values = (right or [None] * len(WEEKS))[index]
        if values:
            for col, value in zip((COL_PAX, COL_TICKET, COL_FLIGHT), values):
                if value is not None:
                    ws.cell(row, col, value)

    wb.save(path)
    return path


def by_name(rows: list[dict]) -> dict[str, dict]:
    return {row["name"]: row for row in rows}


class CrosscheckCase(unittest.TestCase):
    def run_checks(self, **kwargs) -> dict[str, dict]:
        with TemporaryDirectory() as tmp:
            path = build(Path(tmp) / "wb.xlsx", **kwargs)
            return by_name(crosscheck.checks(path))


class TestTwoSides(CrosscheckCase):
    def test_identical_sides_pass(self):
        same = [(0.05, 0.02, 0.03)] * 3
        rows = self.run_checks(right=same, left=same)
        self.assertEqual(rows["航空左右核对"]["level"], "ok")

    def test_mismatch_is_reported(self):
        rows = self.run_checks(
            right=[(0.05, 0.02, 0.03)] * 3,
            left=[(0.05, 0.02, 0.03), (0.09, 0.02, 0.03), (0.05, 0.02, 0.03)],
        )
        row = rows["航空左右核对"]
        self.assertEqual(row["level"], "warn")
        self.assertIn("8/9-8/15", row["detail"])
        self.assertIn("客运量", row["detail"])
        # 必须说清看板取的是哪边，否则用户不知道该改哪一侧
        self.assertIn("右侧", row["advice"])

    def test_one_side_empty_is_not_a_mismatch(self):
        """右侧为空、左侧有值是正常的回退，不是冲突。"""
        rows = self.run_checks(
            right=[(0.05, 0.02, 0.03), None, (0.05, 0.02, 0.03)],
            left=[(0.05, 0.02, 0.03)] * 3,
        )
        self.assertEqual(rows["航空左右核对"]["level"], "ok")

    def test_no_left_block_data_is_ok(self):
        rows = self.run_checks(right=[(0.05, 0.02, 0.03)] * 3)
        self.assertEqual(rows["航空左右核对"]["level"], "ok")

    def test_label_mismatch_flagged_only_when_fallback_needed(self):
        """右侧缺格 + 左侧配不上 = 这周真的会变空，必须报。"""
        rows = self.run_checks(
            right=[(0.05, 0.02, 0.03), (0.05, 0.02, 0.03), None],
            left=[(0.05, 0.02, 0.03)] * 3,
            left_labels=["8/2-8/8", "8/9-8/15", "8/16~8/22"],
        )
        self.assertEqual(rows["周标签配对"]["level"], "warn")
        self.assertIn("8/16-8/22", rows["周标签配对"]["detail"])

    def test_label_mismatch_ignored_when_right_side_complete(self):
        """右侧填满时左侧配不上不影响结果，报出来只是每周固定噪音。

        真实底稿里 6/21-6/27 就是这种情况：右侧三项齐全，左侧没有那一周。
        """
        rows = self.run_checks(
            right=[(0.05, 0.02, 0.03)] * 3,
            left=[(0.05, 0.02, 0.03)] * 3,
            left_labels=["8/2-8/8", "8/9-8/15", "8/16~8/22"],
        )
        self.assertNotIn("周标签配对", rows)

    def test_leading_zero_still_matches(self):
        """`8/09-8/15` 与 `8/9-8/15` 是同一周，归一后应配得上。"""
        rows = self.run_checks(
            right=[(0.05, 0.02, 0.03)] * 3,
            left=[(0.05, 0.02, 0.03)] * 3,
            left_labels=["8/02-8/08", "8/09-8/15", "8/16-8/22"],
        )
        self.assertNotIn("周标签配对", rows)
        self.assertEqual(rows["航空左右核对"]["level"], "ok")

    def test_only_recent_weeks_are_compared(self):
        """历史周次不进核对范围，否则每周 doctor 都在翻旧账。"""
        self.assertEqual(crosscheck.RECENT_WEEKS, 8)


class TestLatestWeekCompleteness(CrosscheckCase):
    def test_complete_week_passes(self):
        rows = self.run_checks(right=[(0.05, 0.02, 0.03)] * 3)
        self.assertEqual(rows["最新周填写"]["level"], "ok")
        self.assertIn("8/16-8/22", rows["最新周填写"]["detail"])

    def test_missing_aviation_is_reported(self):
        rows = self.run_checks(right=[(0.05, 0.02, 0.03), (0.05, 0.02, 0.03), None])
        row = rows["最新周填写"]
        self.assertEqual(row["level"], "warn")
        self.assertIn("缺 3 项", row["detail"])
        self.assertIn("W 列航空客运量", row["detail"])

    def test_missing_hotel_is_reported(self):
        rows = self.run_checks(
            right=[(0.05, 0.02, 0.03)] * 3,
            hotel=[(0.01, 0.02, 0.03), (0.01, 0.02, 0.03), (None, 0.02, None)],
        )
        row = rows["最新周填写"]
        self.assertEqual(row["level"], "warn")
        self.assertIn("S 列酒店入住率", row["detail"])
        self.assertIn("U 列酒店RevPAR", row["detail"])
        self.assertNotIn("T 列", row["detail"])


class TestReadSidesSeparately(unittest.TestCase):
    """`parse()` 会把回退合并掉，核对必须能拿到未合并的两侧原值。"""

    def test_weekly_sides_returns_raw_values(self):
        with TemporaryDirectory() as tmp:
            path = build(
                Path(tmp) / "wb.xlsx",
                right=[(0.05, 0.02, 0.03), None, (0.05, 0.02, 0.03)],
                left=[(0.07, 0.02, 0.03)] * 3,
            )
            right, left = excel.weekly_sides(path)
            self.assertEqual(right["aviationPax"], [0.05, None, 0.05])
            self.assertEqual(left["8/9-8/15"][0], 0.07)

            # 合并后第二周应回退到左侧，第一周仍用右侧——原有行为不变
            merged = excel.parse(path)["weekly"]
            self.assertEqual(merged["aviationPax"], [0.05, 0.07, 0.05])


if __name__ == "__main__":
    unittest.main()
