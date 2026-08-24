"""底稿结构校验的测试。

校验本身也必须被校验：如果它对列位挪动无感，那它给出的 ✓ 是虚假的安全感——
比没有检查更糟。

注意：`ws.cell(r, c, None)` 在 openpyxl 里是**空操作**（value 为 None 时不赋值），
清空单元格必须写 `ws.cell(r, c).value = None`。
"""

from __future__ import annotations

import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

import openpyxl

from modules.industry_data import layout

YEAR_ROW = 5  # 合成底稿里 2026 年块的行号，刻意不用真实的 165，验证契约是相对偏移


def build_workbook(path: Path, *, mutate=None) -> Path:
    """按契约造一份最小可用底稿。`mutate(ws)` 可在写好后破坏它。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = layout.SHEET

    ws.cell(YEAR_ROW, 2, "2026年")
    ws.cell(YEAR_ROW, 18, "2026年")

    group = YEAR_ROW + layout.OFF_GROUP
    for col, text in layout.LEFT_GROUPS.items():
        ws.cell(group, col, text)
    for col, text in layout.RIGHT_GROUPS.items():
        ws.cell(group, col, text)

    header = YEAR_ROW + layout.OFF_HEADER
    ws.cell(header, 2, "月度")
    for col, (text, _field) in layout.LEFT_COLUMNS.items():
        ws.cell(header, col, text)
    for col, (text, _field) in layout.RIGHT_COLUMNS.items():
        ws.cell(header, col, text)

    start = YEAR_ROW + layout.OFF_MONTH_START
    for index in range(12):
        # 第 7 个月刻意带后缀，复现底稿真实写法「7月 (preliminary)」
        label = "7月 (preliminary)" if index == 6 else f"{index + 1}月"
        ws.cell(start + index, 2, label)
        ws.cell(start + index, 3, 0.01)

    quarter_start = start + 12
    for index, quarter in enumerate(("Q1", "Q2", "Q3", "Q4")):
        ws.cell(quarter_start + index, 2, quarter)

    ws.cell(quarter_start + 4, 2, "QTD周度")
    ws.cell(quarter_start + 5, 2, "1/4-1/10")

    for index in range(3):
        ws.cell(header + 1 + index, 18, f"1/{4 + index * 7}-1/{10 + index * 7}")

    if mutate:
        mutate(ws)
    wb.save(path)
    return path


def levels(checks: list[dict]) -> dict[str, str]:
    return {row["name"]: row["level"] for row in checks}


class TestLayoutContract(unittest.TestCase):
    def _verify(self, mutate=None) -> list[dict]:
        with TemporaryDirectory() as tmp:
            path = build_workbook(Path(tmp) / "wb.xlsx", mutate=mutate)
            return layout.verify(path)

    def test_valid_workbook_passes(self):
        checks = self._verify()
        failed = [row for row in checks if row["level"] == "fail"]
        self.assertEqual(failed, [], f"合规底稿不该有 fail：{failed}")

    def test_month_suffix_is_accepted(self):
        """带后缀的月份行必须算合规——旧实现正是在这里提前终止。"""
        self.assertEqual(levels(self._verify())["月度行"], "ok")

    def test_renamed_sheet_fails(self):
        def mutate(ws):
            ws.title = "国内行业数据2026"

        self.assertEqual(levels(self._verify(mutate))["工作表"], "fail")

    def test_missing_year_block_fails(self):
        def mutate(ws):
            ws.cell(YEAR_ROW, 2, "2027年")

        self.assertEqual(levels(self._verify(mutate))["2026 年块"], "fail")

    def test_shifted_left_column_fails(self):
        """把「国铁」从 I 挪到 J —— 这是最危险的一类改动，必须抓到。"""

        def mutate(ws):
            header = YEAR_ROW + layout.OFF_HEADER
            ws.cell(header, 9).value = None
            ws.cell(header, 10, "国铁")

        self.assertEqual(levels(self._verify(mutate))["月度/季度列位"], "fail")

    def test_shifted_right_column_fails(self):
        """把航空「票价」从 X 挪走。"""

        def mutate(ws):
            header = YEAR_ROW + layout.OFF_HEADER
            ws.cell(header, 24).value = None
            ws.cell(header, 26, "票价")

        self.assertEqual(levels(self._verify(mutate))["周度列位"], "fail")

    def test_swapped_domestic_and_intl_groups_fails(self):
        """国内 / 国际航空下都有「民航局」，只靠指标名分不出——分组名必须一起核。"""

        def mutate(ws):
            group = YEAR_ROW + layout.OFF_GROUP
            ws.cell(group, 7, "国际航空客运量（含港澳台）")
            ws.cell(group, 11, "国内航空客运量")

        self.assertEqual(levels(self._verify(mutate))["月度/季度列位"], "fail")

    def test_missing_month_row_fails(self):
        def mutate(ws):
            ws.cell(YEAR_ROW + layout.OFF_MONTH_START + 8, 2).value = None

        self.assertEqual(levels(self._verify(mutate))["月度行"], "fail")

    def test_missing_quarter_fails(self):
        def mutate(ws):
            ws.cell(YEAR_ROW + layout.OFF_MONTH_START + 12 + 3, 2).value = None

        self.assertEqual(levels(self._verify(mutate))["季度行"], "fail")

    def test_missing_qtd_block_warns_not_fails(self):
        """QTD 只是航空回退源，缺了不该阻塞整条链。"""

        def mutate(ws):
            ws.cell(YEAR_ROW + layout.OFF_MONTH_START + 16, 2).value = None

        self.assertEqual(levels(self._verify(mutate))["QTD周度块"], "warn")

    def test_unreadable_file_fails_gracefully(self):
        with TemporaryDirectory() as tmp:
            broken = Path(tmp) / "broken.xlsx"
            broken.write_bytes(b"not an xlsx")
            self.assertEqual(levels(layout.verify(broken))["底稿可读"], "fail")

    def test_full_width_parens_are_normalised(self):
        """底稿里同一含义写成半角与全角括号两种，都要认。"""

        def mutate(ws):
            ws.cell(YEAR_ROW + layout.OFF_GROUP, 3, "国内酒店（STR）")

        self.assertEqual(levels(self._verify(mutate))["月度/季度列位"], "ok")

    def test_newline_in_header_is_normalised(self):
        """真实底稿的 W 列表头含换行。"""

        def mutate(ws):
            ws.cell(YEAR_ROW + layout.OFF_HEADER, 23, "客运量\n（含出入境）")

        self.assertEqual(levels(self._verify(mutate))["周度列位"], "ok")


class TestColLetter(unittest.TestCase):
    def test_known_columns(self):
        for index, letter in ((1, "A"), (18, "R"), (23, "W"), (26, "Z"), (27, "AA")):
            self.assertEqual(layout.col_letter(index), letter)


if __name__ == "__main__":
    unittest.main()


class TestMonthNumber(unittest.TestCase):
    """月份行标签的匹配规则只有一份，两个域共用（layout.month_number）。

    两个域都因为精确相等匹配踩过坑：
    - industry-data 的旧 parser 在「7月 (preliminary)」处提前终止，月度只读到 6 个月
    - aviation-monthly 的管道报 `Cannot locate 2026年7月 monthly row`，直接写不进去
    """

    def test_plain_labels(self):
        for month in range(1, 13):
            self.assertEqual(layout.month_number(f"{month}月"), month)

    def test_label_with_suffix(self):
        self.assertEqual(layout.month_number("7月 (preliminary)"), 7)
        self.assertEqual(layout.month_number("7月（初值）"), 7)
        self.assertEqual(layout.month_number(" 12月 "), 12)

    def test_december_not_confused_with_january(self):
        """「12月」不能被当成 1 月——按数字比对而不是前缀比对。"""
        self.assertEqual(layout.month_number("12月"), 12)
        self.assertNotEqual(layout.month_number("12月"), 1)

    def test_non_month_labels(self):
        for label in ("季度", "Q1", "QTD周度", "2026年", "", None, "月度", "13月", "0月"):
            self.assertIsNone(layout.month_number(label), label)


if __name__ == "__main__":
    unittest.main()
