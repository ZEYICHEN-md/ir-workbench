"""把中金表算出的酒店数据与底稿现值对照，出待写入清单（只读）。

写入是另一件事（须走 Excel COM + 备份 + 回读，见 aviation_monthly 的做法）。
本模块只负责「算出来、摆出来」，让人先核对。
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl

from workbench.result import Result

from . import layout, str_source
from .paths import DOMAIN

#: 底稿月度区酒店三列
MONTH_COLS = {"hotelOccupancy": 3, "hotelADR": 4, "hotelRevPAR": 5}
#: 底稿右侧周度区酒店三列
WEEK_COLS = {"hotelOccupancy": 19, "hotelADR": 20, "hotelRevPAR": 21}
COL_WEEK_AXIS = 18

#: 视为「值不同」的**绝对**容差，单位是小数同比（5e-5 = 0.005 个百分点）。
#:
#: 为什么用绝对容差而不是相对容差：底稿里的手填值常是整数（券商发布 `-1%`，
#: 底稿存 `-0.01`），而重算值是 `-0.010046`。按相对容差（1e-6）会把这种差异报成冲突，
#: 而它在业务上根本不算不同——看板展示只到一位小数（GLOSSARY「看板展示精度」）。
#: 阈值取 0.005pp，即「四舍五入到 0.01 个百分点后不同才算冲突」。
TOLERANCE = 5e-5


@dataclass
class Cell:
    """一处待写入。"""

    address: str
    where: str          # 「2026年7月」/「周 8/9-8/15」
    metric: str         # 人读指标名
    old: float | None
    new: float

    @property
    def kind(self) -> str:
        if self.old is None:
            return "新增"
        return "冲突" if _differs(self.old, self.new) else "一致"

    def describe(self) -> str:
        old = "空" if self.old is None else f"{self.old * 100:+.2f}%"
        return f"{self.address} {self.where} {self.metric}：{old} → {self.new * 100:+.2f}%"


def _differs(old: float, new: float) -> bool:
    return abs(old - new) > TOLERANCE


def _find_year_row(sheet, year: int) -> int:
    for row in range(1, sheet.max_row + 1):
        value = sheet.cell(row, 2).value
        if value and str(value).strip().startswith(str(year)):
            return row
    raise str_source.StrSourceError(f"底稿里找不到 {year} 年块（B 列「{year}年」）")


def _month_rows(sheet, year_row: int) -> dict[int, int]:
    """月份号 → 行号。标签可带后缀（`7月 (preliminary)`），规则与读表共用一份。"""
    start = year_row + layout.OFF_MONTH_START
    out: dict[int, int] = {}
    for row in range(start, start + 12):
        number = layout.month_number(sheet.cell(row, 2).value)
        if number:
            out[number] = row
    return out


def _week_rows(sheet, year_row: int) -> dict[str, int]:
    """周标签 → 行号。用底稿右侧 R 列周轴。"""
    header = None
    for row in range(year_row, year_row + 10):
        value = sheet.cell(row, COL_WEEK_AXIS).value
        if value and str(value).strip() == "周":
            header = row
            break
    if header is None:
        raise str_source.StrSourceError("底稿里找不到右侧周度表头（R 列「周」）")

    from . import excel as excel_mod

    out: dict[str, int] = {}
    for row in range(header + 1, header + 80):
        label = sheet.cell(row, COL_WEEK_AXIS).value
        if label is None or str(label).strip() == "":
            if out:
                break
            continue
        text = str(label).strip()
        if text.startswith(("Note", "注", "**")):
            break
        out[excel_mod.norm_week(text)] = row
    return out


def build(workbook: Path, source: Path, year: int) -> tuple[list[Cell], list[str]]:
    """算出待写入清单。返回 (cells, notes)。"""
    weeks = str_source.load(source)
    notes = [
        f"中金表：{source.name} · tab「{str_source.SHEET}」",
        f"覆盖 {weeks[0].label} 至 {weeks[-1].label}，共 {len(weeks)} 周",
    ]

    book = openpyxl.load_workbook(workbook, data_only=True)
    try:
        sheet = book[layout.SHEET]
        year_row = _find_year_row(sheet, year)
        month_rows = _month_rows(sheet, year_row)
        week_rows = _week_rows(sheet, year_row)
    finally:
        book.close()

    cells: list[Cell] = []

    # --- 月度：只处理被周数据完整覆盖的月份 ---
    for agg in str_source.complete_months(weeks, year):
        row = month_rows.get(agg.month)
        if row is None:
            notes.append(f"⚠ 底稿里找不到 {year}年{agg.month}月 行，跳过")
            continue
        yoy = agg.yoy()
        for field, name in str_source.METRICS:
            col = MONTH_COLS[field]
            book_value = _read(workbook, row, col)
            cells.append(
                Cell(
                    address=f"{layout.col_letter(col)}{row}",
                    where=f"{year}年{agg.month}月",
                    metric=name,
                    old=book_value,
                    new=yoy[field],
                )
            )

    # --- 周度：直接取表里的 K/L/M ---
    for label, values in str_source.weekly_yoy(weeks, year):
        row = week_rows.get(label)
        if row is None:
            notes.append(f"⚠ 底稿周轴里没有「{label}」——新周次需先在底稿加行，本工具不加行")
            continue
        for field, name in str_source.METRICS:
            new = values[field]
            if new is None:
                continue
            col = WEEK_COLS[field]
            cells.append(
                Cell(
                    address=f"{layout.col_letter(col)}{row}",
                    where=f"周 {label}",
                    metric=name,
                    old=_read(workbook, row, col),
                    new=new,
                )
            )

    return cells, notes


_cache: dict[Path, openpyxl.Workbook] = {}


def _read(workbook: Path, row: int, col: int) -> float | None:
    """读底稿一格。工作簿缓存一次，避免每格重开。"""
    book = _cache.get(workbook)
    if book is None:
        book = openpyxl.load_workbook(workbook, data_only=True)
        _cache[workbook] = book
    value = book[layout.SHEET].cell(row, col).value
    return float(value) if isinstance(value, (int, float)) else None


def run(workbook: Path, source: Path, year: int = 2026) -> Result:
    """只读：算出并摆出待写入清单。不动任何文件。"""
    _cache.clear()
    try:
        cells, notes = build(workbook, source, year)
    except str_source.StrSourceError as error:
        return Result(
            status="blocked",
            summary=f"读不出中金表或底稿结构不符：{error}",
            domain=DOMAIN,
            next_steps=["确认中金表 tab 名与列位没改版；改版了先更新 str_source 的契约再跑。"],
        )

    additions = [c for c in cells if c.kind == "新增"]
    conflicts = [c for c in cells if c.kind == "冲突"]
    same = [c for c in cells if c.kind == "一致"]

    checks = [
        {"name": "中金表", "level": "ok", "detail": notes[0].split("：", 1)[1]},
        {"name": "覆盖范围", "level": "ok", "detail": notes[1]},
        {
            "name": "对照结果",
            "level": "ok",
            "detail": f"新增 {len(additions)} · 冲突 {len(conflicts)} · 已一致 {len(same)}",
        },
    ]
    checks += [
        {"name": "待填空格", "level": "ok", "detail": c.describe()} for c in additions[:20]
    ]
    checks += [
        {"name": "值不一致", "level": "warn", "detail": c.describe()} for c in conflicts[:20]
    ]

    warnings = [note for note in notes[2:]]
    if conflicts:
        warnings.append(
            f"有 {len(conflicts)} 处底稿现值与中金表算出的不同，**默认不覆盖**。"
            "2026 年 1–6 月的月度值口径此前不统一（有的来自 STR 官方月报、有的来自券商、"
            "还有过券商预测值），所以历史月份出现冲突是预期的。要统一历史口径是单独一件事，"
            "会改动看板上的历史曲线，须另行决定。"
        )

    return Result(
        status="partial" if (additions or conflicts) else "success",
        summary=(
            f"已算出中金表对底稿的差异（可填空格 {len(additions)} · 值不一致 {len(conflicts)}），**未写入**。"
            if additions or conflicts
            else "底稿与中金表一致，无需写入。"
        ),
        domain=DOMAIN,
        checks=checks,
        warnings=warnings,
        next_steps=[
            "逐条核对上面的清单。",
            "写入时默认只填空格、不覆盖已有值（同飞书投影的规则）。",
            "写底稿须走 Excel COM + 备份 + 回读，且要你明确说「写入」。",
        ],
        data={
            "additions": len(additions),
            "conflicts": len(conflicts),
            "same": len(same),
        },
    )
