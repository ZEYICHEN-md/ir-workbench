"""指标底稿的读表契约与结构校验。

为什么要有这个：`excel.py` 按固定列号取值。**列一挪，取到的数就是别的指标，而且不会报错。**
这是最危险的一类失败——技术上合法、业务上全错（旧管道那个「7月 (preliminary)」缺陷就是
同一家族）。所以把每个指标列的表头写成显式契约，在跑之前核对。

表头文本取自 `国内行业数据_0817.xlsx`，已逐格核实。
"""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl

#: 相对 2026 块表头行的偏移
OFF_GROUP = 1  # 分组名行（如「国内酒店(STR)」）
OFF_HEADER = 2  # 指标名行（如「入住率」）；同一行 B 列 =「月度」，R 列 =「周」
OFF_MONTH_START = 3  # 1 月行

#: 左侧月度 / 季度区：列号 → (指标名行应含的文本, 该列对应的快照字段)
LEFT_COLUMNS: dict[int, tuple[str, str]] = {
    3: ("入住率", "hotelOccupancy"),
    4: ("ADR", "hotelADR"),
    5: ("RevPAR", "hotelRevPAR"),
    7: ("民航局", "domAviationCAAC"),
    8: ("三大航", "domAviationBig3"),
    9: ("国铁", "railway"),
    11: ("民航局", "intlAviationCAAC"),
    12: ("三大航", "intlAviationBig3"),
    13: ("航班管家", "intlCapacity"),
}

#: 左侧分组名行，用于区分「国内航空」与「国际航空」下重名的「民航局 / 三大航」
LEFT_GROUPS: dict[int, str] = {
    3: "国内酒店",
    7: "国内航空客运量",
    9: "铁路客运量",
    11: "国际航空客运量",
    13: "国际航班运力",
}

#: 右侧周度区：列号 → (指标名行应含的文本, 快照字段)
RIGHT_COLUMNS: dict[int, tuple[str, str]] = {
    18: ("周", "weeks"),
    19: ("入住率", "hotelOccupancy"),
    20: ("ADR", "hotelADR"),
    21: ("RevPAR", "hotelRevPAR"),
    23: ("客运量", "aviationPax"),
    24: ("票价", "aviationTicket"),
    25: ("客运航班量", "aviationFlight"),
}

RIGHT_GROUPS: dict[int, str] = {19: "国内酒店", 23: "航空"}

SHEET = "国内行业数据"
YEAR = 2026


def col_letter(index: int) -> str:
    letters = ""
    while index:
        index, rem = divmod(index - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


_MONTH_LABEL = re.compile(r"^(\d{1,2})月")


def month_number(label) -> int | None:
    """从底稿月份行标签取月份号；不是月份行则返回 None。

    **标签可以带后缀**——底稿里 7 月写作 `7月 (preliminary)`。用精确相等匹配会在这里
    找不到行，而两个域都踩过这个坑：`industry-data` 的旧 parser 在此提前终止导致月度
    只读到 6 个月；`aviation-monthly` 的管道则报 `Cannot locate 2026年7月 monthly row`
    直接写不进去。所以这条规则只留一份，两边都用它。
    """
    match = _MONTH_LABEL.match(_norm(label))
    if not match:
        return None
    value = int(match.group(1))
    return value if 1 <= value <= 12 else None


def _norm(value) -> str:
    """归一化表头：去空白与换行，全角括号转半角。

    底稿里同一含义会写成「国内酒店(STR)」与「国内酒店（STR）」两种，
    表头也含换行（`客运量\\n（含出入境）`）。
    """
    text = "" if value is None else str(value)
    text = re.sub(r"\s+", "", text)
    return text.replace("（", "(").replace("）", ")")


def _ok(name: str, detail: str) -> dict:
    return {"name": name, "level": "ok", "detail": detail}


def _fail(name: str, detail: str, advice: str) -> dict:
    return {"name": name, "level": "fail", "detail": detail, "advice": advice}


def verify(workbook: Path) -> list[dict]:
    """核对底稿结构。返回 checks 列表；有 `level == "fail"` 即视为不合契约。"""
    try:
        wb = openpyxl.load_workbook(workbook, data_only=True, read_only=True)
    except Exception as error:  # noqa: BLE001 —— 打不开就是打不开，原因原样报出
        return [
            _fail("底稿可读", f"{type(error).__name__}: {error}", "确认文件没被 Excel 锁住或损坏。")
        ]

    try:
        if SHEET not in wb.sheetnames:
            return [
                _fail(
                    "工作表",
                    f"缺少「{SHEET}」，现有：{'、'.join(wb.sheetnames)}",
                    f"底稿必须有名为「{SHEET}」的工作表；被改名了就改回来。",
                )
            ]
        ws = wb[SHEET]
        grid = _read_region(ws)
    finally:
        wb.close()

    checks: list[dict] = [_ok("工作表", SHEET)]

    year_row = None
    for row in sorted(grid):
        if _norm(grid[row].get(2)).startswith(str(YEAR)):
            year_row = row
            break
    if year_row is None:
        return checks + [
            _fail(
                f"{YEAR} 年块",
                f"B 列找不到「{YEAR}年」",
                f"底稿里应有 B 列写「{YEAR}年」的区块；换年度了要先更新读表契约。",
            )
        ]
    checks.append(_ok(f"{YEAR} 年块", f"第 {year_row} 行"))

    group_row = grid.get(year_row + OFF_GROUP, {})
    header_row = grid.get(year_row + OFF_HEADER, {})

    checks.extend(_verify_columns("月度/季度", LEFT_COLUMNS, LEFT_GROUPS, header_row, group_row))
    checks.extend(_verify_columns("周度", RIGHT_COLUMNS, RIGHT_GROUPS, header_row, group_row))
    checks.append(_verify_month_rows(grid, year_row))
    checks.append(_verify_quarter_rows(grid, year_row))
    checks.append(_verify_qtd_block(grid, year_row))

    return checks


def _read_region(ws) -> dict[int, dict[int, object]]:
    """一次性读出可能相关的区域，避免 read_only 模式下随机访问变慢。

    行号用 enumerate 自己数：read_only 模式下空行会给出 `EmptyCell`，它没有 `.row`。
    """
    grid: dict[int, dict[int, object]] = {}
    for row_index, row in enumerate(ws.iter_rows(min_col=1, max_col=26, values_only=True), start=1):
        cells = {col: value for col, value in enumerate(row, start=1) if value not in (None, "")}
        if cells:
            grid[row_index] = cells
    return grid


def _verify_columns(
    where: str,
    columns: dict[int, tuple[str, str]],
    groups: dict[int, str],
    header_row: dict,
    group_row: dict,
) -> list[dict]:
    bad: list[str] = []
    for col, (expected, field) in columns.items():
        actual = _norm(header_row.get(col))
        if expected not in actual:
            bad.append(f"{col_letter(col)} 列应含「{expected}」（{field}），实为「{actual or '空'}」")
    for col, expected in groups.items():
        actual = _norm(group_row.get(col))
        if expected not in actual:
            bad.append(f"{col_letter(col)} 列分组应含「{expected}」，实为「{actual or '空'}」")

    if bad:
        return [
            _fail(
                f"{where}列位",
                "；".join(bad),
                "列被挪动或改名了。**先更新 modules/industry_data/layout.py 的契约与 excel.py 的列号，"
                "再重跑**——直接跑会把别的指标当成这个指标读进来，而且不会报错。",
            )
        ]
    return [_ok(f"{where}列位", f"{len(columns)} 列表头与分组均符合契约")]


def _verify_month_rows(grid: dict, year_row: int) -> dict:
    start = year_row + OFF_MONTH_START
    found: list[str] = []
    for row in range(start, start + 12):
        label = _norm(grid.get(row, {}).get(2))
        match = re.match(r"^(\d{1,2})月", label)
        if not match:
            break
        found.append(f"{int(match.group(1))}月")
    if len(found) != 12:
        return _fail(
            "月度行",
            f"只识别到 {len(found)} 个月（{'、'.join(found) or '无'}）",
            "月份行应为 12 行，标签可带后缀（如「7月 (preliminary)」）但必须以「N月」开头。"
            "少于 12 个通常是有一行的标签写法变了。",
        )
    return _ok("月度行", "1月–12月 齐全")


def _verify_quarter_rows(grid: dict, year_row: int) -> dict:
    start = year_row + OFF_MONTH_START
    found = [
        _norm(grid.get(row, {}).get(2)).upper()
        for row in range(start, start + 20)
        if re.fullmatch(r"Q[1-4]", _norm(grid.get(row, {}).get(2)).upper())
    ]
    if len(found) != 4:
        return _fail(
            "季度行",
            f"只识别到 {len(found)} 行（{'、'.join(found) or '无'}）",
            "季度区应有 Q1–Q4 四行。",
        )
    return _ok("季度行", "Q1–Q4 齐全")


def _verify_qtd_block(grid: dict, year_row: int) -> dict:
    start = year_row + OFF_MONTH_START
    for row in range(start, start + 40):
        if "QTD" in _norm(grid.get(row, {}).get(2)).upper():
            return _ok("QTD周度块", f"第 {row} 行（航空回退源）")
    return {
        "name": "QTD周度块",
        "level": "warn",
        "detail": "未找到",
        "advice": "左侧「QTD周度」块是右侧航空列为空时的回退源；缺了不阻塞，但右侧缺值会直接变空。",
    }
