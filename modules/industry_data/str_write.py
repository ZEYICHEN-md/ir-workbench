"""把中金表算出的酒店数据写进底稿。

## 权威层级（本模块存在的前提）

底稿是唯一指标真源（ADR 0001），但**底稿内部还有一层权威**：

> **人手填的数字最高权威。**周度聚合只是获取数据的一种方式，不是新的权威源。

所以写入规则是**仅填空、绝不覆盖**：空格才写，已有值一律保留，即使与算出来的不同。
历史口径不需要统一——1–6 月的月度值来自 STR 官方月报/券商/一度还有券商预测值，
那都是当时 IR 经理的判断，自动化没有资格改它。

## 四重保护（照搬 aviation_monthly 的做法，那套已在真实写入中验证过）

1. **写入前归档**：当前底稿整份复制进 `data/workbooks/archived/`，文件名带时间戳。
   人能直接双击打开核对，不要求会用 git。
2. **不能用 openpyxl 保存底稿**：它含历史外部链接，openpyxl 保存会重写链接与 XML，
   导致 Excel 打不开。一律走 **Excel COM**。
3. **写完回读**：重新打开核对每一格确实是目标值，并核对不该动的格没被动。
4. **占用检查**：底稿被 Excel 打开时（有 `~$` 锁文件）直接拒绝写入——否则会与人的
   编辑冲突，或者被人保存时覆盖掉。
"""

from __future__ import annotations

import datetime as dt
import shutil
from dataclasses import dataclass
from pathlib import Path

import openpyxl

from workbench.result import Result

from . import layout, str_plan, str_source
from .paths import DOMAIN, DomainPaths

#: 写入的溯源批注模板
COMMENT = "自动写入 · 来源 {source} · tab「Mainland China (STR)」· {basis} · {stamp}"

#: 周度表格的横向范围。**只到 Y**——Z 及以后是火车票区，边框规律与这一块不同
#: （末行 Y 是 medium 底框而 Z 是 thin），不属于本管道职责，一律不碰。
WEEK_BLOCK_FIRST_COL = 18  # R
WEEK_BLOCK_LAST_COL = 25   # Y

#: Excel COM 常量（避免依赖 win32com.client.constants，那个要求 makepy 生成类型库）
XL_PASTE_FORMATS = -4122
XL_EDGE_BOTTOM = 9
XL_CONTINUOUS = 1
XL_THIN = 2
XL_MEDIUM = -4138
XL_UP = -4162
XL_SHIFT_DOWN = -4121


@dataclass
class WriteOutcome:
    address: str
    where: str
    metric: str
    value: float
    verified: float | None


def _lock_file(workbook: Path) -> Path | None:
    from workbench.archive import lock_file

    return lock_file(workbook)


def archive(paths: DomainPaths, workbook: Path, tag: str) -> Path:
    """写入前把整份底稿复制进归档目录。实现在共享层（两个域都写底稿）。"""
    from workbench.archive import archive_workbook

    return archive_workbook(paths.base, workbook, tag)


def _block(sheet, row: int):
    """该行的周度区范围 R:Y。"""
    return sheet.Range(
        sheet.Cells(row, WEEK_BLOCK_FIRST_COL), sheet.Cells(row, WEEK_BLOCK_LAST_COL)
    )


def _set_bottom(sheet, row: int, weight: int) -> None:
    """设置该行周度区的下边框粗细。"""
    border = _block(sheet, row).Borders(XL_EDGE_BOTTOM)
    border.LineStyle = XL_CONTINUOUS
    border.Weight = weight


def _last_week_row(sheet, header_row: int) -> int:
    """周轴最后一行。以 R 列连续非空判断，与读表逻辑一致。"""
    row = header_row
    while True:
        value = sheet.Cells(row + 1, WEEK_BLOCK_FIRST_COL).Value
        if value is None or str(value).strip() == "":
            return row
        row += 1


def _append_week_row(sheet, header_row: int, label: str) -> int:
    """在周轴末尾加一行，格式对齐上一行。返回新行号。

    格式处理三步（这是使用者明确要求的表格外观）：
    1. 整块 R:Y 从上一行复制格式 —— 边框、填充、字体、数字格式一次带全，
       不逐属性猜（逐属性复制迟早漏掉某个，比如 0% 数字格式或浅灰填充）；
    2. 新行下边框设为 **medium**（它成了表格最后一行）；
    3. 上一行下边框改回 **thin**（它不再是最后一行）。

    是否插入行看下一行空不空：空就直接写（零移动，最安全）；被占用才 Insert
    （底稿末行下方约三行之后是 `Note：春节数据为民航局日均（含出境）` 那行注释，
    连续加几周就会撞上它，此时下移注释是正确行为）。
    """
    last = _last_week_row(sheet, header_row)
    target = last + 1

    occupied = any(
        sheet.Cells(target, col).Value not in (None, "")
        for col in range(WEEK_BLOCK_FIRST_COL, WEEK_BLOCK_LAST_COL + 1)
    )
    if occupied:
        # 只下移周度区，不动整行——左侧月度/季度区共享行号，整行插入会牵动它
        _block(sheet, target).Insert(Shift=XL_SHIFT_DOWN)

    _block(sheet, last).Copy()
    _block(sheet, target).PasteSpecial(Paste=XL_PASTE_FORMATS)
    sheet.Application.CutCopyMode = False

    sheet.Cells(target, WEEK_BLOCK_FIRST_COL).Value = label
    _set_bottom(sheet, target, XL_MEDIUM)
    _set_bottom(sheet, last, XL_THIN)
    return target


def _write_via_com(
    workbook: Path,
    cells: list[tuple[str, float, str]],
    new_weeks: list[tuple[str, dict[str, float | None], str]] | None = None,
    year: int = 2026,
) -> dict[str, dict[str, float]]:
    """用 Excel COM 加行、写值、全量重算。

    返回新建周次实际落在哪些格（`{周次: {地址: 值}}`），供回读核对。
    """
    try:
        import pythoncom
        import win32com.client
    except ImportError as error:  # pragma: no cover —— doctor 已预检
        raise RuntimeError(f"需要 pywin32 才能写底稿：{error}") from error

    from . import str_plan

    written_new: dict[str, dict[str, float]] = {}
    app = None
    book = None
    pythoncom.CoInitialize()
    try:
        app = win32com.client.DispatchEx("Excel.Application")
        app.Visible = False
        app.DisplayAlerts = False
        app.AskToUpdateLinks = False
        book = app.Workbooks.Open(
            str(workbook.resolve()),
            UpdateLinks=0,
            ReadOnly=False,
            IgnoreReadOnlyRecommended=True,
            AddToMru=False,
        )
        sheet = book.Worksheets(layout.SHEET)

        # --- 先加行。加在末尾，所以已有行的地址不变，下面填值仍可用原地址 ---
        if new_weeks:
            header_row = _week_header_row(sheet, year)
            for label, values, comment in new_weeks:
                row = _append_week_row(sheet, header_row, label)
                placed: dict[str, float] = {}
                for field, _name in str_source.METRICS:
                    value = values.get(field)
                    if value is None:
                        continue
                    col = str_plan.WEEK_COLS[field]
                    target = sheet.Cells(row, col)
                    target.Value = value
                    if target.Comment is not None:
                        target.Comment.Delete()
                    target.AddComment(comment)
                    placed[f"{layout.col_letter(col)}{row}"] = value
                written_new[label] = placed

        # --- 再填已有行的空格 ---
        for address, value, comment in cells:
            target = sheet.Range(address)
            target.Value = value
            # 批注是溯源的一部分：过一个月没人记得这格是自动填的还是手填的
            if target.Comment is not None:
                target.Comment.Delete()
            target.AddComment(comment)

        app.CalculateFullRebuild()
        book.Save()
    except Exception as error:
        raise RuntimeError(f"Excel COM 写入失败：{error}") from error
    finally:
        if book is not None:
            try:
                book.Close(SaveChanges=False)
            except Exception:
                pass
        if app is not None:
            try:
                app.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()
    return written_new


def _week_header_row(sheet, year: int) -> int:
    """指定年度块的右侧周轴表头行（R 列 == 「周」）。

    **必须先定位年度块**：底稿里每个年度块都有自己的周轴，实测有 5 个
    （2023/2024/2025×2/2026 分别在 r25/r47/r75/r137/r167）。
    早先的实现从第 1 行找第一个「周」，会命中 2023 年块 —— 加行就加到了错的年份里。
    这个缺陷是端到端实测抓到的，单元测试的合成底稿只有一个年度块，测不出来。
    """
    year_row = None
    for row in range(1, 1000):
        value = sheet.Cells(row, 2).Value
        if value is not None and str(value).strip().startswith(f"{year}年"):
            year_row = row
            break
    if year_row is None:
        raise RuntimeError(f"底稿里找不到 {year} 年块（B 列「{year}年」）")

    for row in range(year_row, year_row + 10):
        value = sheet.Cells(row, WEEK_BLOCK_FIRST_COL).Value
        if value is not None and str(value).strip() == "周":
            return row
    raise RuntimeError(f"{year} 年块里找不到右侧周度表头（R 列「周」）")


def _verify(workbook: Path, expected: list[tuple[str, float]], untouched: list[tuple[str, float | None]]
            ) -> tuple[list[tuple[str, float, float | None]], list[str]]:
    """回读核对。返回 (实测值, 问题清单)。"""
    book = openpyxl.load_workbook(workbook, data_only=True)
    problems: list[str] = []
    actual: list[tuple[str, float, float | None]] = []
    try:
        sheet = book[layout.SHEET]
        for address, value in expected:
            got = sheet[address].value
            got = float(got) if isinstance(got, (int, float)) else None
            actual.append((address, value, got))
            if got is None or abs(got - value) > 1e-9:
                problems.append(f"{address} 应为 {value!r}，回读得到 {got!r}")
        for address, before in untouched:
            got = sheet[address].value
            got = float(got) if isinstance(got, (int, float)) else None
            if (before is None) != (got is None) or (
                before is not None and got is not None and abs(got - before) > 1e-9
            ):
                problems.append(f"{address} 本不该被改动：{before!r} → {got!r}")
    finally:
        book.close()
    return actual, problems


def _verify_borders(workbook: Path, new_labels: list[str], year: int = 2026) -> list[str]:
    """回读核对边框：只有最后一行是粗底框，其余都是细的。

    这条要单独验，因为格式错了不影响数值——merge、看板、上线全都会正常，
    只有人打开 Excel 才会看见表格烂了。没有检查就没人会发现。
    """
    if not new_labels:
        return []

    from . import excel as excel_mod

    problems: list[str] = []
    book = openpyxl.load_workbook(workbook, data_only=False)
    try:
        sheet = book[layout.SHEET]
        # 同样必须先定位年度块——底稿有 5 个周轴
        year_row = next(
            (r for r in range(1, sheet.max_row + 1)
             if sheet.cell(r, 2).value
             and str(sheet.cell(r, 2).value).strip().startswith(f"{year}年")),
            None,
        )
        if year_row is None:
            return [f"回读时找不到 {year} 年块"]
        header = next(
            (r for r in range(year_row, year_row + 10)
             if str(sheet.cell(r, WEEK_BLOCK_FIRST_COL).value or "").strip() == "周"),
            None,
        )
        if header is None:
            return ["回读时找不到周轴表头"]

        rows: list[int] = []
        for row in range(header + 1, header + 120):
            value = sheet.cell(row, WEEK_BLOCK_FIRST_COL).value
            if value is None or str(value).strip() == "":
                break
            rows.append(row)

        labels = {excel_mod.norm_week(str(sheet.cell(r, WEEK_BLOCK_FIRST_COL).value)) for r in rows}
        for label in new_labels:
            if label not in labels:
                problems.append(f"新建的周次「{label}」回读时不在周轴里")

        last = rows[-1]
        for col in range(WEEK_BLOCK_FIRST_COL, WEEK_BLOCK_LAST_COL + 1):
            style = sheet.cell(last, col).border.bottom.style
            if style != "medium":
                problems.append(
                    f"末行 {layout.col_letter(col)}{last} 的下边框应为 medium（粗），实为 {style!r}"
                )
        if len(rows) >= 2:
            prev = rows[-2]
            for col in range(WEEK_BLOCK_FIRST_COL, WEEK_BLOCK_LAST_COL + 1):
                style = sheet.cell(prev, col).border.bottom.style
                if style == "medium":
                    problems.append(
                        f"倒数第二行 {layout.col_letter(col)}{prev} 的下边框仍是 medium，"
                        "应已改回 thin（它不再是最后一行）"
                    )
    finally:
        book.close()
    return problems


def run(paths: DomainPaths, workbook: Path, source: Path, year: int, *, yes: bool = False) -> Result:
    """写入底稿。默认 dry-run；`yes=True` 才真写。**只填空格，不覆盖已有值。**"""
    try:
        cells, notes, new_weeks = str_plan.build(workbook, source, year)
    except str_source.StrSourceError as error:
        return Result(
            status="blocked",
            summary=f"读不出中金表或底稿结构不符：{error}",
            domain=DOMAIN,
            next_steps=["确认中金表 tab 名与列位没改版；改版了先更新 str_source 的契约。"],
        )

    additions = [c for c in cells if c.kind == "新增"]
    conflicts = [c for c in cells if c.kind == "冲突"]

    if not additions and not new_weeks:
        return Result(
            status="success",
            summary="没有要加的周次，也没有空格要填，底稿未改动。",
            domain=DOMAIN,
            checks=[
                {"name": "可填空格", "level": "ok", "detail": "0"},
                {
                    "name": "值不一致",
                    "level": "warn" if conflicts else "ok",
                    "detail": f"{len(conflicts)} 处（人手填的值最高权威，不覆盖）",
                },
            ],
            warnings=notes[2:],
        )

    plan_checks = [
        {"name": "中金表", "level": "ok", "detail": source.name},
        {
            "name": "跳过（已有值）",
            "level": "ok",
            "detail": f"{len(conflicts)} 处 —— 人手填的最高权威，不覆盖",
        },
    ]
    for week in new_weeks:
        shown = "、".join(
            f"{name} {week.values[field] * 100:+.2f}%"
            for field, name in str_source.METRICS
            if week.values.get(field) is not None
        )
        plan_checks.append({
            "name": "将新建周次",
            "level": "ok",
            "detail": f"{week.label} —— {shown}（格式对齐上一行，粗底框下移）",
        })
    if additions:
        plan_checks.append({"name": "将填空格", "level": "ok", "detail": f"{len(additions)} 处"})
    plan_checks += [{"name": "待写入", "level": "ok", "detail": c.describe()} for c in additions[:20]]

    if not yes:
        bits = []
        if new_weeks:
            bits.append(f"新建 {len(new_weeks)} 个周次")
        if additions:
            bits.append(f"填 {len(additions)} 处空格")
        return Result(
            status="partial",
            summary="将" + "、".join(bits) + "，**未写入**。",
            domain=DOMAIN,
            checks=plan_checks,
            warnings=notes[2:],
            next_steps=[
                "核对上面每一格的位置与数值。",
                "确认后回一句「写入」，Agent 才会动底稿。",
                "写入前会把整份底稿归档到 data/workbooks/archived/，可随时取回。",
            ],
            data={
                "additions": len(additions),
                "conflicts": len(conflicts),
                "new_weeks": [w.label for w in new_weeks],
            },
        )

    # --- 以下是真写 ---
    lock = _lock_file(workbook)
    if lock:
        return Result(
            status="blocked",
            summary="底稿正在 Excel 里打开，拒绝写入。",
            domain=DOMAIN,
            missing=[f"锁文件：{lock.name}"],
            next_steps=["先在 Excel 里保存并关闭底稿，再让我写入——否则会与你的编辑冲突。"],
        )

    backup = archive(paths, workbook, "str")
    stamp = dt.datetime.now().strftime("%Y-%m-%d %H:%M")
    payload: list[tuple[str, float, str]] = []
    for cell in additions:
        basis = "周度取表内 K/L/M" if cell.where.startswith("周") else "月度按天加权聚合"
        payload.append((
            cell.address,
            cell.new,
            COMMENT.format(source=source.name, basis=basis, stamp=stamp),
        ))
    week_payload = [
        (
            week.label,
            week.values,
            COMMENT.format(source=source.name, basis="周度取表内 K/L/M", stamp=stamp),
        )
        for week in new_weeks
    ]

    try:
        placed = _write_via_com(workbook, payload, week_payload, year=year)
    except RuntimeError as error:
        shutil.copy2(backup, workbook)  # 原子性兜底：写坏就整份还原
        return Result(
            status="failed",
            summary=f"写入失败，已从归档还原底稿：{error}",
            domain=DOMAIN,
            checks=[{"name": "已还原", "level": "ok", "detail": backup.name}],
            next_steps=["确认 Excel 可用（ir doctor 会查 pywin32 与 COM），再重试。"],
        )

    expected = [(c.address, c.new) for c in additions]
    for cells_of_week in placed.values():
        expected += list(cells_of_week.items())

    actual, problems = _verify(
        workbook,
        expected,
        [(c.address, c.old) for c in conflicts],
    )
    problems += _verify_borders(workbook, list(placed), year)

    if problems:
        shutil.copy2(backup, workbook)
        return Result(
            status="failed",
            summary=f"回读核对不通过（{len(problems)} 处），已从归档还原底稿。",
            domain=DOMAIN,
            missing=problems[:10],
            checks=[{"name": "已还原", "level": "ok", "detail": backup.name}],
            next_steps=["把上面的差异贴给维护者——这说明写入路径有问题，不要绕过。"],
        )

    checks = [
        {"name": "写入前归档", "level": "ok", "detail": f"archived/{backup.name}"},
        {"name": "未覆盖", "level": "ok", "detail": f"{len(conflicts)} 处已有值保持原样"},
    ]
    for label, cells_of_week in placed.items():
        row = next(iter(cells_of_week)) if cells_of_week else "?"
        checks.append({
            "name": "已新建周次",
            "level": "ok",
            "detail": f"{label}（{row} 起）· 格式已对齐 · 粗底框已下移",
        })
    if additions:
        checks.append(
            {"name": "已填空格", "level": "ok", "detail": f"{len(additions)} 处，回读全部一致"}
        )
    checks += [
        {"name": "已写入", "level": "ok", "detail": f"{a} = {v * 100:+.2f}%"}
        for a, v, _got in actual[:20]
    ]

    bits = []
    if placed:
        bits.append(f"新建 {len(placed)} 个周次")
    if additions:
        bits.append(f"填 {len(additions)} 处空格")
    return Result(
        status="success",
        summary="已" + "、".join(bits) + "，格式与边框已处理，并加了来源批注。",
        domain=DOMAIN,
        checks=checks,
        warnings=notes[2:],
        next_steps=["底稿变了，接着跑 ir industry merge 重建快照。"],
        data={
            "written": len(additions),
            "new_weeks": list(placed),
            "skipped": len(conflicts),
            "backup": backup.name,
        },
    )
