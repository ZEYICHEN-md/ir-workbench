"""读指标底稿 Excel 的 2026 块。

布局（勿与旧版混淆）：
  - 周轴：R 列（周标签）
  - 酒店 STR：S / T / U（入住率 / ADR / RevPAR）
  - 航空：W / X / Y = 客运量 / 票价 / 客运航班量（固定列，全年含 Q1）
  - 左侧「QTD周度」G / H / I：仅作航空回退
  - 火车票 Z+：本管道不读

**本模块不再提供「按文件名取最新」**——底稿由 `ir config set industry` 显式锁定
（ADR 0001）。原 `find_latest_excel` / `_excel_version_key` 已随之删除。
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl

from . import layout

MONTH_LABELS = [f"{i}月" for i in range(1, 13)]

#: 2026 右侧周度航空固定列
RIGHT_AV_PAX_COL = 23
RIGHT_AV_TICKET_COL = 24
RIGHT_AV_FLIGHT_COL = 25

#: 左侧「QTD周度」六项列：B 周标签，C/D/E 酒店，G/H/I 航空
LEFT_QTD_LABEL_COL = 2
LEFT_QTD_COLS: tuple[tuple[str, int], ...] = (
    ("hotelOccupancy", 3),
    ("hotelADR", 4),
    ("hotelRevPAR", 5),
    ("aviationPax", 7),
    ("aviationTicket", 8),
    ("aviationFlight", 9),
)

#: 右侧周轴中间容忍的最大连续空行数。
#:
#: 给左侧「QTD周度」块加一周时，如果整行插入，右侧周轴就会被punch出一个空行——
#: 这是**每期都可能发生的常规人工操作**，不是异常。旧实现遇到空行即认为「块到此结束」，
#: 于是静默丢掉下方所有周次（实测 0830 底稿因此把数据截至日从 8/15 读成 6/13）。
#: 所以空行要跳过继续往下读；只有连续空到这个数以上，才认为块真的结束了。
MAX_AXIS_GAP = 3

_WEEK_RANGE = re.compile(r"(\d{1,2})/(\d{1,2})\s*-\s*(\d{1,2})/(\d{1,2})")

#: 指标序列名 → 月度/季度所在列
_MONTHLY_COLS: dict[str, int] = {
    "hotelOccupancy": 3,
    "hotelADR": 4,
    "hotelRevPAR": 5,
    "domAviationCAAC": 7,
    "domAviationBig3": 8,
    "railway": 9,
    "intlAviationCAAC": 11,
    "intlAviationBig3": 12,
    "intlCapacity": 13,
}


def _num(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace("%", "")
        if not text:
            return None
        try:
            return float(text) / 100.0 if "%" in value else float(text)
        except ValueError:
            return None
    return None


def infer_data_update(weeks: list, year: int = 2026) -> str | None:
    """最新一周的结束日，ISO 格式。跳过「春节 / 日均」。跨年（12→1 月）按 year+1。"""
    for label in reversed(weeks or []):
        text = str(label).strip()
        if "春节" in text or "日均" in text:
            continue
        match = _WEEK_RANGE.search(text)
        if not match:
            continue
        start_m, _start_d, end_m, end_d = (int(x) for x in match.groups())
        end_year = year + 1 if end_m < start_m else year
        return f"{end_year:04d}-{end_m:02d}-{end_d:02d}"
    return None


def norm_week(label: str) -> str:
    """去掉 M/D 的前导零；「春节(日均)」等原样保留。"""
    if not label or "春节" in label or "日均" in label:
        return label.strip() if isinstance(label, str) else label
    text = str(label).strip()

    def strip_part(part: str) -> str:
        if "/" not in part:
            return part
        left, right = part.split("/", 1)
        return f"{int(left)}/{int(right)}"

    if "-" in text:
        left, right = text.split("-", 1)
        return f"{strip_part(left.strip())}-{strip_part(right.strip())}"
    return text


class ExcelLayoutError(RuntimeError):
    """底稿结构与预期不符。按契约应停止，而不是猜测。"""


def _open_2026_block(xlsx_path: Path):
    """打开底稿并定位 2026 块。返回 (sheet, year_row)。"""
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    try:
        sheet = workbook["国内行业数据"]
    except KeyError:
        raise ExcelLayoutError("底稿缺少工作表「国内行业数据」") from None

    for row in range(1, sheet.max_row + 1):
        value = sheet.cell(row, 2).value
        if value and str(value).strip().startswith("2026"):
            return sheet, row
    raise ExcelLayoutError("底稿里找不到 2026 年块（B 列「2026年」）")


def weekly_sides(xlsx_path: Path) -> tuple[dict, dict[str, tuple]]:
    """右侧周度原值 + 左侧「QTD周度」航空原值，**不做回退合并**。

    `parse()` 返回的是合并后的结果，右侧优先会静默盖掉左侧，因此看不出两边是否一致。
    交叉核对必须拿两侧原值，所以单独开这个入口。
    """
    sheet, year_row = _open_2026_block(xlsx_path)
    return _read_right_weekly(sheet, year_row), _parse_weekly_aviation_fallback(sheet, year_row)


def parse(xlsx_path: Path) -> dict:
    """读出与指标快照同形状的字典（weekly / monthly / quarterly / meta）。"""
    sheet, year_row = _open_2026_block(xlsx_path)

    monthly = _parse_monthly(sheet, year_row)
    quarterly = _parse_quarterly(sheet, year_row)
    weekly, diagnostics = _parse_weekly(sheet, year_row)

    return {
        "meta": {"sourceExcel": xlsx_path.name},
        "weekly": weekly,
        "monthly": monthly,
        "quarterly": quarterly,
        # 读表过程中的提醒（周轴空行、右侧漏填靠左侧兜上……）。给人看，不进快照。
        "diagnostics": diagnostics,
    }


def _parse_monthly(sheet, year_row: int) -> dict:
    month_start = year_row + 3
    months: list[str] = []
    series: dict[str, list] = {name: [] for name in _MONTHLY_COLS}

    for row in range(month_start, month_start + 12):
        # 月份行标签可能带后缀（如「7月 (preliminary)」）；匹配规则在 layout 里只有一份
        number = layout.month_number(sheet.cell(row, 2).value)
        if number is None:
            break
        months.append(f"{number}月")
        for name, col in _MONTHLY_COLS.items():
            series[name].append(_num(sheet.cell(row, col).value))

    return {"months": months, **series}


def _parse_quarterly(sheet, year_row: int) -> dict:
    month_start = year_row + 3
    quarterly: dict[str, dict] = {}
    for row in range(month_start, month_start + 20):
        label = sheet.cell(row, 2).value
        if not label:
            continue
        text = str(label).strip().upper()
        if not re.match(r"^Q[1-4]$", text):
            continue
        values = {name: _num(sheet.cell(row, col).value) for name, col in _MONTHLY_COLS.items()}
        if any(v is not None for v in values.values()):
            quarterly[text.lower()] = values
    return quarterly


def _read_right_weekly(sheet, year_row: int) -> dict:
    """右侧周度区原值（R 周轴 + S/T/U 酒店 + W/X/Y 航空），未经回退合并。"""
    # 右侧周轴表头：R 列 == "周"
    header_row = None
    for row in range(year_row, year_row + 10):
        value = sheet.cell(row, 18).value
        if value and str(value).strip() == "周":
            header_row = row
            break
    if header_row is None:
        raise ExcelLayoutError("底稿里找不到右侧周度表头（R 列「周」）")

    weeks: list[str] = []
    occupancy: list = []
    adr: list = []
    revpar: list = []
    pax_right: list = []
    ticket_right: list = []
    flight_right: list = []

    scan_end = header_row + 80
    blank_run = 0
    blank_rows: list[int] = []
    last_week_row = header_row

    for row in range(header_row + 1, scan_end):
        label = sheet.cell(row, 18).value
        text = "" if label is None else str(label).strip()

        # 注释行是块尾的硬边界。必须在「春节 / 日均」判断之前——底稿末尾那行
        # `Notes：春节数据为民航局日均（含出境）` 同时含这两个词，顺序反了会被当成周次。
        if text.startswith(("Note", "注", "**")):
            break

        if text == "":
            if not weeks:
                continue
            blank_run += 1
            blank_rows.append(row)
            # 连续空到这个数，才认为块真的结束（单个空行通常是插行带出来的洞）
            if blank_run > MAX_AXIS_GAP:
                break
            continue

        if not re.search(r"\d+/\d+", text) and "春节" not in text and "日均" not in text:
            break

        blank_run = 0
        last_week_row = row
        weeks.append(text)
        occupancy.append(_num(sheet.cell(row, 19).value))
        adr.append(_num(sheet.cell(row, 20).value))
        revpar.append(_num(sheet.cell(row, 21).value))
        pax_right.append(_num(sheet.cell(row, RIGHT_AV_PAX_COL).value))
        ticket_right.append(_num(sheet.cell(row, RIGHT_AV_TICKET_COL).value))
        flight_right.append(_num(sheet.cell(row, RIGHT_AV_FLIGHT_COL).value))

    # 末周之后的空行是块尾留白，不是洞；只有夹在周次之间的才算。
    holes = [row for row in blank_rows if row < last_week_row]

    return {
        "weeks": weeks,
        "axisHoles": holes,
        "hotelOccupancy": occupancy,
        "hotelADR": adr,
        "hotelRevPAR": revpar,
        "aviationPax": pax_right,
        "aviationTicket": ticket_right,
        "aviationFlight": flight_right,
    }


def _week_start(label: str) -> tuple[int, int] | None:
    """周标签的起始 (月, 日)，用于判断先后。「春节(日均)」这类返回 None。"""
    match = _WEEK_RANGE.search(str(label))
    if not match:
        return None
    start_m, start_d, _end_m, _end_d = (int(x) for x in match.groups())
    return start_m, start_d


def _read_left_qtd(sheet, year_row: int) -> dict:
    """左侧「QTD周度」块：B 周标签 + C/D/E 酒店 + G/H/I 航空，六项与右侧同构。

    这一块的表头（第 185 行一带）写的就是「入住率 / ADR / RevPAR / 机票客运量 / 票价 /
    客运航班量」，与右侧 S/T/U/W/X/Y 一一对应，所以它能给右侧当完整的兜底，
    不只是航空兜底。
    """
    month_start = year_row + 3
    qtd_header = None
    for row in range(month_start, month_start + 40):
        value = sheet.cell(row, LEFT_QTD_LABEL_COL).value
        if value and "QTD" in str(value):
            qtd_header = row
            break
    if qtd_header is None:
        return {"weeks": [], **{name: [] for name, _ in LEFT_QTD_COLS}}

    weeks: list[str] = []
    series: dict[str, list] = {name: [] for name, _ in LEFT_QTD_COLS}
    for row in range(qtd_header + 1, qtd_header + 40):
        label = sheet.cell(row, LEFT_QTD_LABEL_COL).value
        text = "" if label is None else str(label).strip()
        if text.startswith(("Note", "注", "**")):
            break
        if text == "":
            if weeks:
                break
            continue
        if not re.search(r"\d+/\d+", text) and "春节" not in text:
            break
        weeks.append(text)
        for name, col in LEFT_QTD_COLS:
            series[name].append(_num(sheet.cell(row, col).value))
    return {"weeks": weeks, **series}


_WEEKLY_SERIES = (
    "hotelOccupancy",
    "hotelADR",
    "hotelRevPAR",
    "aviationPax",
    "aviationTicket",
    "aviationFlight",
)


def _parse_weekly(sheet, year_row: int) -> tuple[dict, list[str]]:
    """合并右侧周度区与左侧「QTD周度」块。右侧优先，左侧兜底。

    兜底分两层，都是为了兜住同一个人工习惯——「加了左侧，右侧忘了跟上」：

    1. **逐格兜底**：右侧某格为空、左侧同周同指标有值 → 用左侧的（六项全兜，不只航空）。
    2. **补周兜底**：某一周只在左侧有、右侧周轴上没有 → 把这一周接到轴末尾。
       只接**晚于右侧末周**的；更早的缺口位置在中间，接到末尾会打乱时间顺序，只报不接。

    返回 (weekly, diagnostics)。diagnostics 是给人看的提醒，不进快照。
    """
    right = _read_right_weekly(sheet, year_row)
    left = _read_left_qtd(sheet, year_row)

    weeks: list[str] = list(right["weeks"])
    series: dict[str, list] = {
        name: list(right[name]) + [None] * (len(weeks) - len(right[name]))
        for name in _WEEKLY_SERIES
    }

    left_by_week: dict[str, dict] = {}
    for index, label in enumerate(left["weeks"]):
        key = label if "春节" in label else norm_week(label)
        left_by_week[key] = {name: left[name][index] for name in _WEEKLY_SERIES}

    diagnostics: list[str] = []

    if right["axisHoles"]:
        rows = "、".join(f"第 {row} 行" for row in right["axisHoles"])
        diagnostics.append(
            f"右侧周轴中间有空行（{rows}），已跳过继续往下读。"
            "通常是给左侧「QTD周度」加一周时整行插入带出来的，建议删掉这一行。"
        )

    # 第 1 层：逐格兜底
    filled: dict[str, list[str]] = {}
    for index, label in enumerate(weeks):
        key = label if "春节" in label else norm_week(label)
        source = left_by_week.get(key) or left_by_week.get(label)
        if not source:
            continue
        for name in _WEEKLY_SERIES:
            if series[name][index] is None and source[name] is not None:
                series[name][index] = source[name]
                filled.setdefault(label, []).append(name)

    # 第 2 层：补右侧漏掉的整周
    right_keys = {label if "春节" in label else norm_week(label) for label in weeks}
    last_start = None
    for label in reversed(weeks):
        last_start = _week_start(label)
        if last_start:
            break

    appended: list[str] = []
    skipped_earlier: list[str] = []
    for index, label in enumerate(left["weeks"]):
        key = label if "春节" in label else norm_week(label)
        if key in right_keys:
            continue
        values = {name: left[name][index] for name in _WEEKLY_SERIES}
        if all(value is None for value in values.values()):
            continue
        start = _week_start(label)
        if last_start is not None and (start is None or start <= last_start):
            skipped_earlier.append(label)
            continue
        weeks.append(key)
        for name in _WEEKLY_SERIES:
            series[name].append(values[name])
        right_keys.add(key)
        last_start = start
        appended.append(key)

    if filled:
        detail = "；".join(
            f"{label} 的 {len(names)} 项" for label, names in list(filled.items())[:6]
        )
        diagnostics.append(
            f"有 {len(filled)} 周右侧有缺格、已用左侧「QTD周度」的值兜上（{detail}）。"
            "看板不会缺数，但右侧仍建议补齐——它才是这一块的主填写面。"
        )
    if appended:
        diagnostics.append(
            f"这些周只在左侧「QTD周度」里有，右侧周轴上没有，已按左侧的值接到轴末尾："
            + "、".join(appended)
            + "。请在右侧补上对应行，否则下一期插行时容易再错位。"
        )
    if skipped_earlier:
        diagnostics.append(
            "这些周只在左侧有，且早于右侧末周，**没有**并入（接到末尾会打乱时间顺序）："
            + "、".join(skipped_earlier)
            + "。要补得在右侧对应位置手工插行。"
        )

    return {"weeks": weeks, **series}, diagnostics


def _parse_weekly_aviation_fallback(sheet, year_row: int) -> dict[str, tuple]:
    """左侧「QTD周度」G/H/I，按周标签索引。供 `weekly_sides()` 做左右交叉核对。"""
    left = _read_left_qtd(sheet, year_row)
    return {
        (label if "春节" in label else norm_week(label)): (
            left["aviationPax"][index],
            left["aviationTicket"][index],
            left["aviationFlight"][index],
        )
        for index, label in enumerate(left["weeks"])
    }
