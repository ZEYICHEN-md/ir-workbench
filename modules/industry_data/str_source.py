"""从中金旅游周度数据表读 STR 酒店数据，供给底稿的周度与月度酒店三列。

## 为什么要这个模块

底稿的酒店 STR 数据此前全靠手抄：周度从中金表的 `Mainland China (STR)` tab 抄
K/L/M 三列同比，月度则要么等 STR 官方月报、要么等券商报告。前者慢，后者不可复现。
把「怎么从周报算月度」写成代码之后，拿到中金表就能自己算，不用等别人。

已验证：中金表 K/L/M 与底稿周度值**逐位一致**（如 2026/08/09-08/15 的
`-0.033643 / -0.0073863 / -0.040781`），所以自动化是在精确复现既有手工操作，不是换口径。

## 月度为什么不能简单平均周度同比

STR 的月度是按整月总量算的。正确做法回到绝对值（E–J 列有 Occ/ADR/RevPAR 的当年与上年）：

- 月度 Occ    = Σ(天数 × Occ) / Σ天数           —— 按天加权
- 月度 RevPAR = Σ(天数 × RevPAR) / Σ天数        —— 同理，分母是可售房晚
- 月度 ADR    = Σ(天数 × RevPAR) / Σ(天数 × Occ/100)
                = 总房间收入 / 总售出房晚        —— **不能按天平均**

前提是可售房晚在月内日均大致稳定（酒店供给短期不会跳变）。周是周日至周六，与月边界
不对齐，所以按**天**归属：跨月的周只把落在本月的天数计入。

## 基期口径：用「上年」列，不用去年同月

表里 F/H/J 是 STR 自己的「上年」（可比周）。实测另一条路——聚合去年同月的「当前年度」
列——与市场口径差很远（2026 年 2 月 ADR 差 8.5 个百分点），因为 STR 的 LY 定义是可比周
而非日历月。所以一律用「上年」列。

该口径已被券商报告独立印证：2026 年 6 月算出 `-4.13% / +1.88% / -2.32%`，券商发布
`-4% / +2% / -2%`；7 月算出 `-2.99% / -1.00% / -3.96%`，券商发布 `-3% / -1% / -4%`。
"""

from __future__ import annotations

import calendar
import datetime as dt
import re
from dataclasses import dataclass
from pathlib import Path

import openpyxl

#: 中金表里的 STR 中国大陆 tab
SHEET = "Mainland China (STR)"

#: 列位（1-based），取自表头 r3/r4
COL_WEEK = 1
COL_OCC_TY, COL_OCC_LY = 5, 6
COL_ADR_TY, COL_ADR_LY = 7, 8
COL_REV_TY, COL_REV_LY = 9, 10
#: K/L/M —— 表里自带的周度同比，底稿周度直接取这三列
COL_YOY_OCC, COL_YOY_ADR, COL_YOY_REV = 11, 12, 13

_WEEK_RE = re.compile(r"(\d{4})/(\d{1,2})/(\d{1,2})\s*-\s*(\d{4})/(\d{1,2})/(\d{1,2})")

#: 快照字段名 → 人读名
METRICS = (("hotelOccupancy", "入住率"), ("hotelADR", "ADR"), ("hotelRevPAR", "RevPAR"))


class StrSourceError(RuntimeError):
    """中金表结构与预期不符。按契约应停止，不猜。"""


@dataclass(frozen=True)
class WeekRow:
    row: int
    label: str
    start: dt.date
    end: dt.date
    occ_ty: float
    occ_ly: float
    adr_ty: float
    adr_ly: float
    rev_ty: float
    rev_ly: float
    yoy_occ: float | None
    yoy_adr: float | None
    yoy_rev: float | None

    def days_in(self, year: int, month: int) -> int:
        """这一周有几天落在指定月份。跨月的周只计入落在本月的天数。"""
        first = dt.date(year, month, 1)
        last = dt.date(year, month, calendar.monthrange(year, month)[1])
        lo, hi = max(self.start, first), min(self.end, last)
        return (hi - lo).days + 1 if lo <= hi else 0

    @property
    def short_label(self) -> str:
        """底稿周轴用的写法：`8/9-8/15`（无前导零，不含年份）。"""
        return f"{self.start.month}/{self.start.day}-{self.end.month}/{self.end.day}"


def _num(value) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    return None


def load(path: Path) -> list[WeekRow]:
    """读出全部周行。结构不符直接抛错，不返回半份数据。"""
    try:
        workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as error:  # noqa: BLE001 —— 打不开就原样报出
        raise StrSourceError(f"打不开中金表：{type(error).__name__}: {error}") from None

    try:
        if SHEET not in workbook.sheetnames:
            raise StrSourceError(
                f"中金表缺少工作表「{SHEET}」；现有 {len(workbook.sheetnames)} 个 tab。"
                "表结构改版时须先更新本模块的 SHEET 与列位契约。"
            )
        sheet = workbook[SHEET]
        rows: list[WeekRow] = []
        for index, values in enumerate(sheet.iter_rows(values_only=True), start=1):
            if not values or values[0] is None:
                continue
            match = _WEEK_RE.match(str(values[0]).strip())
            if not match:
                continue

            def col(number: int):
                return values[number - 1] if number - 1 < len(values) else None

            absolutes = [_num(col(c)) for c in (
                COL_OCC_TY, COL_OCC_LY, COL_ADR_TY, COL_ADR_LY, COL_REV_TY, COL_REV_LY)]
            if any(v is None for v in absolutes):
                continue  # 表尾常有空行或小计行

            y1, m1, d1, y2, m2, d2 = (int(x) for x in match.groups())
            rows.append(
                WeekRow(
                    row=index,
                    label=str(values[0]).strip(),
                    start=dt.date(y1, m1, d1),
                    end=dt.date(y2, m2, d2),
                    occ_ty=absolutes[0],
                    occ_ly=absolutes[1],
                    adr_ty=absolutes[2],
                    adr_ly=absolutes[3],
                    rev_ty=absolutes[4],
                    rev_ly=absolutes[5],
                    yoy_occ=_num(col(COL_YOY_OCC)),
                    yoy_adr=_num(col(COL_YOY_ADR)),
                    yoy_rev=_num(col(COL_YOY_REV)),
                )
            )
    finally:
        workbook.close()

    if not rows:
        raise StrSourceError(
            f"在「{SHEET}」里没读到任何周行。A 列应为 `YYYY/MM/DD-YYYY/MM/DD` 格式。"
        )
    return rows


@dataclass(frozen=True)
class MonthAggregate:
    year: int
    month: int
    days: int
    expected_days: int
    occ_ty: float
    occ_ly: float
    adr_ty: float
    adr_ly: float
    rev_ty: float
    rev_ly: float
    #: 参与聚合的 (周标签, 计入天数)
    parts: tuple[tuple[str, int], ...]

    @property
    def complete(self) -> bool:
        """本月的每一天是否都被某一周覆盖到。不完整则不该入库。"""
        return self.days == self.expected_days

    def yoy(self) -> dict[str, float]:
        return {
            "hotelOccupancy": self.occ_ty / self.occ_ly - 1.0,
            "hotelADR": self.adr_ty / self.adr_ly - 1.0,
            "hotelRevPAR": self.rev_ty / self.rev_ly - 1.0,
        }


def aggregate_month(weeks: list[WeekRow], year: int, month: int) -> MonthAggregate | None:
    """按天加权聚合出该月的 Occ / ADR / RevPAR（当年与上年）。无数据返回 None。"""
    total_days = 0
    sums = {"rev_ty": 0.0, "rev_ly": 0.0, "occ_ty": 0.0, "occ_ly": 0.0}
    parts: list[tuple[str, int]] = []

    for week in weeks:
        days = week.days_in(year, month)
        if not days:
            continue
        total_days += days
        sums["rev_ty"] += days * week.rev_ty
        sums["rev_ly"] += days * week.rev_ly
        sums["occ_ty"] += days * week.occ_ty
        sums["occ_ly"] += days * week.occ_ly
        parts.append((week.short_label, days))

    if not total_days:
        return None

    return MonthAggregate(
        year=year,
        month=month,
        days=total_days,
        expected_days=calendar.monthrange(year, month)[1],
        occ_ty=sums["occ_ty"] / total_days,
        occ_ly=sums["occ_ly"] / total_days,
        # ADR = 总房间收入 / 总售出房晚；Occ 是百分数所以除 100
        adr_ty=sums["rev_ty"] / (sums["occ_ty"] / 100.0),
        adr_ly=sums["rev_ly"] / (sums["occ_ly"] / 100.0),
        rev_ty=sums["rev_ty"] / total_days,
        rev_ly=sums["rev_ly"] / total_days,
        parts=tuple(parts),
    )


def weekly_yoy(weeks: list[WeekRow], year: int) -> list[tuple[str, dict[str, float | None]]]:
    """某年各周的同比，直接取表里的 K/L/M（与底稿现有值逐位一致）。"""
    out: list[tuple[str, dict[str, float | None]]] = []
    for week in weeks:
        # 周次归属按结束日所在年份，与底稿周轴一致（跨年周 12/27-1/2 归入新年）
        if week.end.year != year:
            continue
        out.append((
            week.short_label,
            {
                "hotelOccupancy": week.yoy_occ,
                "hotelADR": week.yoy_adr,
                "hotelRevPAR": week.yoy_rev,
            },
        ))
    return out


def complete_months(weeks: list[WeekRow], year: int) -> list[MonthAggregate]:
    """该年所有**已被周数据完整覆盖**的月份。不完整的月不返回——半个月的聚合没有意义。"""
    out: list[MonthAggregate] = []
    for month in range(1, 13):
        agg = aggregate_month(weeks, year, month)
        if agg and agg.complete:
            out.append(agg)
    return out
