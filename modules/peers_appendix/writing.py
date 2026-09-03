"""Mechanical Model → brief → gated texts → ticker-specific Word apply flow.

Scripts do not compose earnings narrative.  They expose model facts and stable
template slots; a human/agent writes ``texts.json`` from current-quarter
materials, then the gates verify completeness and grounding before Word is
patched.
"""

from __future__ import annotations

import json
import re
import shutil
import tempfile
from dataclasses import dataclass
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from .ir_snapshot import check_must_cover
from .model_common import detect_layout


BKNG_ROWS = {
    "gbv_mn": 6,
    "gbv_yoy": 7,
    "gbv_yoy_cc": 8,
    "agency_gbv": 14,
    "agency_gbv_yoy": 15,
    "agency_gbv_share": 16,
    "merchant_gbv": 17,
    "merchant_gbv_yoy": 18,
    "merchant_gbv_share": 19,
    "revenue_mn": 30,
    "revenue_yoy": 31,
    "agency_rev": 39,
    "agency_rev_yoy": 40,
    "agency_rev_share": 41,
    "merchant_rev": 42,
    "merchant_rev_yoy": 43,
    "merchant_rev_share": 44,
    "ads_rev": 45,
    "ads_rev_yoy": 46,
    "ads_rev_share": 47,
    "take_rate": 49,
    "take_rate_ttm": 53,
    "mkt_exp": 60,
    "sales_exp": 61,
    "personnel_exp": 62,
    "ga_exp": 63,
    "it_exp": 64,
    "adj_ebitda_mn": 86,
    "adj_ebitda_margin": 88,
    "adj_eps": 122,
    "room_nights_mn": 235,
    "room_nights_yoy": 236,
    "air_tickets_mn": 238,
    "air_tickets_yoy": 239,
    "rental_days_mn": 241,
    "rental_days_yoy": 242,
}

EXPE_ROWS = {
    "gbv_mn": 5,
    "gbv_yoy": 6,
    "gbv_yoy_cc": 7,
    "agency_gbv": 33,
    "agency_gbv_yoy": 34,
    "agency_gbv_share": 35,
    "merchant_gbv": 37,
    "merchant_gbv_yoy": 38,
    "merchant_gbv_share": 39,
    "revenue_mn": 59,
    "revenue_yoy": 60,
    "us_rev": 66,
    "us_rev_yoy": 67,
    "us_rev_share": 68,
    "intl_rev": 71,
    "intl_rev_yoy": 72,
    "intl_rev_share": 73,
    "hotel_rev": 113,
    "hotel_rev_yoy": 114,
    "hotel_rev_share": 116,
    "air_rev": 117,
    "air_rev_yoy": 118,
    "air_rev_share": 120,
    "ads_rev": 121,
    "ads_rev_yoy": 122,
    "ads_rev_share": 124,
    "other_rev": 125,
    "other_rev_yoy": 126,
    "other_rev_share": 127,
    "b2c_rev": 130,
    "b2c_rev_yoy": 131,
    "b2c_rev_share": 132,
    "b2b_rev": 134,
    "b2b_rev_yoy": 135,
    "b2b_rev_share": 136,
    "adj_ebitda_mn": 180,
    "adj_ebitda_margin": 182,
    "cos_pct_rev": 231,
    "pd_pct_rev": 232,
    "sm_pct_rev": 233,
    "ga_pct_rev": 234,
    "room_nights_mn": 307,
    "room_nights_yoy": 308,
    "adr": 317,
    "adr_yoy": 318,
    "air_tickets_mn": 323,
    "air_tickets_yoy": 324,
}

# The old extractor had no ABNB map, which made its writing phase un-runnable.
# These are the same stable rows used by the authoritative ABNB model audit.
ABNB_ROWS = {
    "nights_mn": 3,
    "nights_yoy": 4,
    "gbv_mn": 6,
    "gbv_yoy": 7,
    "take_rate": 9,
    "adr": 11,
    "revenue_mn": 15,
    "revenue_yoy": 16,
    "adj_ebitda_mn": 55,
    "adj_ebitda_margin": 56,
    "ops_exp": 22,
    "pd_exp": 23,
    "sm_exp": 24,
    "ga_exp": 25,
}

ROW_MAPS = {"BKNG": BKNG_ROWS, "EXPE": EXPE_ROWS, "ABNB": ABNB_ROWS}


def prior_year_quarter(quarter: str) -> str:
    match = re.fullmatch(r"(20\d{2})Q([1-4])", quarter)
    if not match:
        raise ValueError(f"季度格式不对：{quarter}")
    return f"{int(match.group(1)) - 1}Q{match.group(2)}"


def _safe_div(value, base):
    if value is None or base in (None, 0):
        return None
    return value / base - 1.0


def extract_model_facts(model: Path, ticker: str, quarter: str) -> dict:
    """Read current and prior-year model facts without mutating the workbook."""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    ticker = ticker.upper()
    rows = ROW_MAPS.get(ticker)
    if rows is None:
        raise ValueError(
            f"{ticker} 尚无写作 row map；已支持 {', '.join(ROW_MAPS)}。"
        )
    layout = detect_layout(model, ticker)
    prior = prior_year_quarter(quarter)
    if quarter not in layout["quarters"]:
        raise ValueError(f"{ticker} Model 缺 {quarter}。")
    if prior not in layout["quarters"]:
        raise ValueError(f"{ticker} Model 缺同比基期 {prior}。")
    current_column = layout["quarters"][quarter]
    prior_column = layout["quarters"][prior]
    workbook = load_workbook(model, data_only=True, read_only=True)
    try:
        worksheet = workbook[ticker]
        current = {
            key: worksheet.cell(row, current_column).value
            for key, row in rows.items()
        }
        prior_values = {
            key: worksheet.cell(row, prior_column).value
            for key, row in rows.items()
        }
    finally:
        workbook.close()
    for derived_key, absolute_key in (
        ("adj_ebitda_yoy", "adj_ebitda_mn"),
        ("air_tickets_yoy", "air_tickets_mn"),
    ):
        if derived_key not in current or current.get(derived_key) is None:
            if absolute_key in current:
                current[derived_key] = _safe_div(
                    current.get(absolute_key),
                    prior_values.get(absolute_key),
                )
    margin_change = None
    if (
        current.get("adj_ebitda_margin") is not None
        and prior_values.get("adj_ebitda_margin") is not None
    ):
        margin_change = (
            current["adj_ebitda_margin"]
            - prior_values["adj_ebitda_margin"]
        ) * 100
    return {
        "ticker": ticker,
        "quarter": quarter,
        "quarter_short": quarter[2:],
        "prior_quarter": prior,
        "prior_short": prior[2:],
        "q": int(quarter[-1]),
        "col": get_column_letter(current_column),
        "prior_col": get_column_letter(prior_column),
        "model": str(model),
        "current": current,
        "prior": prior_values,
        "derived": {
            "margin_pp_vs_prior": margin_change,
            "margin_bps_vs_prior": (
                margin_change * 100 if margin_change is not None else None
            ),
        },
    }


EXPE_SLOTS = [
    {"id": "ops_h1", "role": "标题"},
    {
        "id": "ops_lodging_h",
        "role": "住宿标题句",
        "need_from": ["model", "materials"],
    },
    {
        "id": "ops_lodging_geo",
        "role": "分地区",
        "need_from": ["materials"],
    },
    {
        "id": "ops_lodging_adr",
        "role": "住宿交易额与 ADR",
        "need_from": ["model", "materials"],
    },
    {"id": "ops_air", "role": "机票", "need_from": ["model", "materials"]},
    {"id": "guide_next_h", "role": "下季指引标题"},
    {
        "id": "guide_next_b0",
        "role": "下季收入/交易额",
        "need_from": ["materials"],
    },
    {
        "id": "guide_next_b1",
        "role": "下季利润率",
        "need_from": ["materials"],
    },
    {"id": "guide_fy_h", "role": "全年指引标题"},
    {
        "id": "guide_fy_b0",
        "role": "全年收入/交易额",
        "need_from": ["materials"],
    },
    {
        "id": "guide_fy_b1",
        "role": "全年利润率",
        "need_from": ["materials"],
    },
    {"id": "gbv_h", "role": "交易额标题"},
    {"id": "gbv_total", "role": "总交易额", "need_from": ["model"]},
    {"id": "gbv_mix", "role": "交易额模式占比", "need_from": ["model"]},
    {
        "id": "gbv_yoy_mode",
        "role": "交易额分模式同比",
        "need_from": ["model"],
    },
    {"id": "rev_h", "role": "收入标题"},
    {"id": "rev_total", "role": "总收入", "need_from": ["model"]},
    {"id": "rev_by_line_h", "role": "业务线表标题"},
    {"id": "rev_by_seg_h", "role": "业务类型表标题"},
    {"id": "rev_by_geo_h", "role": "地区标题"},
    {"id": "rev_geo_yoy", "role": "地区同比", "need_from": ["model"]},
    {"id": "rev_geo_share", "role": "地区占比", "need_from": ["model"]},
    {"id": "profit_h", "role": "利润与费用标题"},
    {
        "id": "profit_ebitda",
        "role": "Adj. EBITDA",
        "need_from": ["model"],
    },
    {
        "id": "profit_opex_intro",
        "role": "费用总括",
        "need_from": ["model", "materials"],
    },
    {
        "id": "profit_reason_0",
        "role": "费用原因一",
        "need_from": ["materials"],
    },
    {
        "id": "profit_reason_1",
        "role": "费用原因二",
        "need_from": ["materials"],
    },
    {
        "id": "profit_reason_2",
        "role": "费用原因三",
        "need_from": ["materials"],
        "optional": True,
    },
]

BKNG_SLOT_IDS = [
    "fin_h1", "fin_results_h", "fin_results_body", "fin_guide_h",
    "fin_guide_b0", "fin_guide_b1", "fin_guide_b2", "fin_guide_b3",
    "fin_fy_h", "fin_fy_b0", "fin_fy_b1", "fin_fy_intro",
    "fin_fy_t0", "fin_fy_t1", "fin_fy_t2", "fin_fy_t3",
    "ops_h1", "ops_rn_h", "ops_rn_fact", "ops_adr", "ops_gbv_h",
    "ops_gbv_fact", "ops_gbv_d0", "ops_gbv_d1", "ops_gbv_d2",
    "ops_gbv_d3", "ops_gbv_mode_h", "ops_gbv_agency",
    "ops_gbv_merchant", "ops_rev_h", "ops_rev_fact", "ops_rev_lag",
    "ops_rev_mode_h", "ops_rev_agency", "ops_rev_merchant",
    "ops_rev_ads", "ops_take_h", "ops_take_body", "ops_profit_h",
    "ops_ebitda", "ops_exp_h", "ops_exp_mkt", "ops_exp_sales",
    "ops_exp_personnel", "ops_exp_ga", "ops_exp_it",
]

# Deliberately excludes ABNB strategy slots.  The dedicated apply code is used,
# but this pipeline patches only operations/finance and leaves mentor text alone.
ABNB_SLOT_IDS = [
    "ops_h", "nights_h", "nights_body", "adr_h", "adr_body",
    "adr_regional", "take_h", "take_body", "fin_h", "fin_lede",
    "guide_h", "guide_next_h", "guide_next_rev", "guide_next_gbv",
    "guide_next_adr", "guide_next_ebitda", "guide_fy_h",
    "guide_fy_rev", "guide_fy_margin", "results_h", "gbv_h",
    "gbv_body", "rev_body", "profit_h", "profit_ebitda", "opex_h",
    "opex_body", "opex_sm",
]


def _share(value) -> str | None:
    return None if value is None else f"{value * 100:.0f}%"


def _growth(value) -> str | None:
    if value is None:
        return None
    return f"{value * 100:.1f}%".replace(".0%", "%")


def build_writing_brief(
    facts: dict,
    snapshot: dict,
) -> dict:
    ticker = facts["ticker"]
    if ticker == "EXPE":
        slots = EXPE_SLOTS
        current, prior = facts["current"], facts["prior"]
        tables = [
            {
                "id": "rev_by_line",
                "kind": "share_yoy",
                "header_quarter": facts["quarter_short"],
                "rows": [
                    [
                        "住宿",
                        _share(current.get("hotel_rev_share")),
                        _growth(current.get("hotel_rev_yoy")),
                    ],
                    [
                        "广告",
                        _share(current.get("ads_rev_share")),
                        _growth(current.get("ads_rev_yoy")),
                    ],
                    [
                        "机票",
                        _share(current.get("air_rev_share")),
                        _growth(current.get("air_rev_yoy")),
                    ],
                ],
            },
            {
                "id": "rev_by_seg",
                "kind": "share_yoy",
                "header_quarter": facts["quarter_short"],
                "rows": [
                    [
                        "B2C",
                        _share(current.get("b2c_rev_share")),
                        _growth(current.get("b2c_rev_yoy")),
                    ],
                    [
                        "B2B",
                        _share(current.get("b2b_rev_share")),
                        _growth(current.get("b2b_rev_yoy")),
                    ],
                ],
            },
            {
                "id": "opex_pct",
                "kind": "opex",
                "header_quarter": facts["quarter_short"],
                "prior_quarter": facts["prior_short"],
                "rows": [
                    [
                        "主营",
                        _share(current.get("cos_pct_rev")),
                        _share(prior.get("cos_pct_rev")),
                    ],
                    [
                        "产品",
                        _share(current.get("pd_pct_rev")),
                        _share(prior.get("pd_pct_rev")),
                    ],
                    [
                        "营销",
                        _share(current.get("sm_pct_rev")),
                        _share(prior.get("sm_pct_rev")),
                    ],
                    [
                        "行政",
                        _share(current.get("ga_pct_rev")),
                        _share(prior.get("ga_pct_rev")),
                    ],
                ],
            },
        ]
    elif ticker == "BKNG":
        slots = [
            {
                "id": slot_id,
                "role": slot_id,
                "need_from": ["materials"] if "guide" in slot_id else ["model"],
            }
            for slot_id in BKNG_SLOT_IDS
        ]
        tables = []
    elif ticker == "ABNB":
        slots = [
            {
                "id": slot_id,
                "role": slot_id,
                "need_from": (
                    ["materials"]
                    if "guide" in slot_id or slot_id in {"fin_lede", "opex_sm"}
                    else ["model"]
                ),
            }
            for slot_id in ABNB_SLOT_IDS
        ]
        tables = []
    else:
        raise ValueError(f"{ticker} 没有写作 brief 配置。")
    return {
        "ticker": ticker,
        "quarter": facts["quarter"],
        "scope": "ops_finance",
        "mechanical_only": True,
        "instructions": (
            "只把 Model 数字和当前季度材料整理成事实 brief；"
            "正文必须由人/Agent 依据材料填写 texts.json，系统不自动编叙事。"
        ),
        "facts": facts,
        "guidance": snapshot.get("guidance") or {},
        "must_cover_in_writing": snapshot.get("must_cover_in_writing") or [],
        "slots": slots,
        "tables": tables,
        "texts_skeleton": {
            "ticker": ticker,
            "quarter": facts["quarter"],
            "scope": "ops_finance",
            "paragraphs": [
                {"id": slot["id"], "text": ""} for slot in slots
            ],
            "tables": tables,
        },
    }


def write_brief(brief: dict, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(brief, ensure_ascii=False, indent=2, default=str) + "\n",
        encoding="utf-8",
        newline="\n",
    )
    return path


def load_texts(path: Path) -> dict:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(payload, list):
        payload = {"paragraphs": payload, "tables": []}
    if not isinstance(payload, dict):
        raise ValueError("texts.json 必须是 object 或 paragraph list。")
    if not isinstance(payload.get("paragraphs"), list):
        raise ValueError("texts.json 缺 paragraphs[]。")
    return payload


def validate_strategy_decision(path: Path) -> dict:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError("strategy_decision.json 必须是 object。")
    allowed = {"preserve-template", "mentor-supplied", "out-of-scope"}
    if payload.get("decision") not in allowed:
        raise ValueError(
            "strategy_decision.json 的 decision 必须是 "
            "preserve-template / mentor-supplied / out-of-scope。"
        )
    if payload.get("confirmed_by_human") is not True:
        raise ValueError(
            "strategy_decision.json 必须有 confirmed_by_human: true；"
            "系统不能代替人决定战略段。"
        )
    return payload


def check_writing(
    brief: dict,
    texts: dict,
    quarter: str,
    snapshot: dict,
    materials_dir: Path,
    *,
    min_materials_length: int = 12,
) -> list[str]:
    errors: list[str] = []
    if texts.get("ticker", "").upper() != brief["ticker"]:
        errors.append(
            f"texts.ticker={texts.get('ticker')}，应为 {brief['ticker']}"
        )
    if texts.get("quarter") != quarter:
        errors.append(
            f"texts.quarter={texts.get('quarter')}，应为 {quarter}"
        )
    paragraphs = texts.get("paragraphs") or []
    by_id: dict[str, str] = {}
    for paragraph in paragraphs:
        if not isinstance(paragraph, dict) or not paragraph.get("id"):
            errors.append("texts.paragraphs 有缺 id 的条目")
            continue
        slot_id = str(paragraph["id"])
        if slot_id in by_id:
            errors.append(f"paragraph id 重复：{slot_id}")
        by_id[slot_id] = str(paragraph.get("text") or "")
    for slot in brief.get("slots", []):
        slot_id = slot["id"]
        text = by_id.get(slot_id, "").strip()
        if slot.get("optional") and not text:
            continue
        if not text:
            errors.append(f"missing/empty paragraph {slot_id}")
            continue
        if (
            "materials" in (slot.get("need_from") or [])
            and len(text) < min_materials_length
        ):
            errors.append(
                f"materials slot {slot_id} too short ({len(text)} chars)"
            )
    expected_tables = {
        table["id"]: table for table in brief.get("tables", [])
    }
    actual_tables = {
        table.get("id"): table for table in texts.get("tables", [])
    }
    for table_id, expected in expected_tables.items():
        actual = actual_tables.get(table_id)
        if actual is None:
            errors.append(f"missing table {table_id}")
        elif len(actual.get("rows", [])) != len(expected.get("rows", [])):
            errors.append(f"table {table_id} row count mismatch")
    body = "\n".join(by_id.values())
    for finding in check_must_cover(
        body,
        snapshot,
        materials_dir=materials_dir,
        scope="ops_finance",
    ):
        if finding["severity"] in {"FAIL", "WARN"}:
            errors.append(
                f"{finding['kind']} {finding.get('label')}: "
                f"{finding.get('missing') or finding.get('note')}"
            )
    return errors


@dataclass(frozen=True)
class ApplyRoute:
    ticker: str
    name: str
    start_patterns: tuple[re.Pattern[str], ...]
    anchors: tuple[tuple[str, re.Pattern[str]], ...]
    table_mode: str


def _compile_anchor_pairs(items: list[tuple[str, str]]) -> tuple:
    return tuple((slot, re.compile(pattern)) for slot, pattern in items)


BKNG_ANCHORS = _compile_anchor_pairs([
    ("fin_h1", r"^主要财务数据更新"),
    ("fin_results_h", r"^Q[1-4]业绩"),
    ("fin_results_body", r"^[一二三四]季度Booking业绩"),
    ("fin_guide_h", r"^\d{2}Q[1-4]指引"),
    ("fin_guide_b0", r"^收入同比增长"),
    ("fin_guide_b1", r"^Adj\. EBITDA同比增长"),
    ("fin_guide_b2", r"^[二三四]季度迄今|需求仍"),
    ("fin_guide_b3", r"^指引假设"),
    ("fin_fy_h", r"^公司(更新|维持).+增长指引"),
    ("fin_fy_b0", r"^全年交易额"),
    ("fin_fy_b1", r"^预计外汇波动"),
    ("fin_fy_intro", r"^财务口径下"),
    ("fin_fy_t0", r"^收入同比增长"),
    ("fin_fy_t1", r"^交易额同比增长"),
    ("fin_fy_t2", r"^Adj\. EBITDA同比增长"),
    ("fin_fy_t3", r"^EPS同比增长"),
    ("ops_h1", r"^主要财务和运营数据更新|^主要运营数据更新"),
    ("ops_rn_h", r"^住宿间夜"),
    ("ops_rn_fact", r"^\d{2}Q[1-4]间夜量为"),
    ("ops_adr", r"^ADR"),
    ("ops_gbv_h", r"^交易额及增长"),
    ("ops_gbv_fact", r"^Q[1-4]总交易额为"),
    ("ops_gbv_d0", r"^汇率"),
    ("ops_gbv_d1", r"^\d+%的间夜"),
    ("ops_gbv_d2", r"^恒定汇率下"),
    ("ops_gbv_d3", r"^机票及其他"),
    ("ops_gbv_mode_h", r"^按照业务模式分布"),
    ("ops_gbv_agency", r"^Q[1-4] Agency model交易额"),
    ("ops_gbv_merchant", r"^Q[1-4] Merchant model交易额"),
    ("ops_rev_h", r"^收入及增长"),
    ("ops_rev_fact", r"^Q[1-4]收入为"),
    ("ops_rev_lag", r"^收入(增长)?增速低于|^收入增长慢于"),
    ("ops_rev_mode_h", r"^按照业务模式分布$"),
    ("ops_rev_agency", r"^Agency model收入"),
    ("ops_rev_merchant", r"^Merchant model收入"),
    ("ops_rev_ads", r"^广告及其他收入"),
    ("ops_take_h", r"^Take rate"),
    ("ops_take_body", r"^计算所得"),
    ("ops_profit_h", r"^利润率和费用"),
    ("ops_ebitda", r"^Q[1-4] Adjusted EBITDA"),
    ("ops_exp_h", r"^主要费用"),
    ("ops_exp_mkt", r"^营销费用"),
    ("ops_exp_sales", r"^销售及其他费用"),
    ("ops_exp_personnel", r"^人员费用"),
    ("ops_exp_ga", r"^行政费用"),
    ("ops_exp_it", r"^IT费用"),
])

EXPE_ANCHORS = _compile_anchor_pairs([
    ("ops_h1", r"^业务运营"),
    ("ops_lodging_h", r"^住宿"),
    ("ops_lodging_geo", r"美国市场|EMEA|其他地区"),
    ("ops_lodging_adr", r"^住宿交易额|^ADR"),
    ("ops_air", r"^机票"),
    ("guide_next_h", r"^公司预计\d{2}Q[1-4]"),
    ("guide_next_b0", r"^收入同比增长"),
    ("guide_next_b1", r"^Adj\. EBITDA Margin"),
    ("guide_fy_h", r"^公司预计20\d{2}年"),
    ("guide_fy_b0", r"^全年收入同比增长|^全年交易额"),
    ("guide_fy_b1", r"^Adj\. EBITDA Margin"),
    ("gbv_h", r"^交易额及增长"),
    ("gbv_total", r"^总交易额及增长|^Q[1-4]总交易额"),
    ("gbv_mix", r"Agency model.*Merchant model.*占比"),
    ("gbv_yoy_mode", r"^较\d{2}Q[1-4]"),
    ("rev_h", r"^收入及增长"),
    ("rev_total", r"^Q[1-4]收入为"),
    ("rev_by_line_h", r"^总收入按照业务线"),
    ("rev_by_seg_h", r"^总收入按照业务类型"),
    ("rev_by_geo_h", r"^总收入按照地域"),
    ("rev_geo_yoy", r"^Q[1-4]美国市场收入"),
    ("rev_geo_share", r"^Q[1-4]美国国内"),
    ("profit_h", r"^利润率和费用占比"),
    ("profit_ebitda", r"^Q[1-4] Adj\. EBITDA"),
    ("profit_opex_intro", r"主要费用"),
    ("profit_reason_0", r"^持续提升|^营销效率|^支付"),
    ("profit_reason_1", r"^B2C"),
    ("profit_reason_2", r"^前期降本|^降本增效"),
])

ABNB_ANCHORS = _compile_anchor_pairs([
    ("ops_h", r"^公司运营数据更新"),
    ("nights_h", r"^各板块恢复情况"),
    ("nights_body", r"^\d{2}Q[1-4]间夜及体验预订达到"),
    ("adr_h", r"^ADR$"),
    ("adr_body", r"^整体：.*ADR|^Q[1-4] ADR"),
    ("adr_regional", r"^分地区"),
    ("take_h", r"^Take rate$"),
    ("take_body", r"^计算得出.*Take rate"),
    ("fin_h", r"^公司财务数据更新"),
    ("fin_lede", r"^Airbnb[一二三四]季度各业绩|^Airbnb[一二三四]季度"),
    ("guide_h", r"^业绩指引"),
    ("guide_next_h", r"^公司预计\d{2}Q[1-4]"),
    ("guide_next_rev", r"^收入在|^收入同比"),
    ("guide_next_gbv", r"^预订增速|^GBV|^总交易额增速"),
    ("guide_next_adr", r"^ADR同比"),
    ("guide_next_ebitda", r"^Adj\.? ?EBITDA"),
    ("guide_fy_h", r"^公司预计20\d{2}年全年"),
    ("guide_fy_rev", r"^收入增速预期|^全年收入"),
    ("guide_fy_margin", r"^Adj\.? ?EBITDA [Mm]argin"),
    ("results_h", r"^[一二三四]季度业绩"),
    ("gbv_h", r"^交易额及收入增长|^交易额及增长"),
    ("gbv_body", r"^总交易额及增长|^Q[1-4]总交易额"),
    ("rev_body", r"^总收入及增长|^Q[1-4]收入为"),
    ("profit_h", r"^利润率和费用占比"),
    ("profit_ebitda", r"Adj\.? EBITDA"),
    ("opex_h", r"^主要费用|^Q[1-4]主要费用"),
    ("opex_body", r"运营支持费用和行政费用|运营支持费用收入占比"),
    ("opex_sm", r"主要费用中.*销售及营销费用同比"),
])


def select_apply_route(ticker: str) -> ApplyRoute:
    normalized = ticker.upper()
    if normalized == "EXPE":
        return ApplyRoute(
            normalized,
            "expe_dedicated",
            (re.compile(r"^业务运营"),),
            EXPE_ANCHORS,
            "expe",
        )
    if normalized == "ABNB":
        return ApplyRoute(
            normalized,
            "abnb_dedicated",
            (re.compile(r"^公司运营数据更新"),),
            ABNB_ANCHORS,
            "sequential",
        )
    if normalized == "BKNG":
        return ApplyRoute(
            normalized,
            "bkng_generic",
            (re.compile(r"^主要财务数据更新"),),
            BKNG_ANCHORS,
            "none",
        )
    raise ValueError(
        f"{normalized} 没有 Word apply 路由；不能假装通用支持。"
    )


NS = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
A_BLIP = "{http://schemas.openxmlformats.org/drawingml/2006/main}blip"


def _paragraph_text(paragraph) -> str:
    return "".join(
        node.text or "" for node in paragraph.findall(".//w:t", NS)
    )


def _set_paragraph_text(paragraph, text: str) -> None:
    from lxml import etree

    runs = paragraph.findall("w:r", NS)
    if not runs:
        run = etree.SubElement(paragraph, f"{W}r")
        node = etree.SubElement(run, f"{W}t")
        node.set(
            "{http://www.w3.org/XML/1998/namespace}space", "preserve"
        )
        node.text = text
        return
    first = runs[0]
    for run in runs[1:]:
        paragraph.remove(run)
    for child in list(first):
        if child.tag != f"{W}rPr":
            first.remove(child)
    node = etree.SubElement(first, f"{W}t")
    node.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
    node.text = text


def _cell_text(cell) -> str:
    return "".join(node.text or "" for node in cell.findall(".//w:t", NS))


def _set_cell_text(cell, text: str) -> None:
    from lxml import etree

    paragraphs = cell.findall("w:p", NS)
    if not paragraphs:
        paragraphs = [etree.SubElement(cell, f"{W}p")]
    _set_paragraph_text(paragraphs[0], str(text))
    for paragraph in paragraphs[1:]:
        cell.remove(paragraph)


def _patch_expe_tables(children: list, specs: list[dict]) -> int:
    share_specs = [item for item in specs if item.get("kind") == "share_yoy"]
    opex_specs = [item for item in specs if item.get("kind") == "opex"]
    share_index = opex_index = patched = 0
    for element in children:
        if element.tag != f"{W}tbl":
            continue
        rows = element.findall("w:tr", NS)
        if not rows:
            continue
        header_cells = rows[0].findall("w:tc", NS)
        header = "".join(_cell_text(cell) for cell in header_cells)
        first_label = (
            _cell_text(rows[1].findall("w:tc", NS)[0])
            if len(rows) > 1 and rows[1].findall("w:tc", NS)
            else ""
        )
        if (
            share_index < len(share_specs)
            and ("占比" in header or first_label in {"住宿", "B2C"})
            and "As %" not in header
            and "YoY%" not in header
        ):
            spec = share_specs[share_index]
            share_index += 1
            if header_cells:
                _set_cell_text(header_cells[0], spec["header_quarter"])
            for row_index, data in enumerate(spec["rows"], 1):
                if row_index >= len(rows):
                    break
                cells = rows[row_index].findall("w:tc", NS)
                if len(cells) >= 3 and len(data) >= 3:
                    _set_cell_text(cells[1], data[1])
                    _set_cell_text(cells[2], data[2])
            patched += 1
        elif opex_index < len(opex_specs) and (
            "As %" in header or "YoY%" in header
        ):
            spec = opex_specs[opex_index]
            opex_index += 1
            if len(header_cells) >= 3:
                _set_cell_text(header_cells[1], spec["header_quarter"])
                _set_cell_text(header_cells[2], spec["prior_quarter"])
            for row_index, data in enumerate(spec["rows"], 1):
                if row_index >= len(rows):
                    break
                cells = rows[row_index].findall("w:tc", NS)
                if len(cells) >= 3 and len(data) >= 3:
                    _set_cell_text(cells[1], data[1])
                    _set_cell_text(cells[2], data[2])
            patched += 1
    return patched


def _patch_sequential_tables(children: list, specs: list[dict]) -> int:
    tables = [element for element in children if element.tag == f"{W}tbl"]
    patched = 0
    for table, spec in zip(tables, specs):
        rows = table.findall("w:tr", NS)
        for row_index, data in enumerate(spec.get("rows") or []):
            if row_index >= len(rows):
                break
            cells = rows[row_index].findall("w:tc", NS)
            for column_index, value in enumerate(data):
                if column_index < len(cells):
                    _set_cell_text(cells[column_index], value)
        patched += 1
    return patched


def apply_for_ticker(
    ticker: str,
    template: Path,
    texts_path: Path,
    out: Path,
    work: Path,
) -> dict:
    """Apply only the configured company's operations/finance slots."""
    from lxml import etree

    route = select_apply_route(ticker)
    texts = load_texts(texts_path)
    by_id = {
        str(item["id"]): str(item.get("text") or "")
        for item in texts["paragraphs"]
    }
    if work.exists():
        shutil.rmtree(work)
    work.mkdir(parents=True)
    with ZipFile(template) as archive:
        archive.extractall(work)
    document_xml = work / "word" / "document.xml"
    root = etree.parse(str(document_xml)).getroot()
    body = root.find("w:body", NS)
    children = list(body)
    start = next(
        (
            index
            for index, element in enumerate(children)
            if element.tag == f"{W}p"
            and any(
                pattern.search(_paragraph_text(element).strip())
                for pattern in route.start_patterns
            )
        ),
        None,
    )
    if start is None:
        raise ValueError(
            f"{ticker} 模板找不到写作段起点；请提供该公司上季模板。"
        )
    used: set[str] = set()
    for element in children[start:]:
        if element.tag == f"{W}sectPr":
            break
        if element.tag != f"{W}p":
            continue
        if element.findall(f".//{A_BLIP}") and not _paragraph_text(element).strip():
            continue
        raw = _paragraph_text(element).strip()
        if not raw:
            continue
        for slot_id, pattern in route.anchors:
            if slot_id in used or slot_id not in by_id:
                continue
            if pattern.search(raw):
                _set_paragraph_text(element, by_id[slot_id])
                used.add(slot_id)
                break
    tables = texts.get("tables") or []
    table_count = 0
    if route.table_mode == "expe":
        table_count = _patch_expe_tables(children[start:], tables)
    elif route.table_mode == "sequential":
        table_count = _patch_sequential_tables(children[start:], tables)
    missing = [
        slot_id
        for slot_id, value in by_id.items()
        if value.strip() and slot_id not in used
    ]
    if missing:
        raise ValueError(
            f"{ticker} 模板未匹配这些 texts 槽位：{', '.join(missing)}"
        )
    document_xml.write_bytes(
        etree.tostring(
            root,
            xml_declaration=True,
            encoding="UTF-8",
            standalone=True,
        )
    )
    out.parent.mkdir(parents=True, exist_ok=True)
    out.unlink(missing_ok=True)
    with ZipFile(out, "w", ZIP_DEFLATED) as archive:
        for file in work.rglob("*"):
            if file.is_file():
                archive.write(file, file.relative_to(work).as_posix())
    return {
        "ticker": ticker.upper(),
        "route": route.name,
        "matched": len(used),
        "tables": table_count,
        "out": str(out),
    }


def _docx_paragraphs(docx: Path) -> list[str]:
    from lxml import etree

    with ZipFile(docx) as archive:
        root = etree.fromstring(archive.read("word/document.xml"))
    body = root.find("w:body", NS)
    return [
        _paragraph_text(paragraph).strip()
        for paragraph in body.findall("w:p", NS)
        if _paragraph_text(paragraph).strip()
    ]


def _image_sizes(docx: Path) -> dict[str, int]:
    with ZipFile(docx) as archive:
        return {
            name.removeprefix("word/"): len(archive.read(name))
            for name in archive.namelist()
            if name.startswith("word/media/")
        }


def accept_docx(
    ticker: str,
    docx: Path,
    template: Path,
    texts_path: Path,
    quarter: str,
    chart_mapping: dict[str, str],
) -> list[str]:
    """Post-apply acceptance gate, including a ticker-specific re-apply smoke."""
    route = select_apply_route(ticker)
    paragraphs = _docx_paragraphs(docx)
    start = next(
        (
            index
            for index, text in enumerate(paragraphs)
            if any(pattern.search(text) for pattern in route.start_patterns)
        ),
        None,
    )
    if start is None:
        return [f"{ticker} operations/finance section not found"]
    body = "\n".join(paragraphs[start:])
    short = quarter[2:]
    if short not in body and f"Q{quarter[-1]}" not in body:
        errors = [f"finance block missing quarter label {short}"]
    else:
        errors = []
    image_sizes = _image_sizes(docx)
    template_sizes = _image_sizes(template)
    unchanged = []
    for key in chart_mapping:
        normalized = key.replace("\\", "/").removeprefix("word/")
        if not normalized.startswith("media/"):
            normalized = f"media/{Path(normalized).name}"
        size = image_sizes.get(normalized, 0)
        if size < 2000:
            errors.append(f"chart slot {normalized} missing or <2KB")
        if size and size == template_sizes.get(normalized):
            unchanged.append(normalized)
    if chart_mapping and len(unchanged) == len(chart_mapping):
        errors.append("all mapped chart images are unchanged from template")

    with tempfile.TemporaryDirectory() as temporary:
        smoke = Path(temporary) / "smoke.docx"
        work = Path(temporary) / "work"
        try:
            apply_for_ticker(
                ticker, template, texts_path, smoke, work
            )
        except Exception as error:  # noqa: BLE001
            errors.append(f"re-apply smoke failed: {error}")
        else:
            smoke_body = "\n".join(_docx_paragraphs(smoke))
            texts = load_texts(texts_path)
            missing = []
            for paragraph in texts["paragraphs"]:
                text = str(paragraph.get("text") or "").strip()
                key = text[: min(12, len(text))]
                if key and key not in smoke_body:
                    missing.append(str(paragraph.get("id")))
            if missing:
                errors.append(
                    "re-apply smoke missing slots: " + ", ".join(missing[:8])
                )
    return errors
