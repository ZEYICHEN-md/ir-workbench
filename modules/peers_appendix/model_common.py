"""Shared helpers for peers-model Excel operations.

Migrated from the repository-authoritative earnings-summary implementation.
Sheets may store either ``2026Q2`` or the short ``26Q2`` label; the canonical
in-memory form is always ``20YYQn``.
"""

from __future__ import annotations

import re
from collections import Counter
from pathlib import Path

Q_RE = re.compile(r"^(20\d{2})Q([1-4])$")
Q_SHORT_RE = re.compile(r"^(\d{2})Q([1-4])$", re.IGNORECASE)
YEAR_RE = re.compile(r"^(20\d{2})$")
FY_RE = re.compile(r"^FY\s*(20\d{2})", re.IGNORECASE)

XL_PASTE_FORMATS = -4122
XL_CALC_MANUAL = -4135
XL_TO_RIGHT = -4161
INPUT_BLUE_BGR = 16711680


class ModelLayoutError(ValueError):
    pass


def normalize_quarter(label: str) -> str | None:
    value = str(label).strip()
    match = Q_RE.match(value)
    if match:
        return f"{match.group(1)}Q{match.group(2)}"
    match = Q_SHORT_RE.match(value)
    if match:
        return f"20{match.group(1)}Q{match.group(2)}"
    return None


def format_quarter_label(canonical: str, style: str) -> str:
    match = Q_RE.match(canonical)
    if not match:
        raise ValueError(f"季度格式不对：{canonical}")
    if style == "short":
        return f"{match.group(1)[2:]}Q{match.group(2)}"
    return canonical


def next_quarter(quarter: str) -> str:
    canonical = normalize_quarter(quarter) or quarter
    match = Q_RE.match(canonical)
    if not match:
        raise ValueError(f"季度格式不对：{quarter}")
    year, number = int(match.group(1)), int(match.group(2))
    return f"{year + 1}Q1" if number == 4 else f"{year}Q{number + 1}"


def col_letter(number: int) -> str:
    out = ""
    while number:
        number, remainder = divmod(number - 1, 26)
        out = chr(65 + remainder) + out
    return out


def col_num(letters: str) -> int:
    number = 0
    for char in letters.upper():
        if not ("A" <= char <= "Z"):
            raise ValueError(letters)
        number = number * 26 + (ord(char) - 64)
    return number


def is_year_label(value) -> bool:
    if value is None:
        return False
    if (
        isinstance(value, (int, float))
        and 2000 <= float(value) <= 2100
        and float(value) == int(value)
    ):
        return True
    text = str(value).strip()
    return bool(YEAR_RE.match(text) or FY_RE.match(text))


def detect_layout(model: Path, sheet: str) -> dict:
    from openpyxl import load_workbook

    workbook = load_workbook(model, data_only=False, read_only=True)
    try:
        if sheet not in workbook.sheetnames:
            raise ModelLayoutError(f"Model 找不到工作表 {sheet!r}。")
        worksheet = workbook[sheet]
        hits: list[tuple[int, int, str, str]] = []
        for row in range(1, 8):
            for column in range(1, (worksheet.max_column or 1) + 1):
                value = worksheet.cell(row, column).value
                if value is None:
                    continue
                raw = str(value).strip()
                canonical = normalize_quarter(raw)
                if canonical:
                    hits.append((row, column, canonical, raw))
        if not hits:
            raise ModelLayoutError(f"{sheet} 前 7 行没有季度标签。")
        label_row = Counter(row for row, _, _, _ in hits).most_common(1)[0][0]
        row_hits = [
            (column, canonical, raw)
            for row, column, canonical, raw in hits
            if row == label_row
        ]
        quarter_map = {canonical: column for column, canonical, _ in row_hits}
        short_count = sum(
            1
            for _, _, raw in row_hits
            if Q_SHORT_RE.match(raw) and not Q_RE.match(raw)
        )
        label_style = (
            "short"
            if short_count >= max(1, len(row_hits) // 2)
            else "full"
        )
        last_quarter = max(quarter_map, key=quarter_map.get)
        last_column = quarter_map[last_quarter]
        annual_column = None
        for column in range(
            last_column + 1, (worksheet.max_column or last_column) + 1
        ):
            value = worksheet.cell(label_row, column).value
            if value is None or not str(value).strip():
                continue
            if is_year_label(value):
                annual_column = column
                break
        return {
            "label_row": label_row,
            "quarters": quarter_map,
            "last_quarter": last_quarter,
            "last_col": last_column,
            "annual_col": annual_column,
            "gap_cols": (
                annual_column - last_column - 1 if annual_column else None
            ),
            "label_style": label_style,
        }
    finally:
        workbook.close()


def chart_sheet_name(data_sheet: str) -> str:
    return f"{data_sheet} Quarterly Charts"


def prev_quarter_col(layout: dict, quarter: str) -> int | None:
    quarter_map = layout.get("quarters") or {}
    if quarter not in quarter_map:
        return None
    target = quarter_map[quarter]
    previous = None
    for _, column in sorted(quarter_map.items(), key=lambda item: item[1]):
        if column == target:
            break
        previous = column
    return previous


def openpyxl_font_key(cell) -> tuple:
    font = cell.font
    color = font.color
    normalized = None
    if color is not None:
        if color.type == "rgb" and color.rgb:
            normalized = str(color.rgb).upper()
        elif color.type == "theme" and color.theme is not None:
            normalized = ("theme", int(color.theme), float(color.tint or 0))
    if normalized in (None, "FF000000"):
        normalized = None
    return (font.name, font.sz, bool(font.bold), bool(font.italic), normalized)


def required_gap_before_annual(_sheet: str) -> int:
    return 1
