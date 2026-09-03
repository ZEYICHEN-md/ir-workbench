"""Must-pass Model audit: write verification plus adversarial cross-checks.

This is the repository-authoritative gate refactored into an importable module.
Quarter-specific ABNB constants were intentionally removed: disclosed values
belong in ``ir_snapshot.json`` and are compared there for every quarter.
"""

from __future__ import annotations

import json
import re
import time
from pathlib import Path

from .ir_snapshot import check_fill_vs_snapshot, check_model_vs_snapshot
from .model_common import (
    detect_layout,
    openpyxl_font_key,
    prev_quarter_col,
    required_gap_before_annual,
)


def _num(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip().replace(",", ""))
    except ValueError:
        return None


def _eval_simple(value):
    if not isinstance(value, str) or not value.startswith("="):
        return _num(value)
    body = value[1:].replace(" ", "")
    if not re.fullmatch(
        r"-?(?:\(\d+(?:\.\d+)?(?:[+\-*/]\d+(?:\.\d+)?)+\)"
        r"|\d+(?:\.\d+)?(?:[+\-*/]\d+(?:\.\d+)?)+)",
        body,
    ):
        return None
    try:
        return float(eval(body, {"__builtins__": {}}, {}))  # noqa: S307
    except Exception:
        return None


def _near(left, right, tolerance: float) -> bool:
    return (
        left is not None
        and right is not None
        and abs(float(left) - float(right)) <= tolerance
    )


def _finding(
    ok: bool,
    name: str,
    *,
    severity: str = "FAIL",
    kind: str = "adversarial",
    expected=None,
    got=None,
    note: str | None = None,
) -> dict:
    return {
        "severity": "OK" if ok else severity,
        "kind": kind,
        "name": name,
        "expected": expected,
        "got": got,
        "note": note,
    }


AUDIT_ROWS = {
    "BKNG": {6, 14, 17, 30},
    "EXPE": {
        5, 6, 33, 34, 37, 38, 59, 60, 66, 71, 113, 117, 121, 125,
        130, 134, 138, 142, 147, 148, 149, 150, 155, 180, 182, 215,
        307, 308,
    },
    "ABNB": {
        3, 4, 6, 7, 9, 11, 15, 16, 17, 18, 22, 23, 24, 25, 27,
        29, 31, 32, 33, 35, 36, 37, 40, 41, 42, 43, 45, 55, 56,
        59, 60, 61, 62, 99, 116,
    },
}


def read_model_values(
    model: Path, sheet: str, column: int, rows: list[int]
) -> dict[int, object]:
    """Read calculated values through Excel COM; openpyxl cannot recalculate."""
    import pythoncom
    import win32com.client as win32

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    values: dict[int, object] = {}
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(
            str(model.resolve()), UpdateLinks=0, ReadOnly=True
        )
        try:
            excel.CalculateFull()
        except Exception:
            pass
        time.sleep(0.3)
        worksheet = workbook.Worksheets(sheet)
        for row in rows:
            values[row] = worksheet.Cells(row, column).Value
        return values
    finally:
        try:
            if workbook is not None:
                workbook.Close(False)
        except Exception:
            pass
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


def check_fill_landed(fill: dict, values: dict[int, object]) -> list[dict]:
    findings = []
    for item in fill.get("inputs", []):
        if item.get("value") is None:
            continue
        row = int(item["row"])
        raw = item["value"]
        got_raw = values.get(row)
        expected = _eval_simple(raw)
        if expected is None and isinstance(raw, str) and raw.startswith("="):
            findings.append(
                _finding(
                    got_raw not in (None, ""),
                    f"write_verify_formula_row_{row}",
                    severity="WARN",
                    kind="write_verify",
                    expected=raw,
                    got=got_raw,
                    note="复杂公式无法静态求值，只核对非空。",
                )
            )
        elif expected is None:
            findings.append(
                _finding(
                    str(got_raw).strip() == str(raw).strip(),
                    f"write_verify_row_{row}",
                    kind="write_verify",
                    expected=raw,
                    got=got_raw,
                )
            )
        else:
            got = _num(got_raw)
            findings.append(
                _finding(
                    _near(got, expected, 0.51),
                    f"write_verify_row_{row}",
                    kind="write_verify",
                    expected=expected,
                    got=got,
                )
            )
    return findings


def check_layout_and_format(
    model: Path, sheet: str, quarter: str, fill: dict, layout: dict
) -> list[dict]:
    from openpyxl import load_workbook

    findings = []
    gap = layout.get("gap_cols")
    if layout.get("annual_col") is not None:
        needed = required_gap_before_annual(sheet)
        findings.append(
            _finding(
                (gap or 0) >= needed,
                "gap_before_annual",
                kind="layout",
                expected=f">={needed}",
                got=gap,
            )
        )
    workbook = load_workbook(model, data_only=False, read_only=True)
    try:
        worksheet = workbook[sheet]
        column = layout["quarters"][quarter]
        raw_header = worksheet.cell(layout["label_row"], column).value
        style = layout["label_style"]
        header_ok = (
            style == "short"
            and bool(re.fullmatch(r"\d{2}Q[1-4]", str(raw_header)))
        ) or (
            style == "full"
            and bool(re.fullmatch(r"20\d{2}Q[1-4]", str(raw_header)))
        )
        findings.append(
            _finding(
                header_ok,
                "quarter_header_style",
                kind="layout",
                expected=style,
                got=raw_header,
            )
        )
        previous = prev_quarter_col(layout, quarter)
        if previous is not None:
            for item in fill.get("inputs", []):
                if item.get("value") is None:
                    continue
                row = int(item["row"])
                target_font = openpyxl_font_key(worksheet.cell(row, column))
                prior_font = openpyxl_font_key(worksheet.cell(row, previous))
                findings.append(
                    _finding(
                        target_font == prior_font,
                        f"font_row_{row}",
                        kind="format",
                        expected=prior_font,
                        got=target_font,
                    )
                )
    finally:
        workbook.close()
    return findings


def _adversarial_bkng(values: dict[int, object]) -> list[dict]:
    value = {row: _num(raw) for row, raw in values.items()}
    gbv, agency, merchant = value.get(6), value.get(14), value.get(17)
    revenue = value.get(30)
    return [
        _finding(
            None not in (gbv, agency, merchant)
            and _near(agency + merchant, gbv, 1.0),
            "agency+merchant=gbv",
            expected=gbv,
            got=None if agency is None or merchant is None else agency + merchant,
        ),
        _finding(
            revenue is not None,
            "revenue_present",
            expected="non-empty",
            got=revenue,
        ),
    ]


def _adversarial_expe(values: dict[int, object]) -> list[dict]:
    value = {row: _num(raw) for row, raw in values.items()}
    gbv, agency, merchant = value.get(5), value.get(33), value.get(37)
    revenue = value.get(59)
    us_revenue, international_revenue = value.get(66), value.get(71)
    products = [value.get(row) for row in (113, 117, 121, 125)]
    segments = [value.get(row) for row in (130, 134, 138)]
    ebitda, margin = value.get(180), value.get(182)
    findings = [
        _finding(
            None not in (gbv, agency, merchant)
            and _near(agency + merchant, gbv, 1.0),
            "agency+merchant=gbv",
            expected=gbv,
            got=None if agency is None or merchant is None else agency + merchant,
        ),
        _finding(
            None not in (revenue, us_revenue, international_revenue)
            and _near(us_revenue + international_revenue, revenue, 1.0),
            "us+international=revenue",
            expected=revenue,
            got=(
                None
                if us_revenue is None or international_revenue is None
                else us_revenue + international_revenue
            ),
        ),
        _finding(
            all(item is not None for item in products)
            and revenue is not None
            and _near(sum(products), revenue, 1.0),
            "product_sum=revenue",
            expected=revenue,
            got=sum(products) if all(x is not None for x in products) else None,
        ),
        _finding(
            all(item is not None for item in segments)
            and revenue is not None
            and _near(sum(segments), revenue, 1.0),
            "segment_sum=revenue",
            expected=revenue,
            got=sum(segments) if all(x is not None for x in segments) else None,
        ),
        _finding(
            None not in (ebitda, margin, revenue)
            and revenue != 0
            and _near(margin, ebitda / revenue, 0.002),
            "ebitda_margin=ebitda/revenue",
            expected=(
                None
                if revenue in (None, 0) or ebitda is None
                else ebitda / revenue
            ),
            got=margin,
        ),
    ]
    for row, name in (
        (142, "cost_of_sales_sign"),
        (147, "technology_sign"),
        (148, "sales_marketing_sign"),
        (149, "general_admin_sign"),
        (150, "depreciation_sign"),
        (155, "restructuring_sign"),
    ):
        findings.append(
            _finding(
                value.get(row) is not None and value[row] < 0,
                name,
                expected="<0",
                got=value.get(row),
            )
        )
    return findings


def _adversarial_abnb(values: dict[int, object]) -> list[dict]:
    value = {row: _num(raw) for row, raw in values.items()}
    nights, gbv, adr = value.get(3), value.get(6), value.get(11)
    revenue, cost, gross_profit = value.get(15), value.get(17), value.get(18)
    expenses = [value.get(row) for row in (22, 23, 24, 25)]
    total_expenses, operating_income = value.get(27), value.get(29)
    pretax, tax, net_income = value.get(35), value.get(36), value.get(37)
    ebitda, margin, take_rate = value.get(55), value.get(56), value.get(9)
    findings = [
        _finding(nights is not None, "nights_present", got=nights),
        _finding(gbv is not None, "gbv_present", got=gbv),
        _finding(revenue is not None, "revenue_present", got=revenue),
        _finding(ebitda is not None, "ebitda_present", got=ebitda),
        _finding(
            None not in (gbv, nights, adr)
            and nights != 0
            and _near(adr, gbv / nights, 1.0),
            "adr≈gbv/nights",
            expected=(
                None if nights in (None, 0) or gbv is None else gbv / nights
            ),
            got=adr,
        ),
        _finding(
            None not in (revenue, gbv, take_rate)
            and gbv != 0
            and _near(take_rate, revenue / gbv, 0.002),
            "take_rate=revenue/gbv",
            expected=(
                None
                if gbv in (None, 0) or revenue is None
                else revenue / gbv
            ),
            got=take_rate,
        ),
        _finding(
            None not in (revenue, cost, gross_profit)
            and _near(gross_profit, revenue - cost, 1.0),
            "gross_profit=revenue-cost",
            expected=None if revenue is None or cost is None else revenue - cost,
            got=gross_profit,
        ),
        _finding(
            all(item is not None for item in expenses)
            and total_expenses is not None
            and _near(total_expenses, sum(expenses), 1.0),
            "opex_sum",
            expected=sum(expenses) if all(x is not None for x in expenses) else None,
            got=total_expenses,
        ),
        _finding(
            None not in (gross_profit, total_expenses, operating_income)
            and _near(operating_income, gross_profit - total_expenses, 1.0),
            "operating_income=gross_profit-opex",
            expected=(
                None
                if gross_profit is None or total_expenses is None
                else gross_profit - total_expenses
            ),
            got=operating_income,
        ),
        _finding(
            None not in (pretax, tax, net_income)
            and _near(net_income, pretax - tax, 1.0),
            "net_income=pretax-tax",
            expected=None if pretax is None or tax is None else pretax - tax,
            got=net_income,
        ),
        _finding(
            None not in (ebitda, margin, revenue)
            and revenue != 0
            and _near(margin, ebitda / revenue, 0.002),
            "ebitda_margin=ebitda/revenue",
            expected=(
                None
                if revenue in (None, 0) or ebitda is None
                else ebitda / revenue
            ),
            got=margin,
        ),
        _finding(
            value.get(32) is not None and value[32] != 0,
            "interest_expense_present",
            expected="non-zero",
            got=value.get(32),
        ),
    ]
    sbc_parts = [value.get(row) for row in (40, 41, 42, 43)]
    if all(part is not None for part in sbc_parts):
        findings.append(
            _finding(
                value.get(45) is not None
                and _near(value[45], sum(sbc_parts), 1.0),
                "sbc_total=parts",
                expected=sum(sbc_parts),
                got=value.get(45),
            )
        )
    for gaap_row, sbc_row, non_gaap_row, name in (
        (22, 40, 59, "non_gaap_ops=gaap-sbc"),
        (23, 41, 60, "non_gaap_pd=gaap-sbc"),
        (24, 42, 61, "non_gaap_sm=gaap-sbc"),
        (25, 43, 62, "non_gaap_ga=gaap-sbc"),
    ):
        gaap, sbc, non_gaap = (
            value.get(gaap_row),
            value.get(sbc_row),
            value.get(non_gaap_row),
        )
        findings.append(
            _finding(
                None not in (gaap, sbc, non_gaap)
                and _near(non_gaap, gaap - sbc, 1.0),
                name,
                expected=None if gaap is None or sbc is None else gaap - sbc,
                got=non_gaap,
            )
        )
    return findings


ADVERSARIAL = {
    "BKNG": _adversarial_bkng,
    "EXPE": _adversarial_expe,
    "ABNB": _adversarial_abnb,
}


def run_audit(
    model: Path,
    sheet: str,
    quarter: str,
    fill_path: Path,
    snapshot: dict,
) -> dict:
    sheet = sheet.upper()
    fill = json.loads(fill_path.read_text(encoding="utf-8"))
    if str(fill.get("sheet", "")).upper() != sheet:
        raise ValueError(f"fill.sheet={fill.get('sheet')}，应为 {sheet}。")
    if fill.get("quarter") != quarter:
        raise ValueError(f"fill.quarter={fill.get('quarter')}，应为 {quarter}。")
    if not fill.get("inputs"):
        raise ValueError("fill_inputs.json 的 inputs 不能为空。")

    layout = detect_layout(model, sheet)
    if quarter not in layout["quarters"]:
        raise ValueError(f"{sheet} 缺 {quarter}。")
    rows = {int(item["row"]) for item in fill["inputs"]}
    rows |= AUDIT_ROWS.get(sheet, set())
    for meta in snapshot.get("actuals", {}).values():
        if not isinstance(meta, dict):
            continue
        if meta.get("model_row") is not None:
            rows.add(int(meta["model_row"]))
        if meta.get("yoy_model_row") is not None:
            rows.add(int(meta["yoy_model_row"]))
    values = read_model_values(
        model, sheet, layout["quarters"][quarter], sorted(rows)
    )

    findings = check_layout_and_format(model, sheet, quarter, fill, layout)
    findings.extend(check_fill_landed(fill, values))
    findings.extend(check_fill_vs_snapshot(fill, snapshot))
    findings.extend(check_model_vs_snapshot(values, snapshot))
    adversarial = ADVERSARIAL.get(sheet)
    if adversarial is None:
        findings.append(
            _finding(
                False,
                "no_ticker_adversarial_suite",
                severity="FAIL",
                note=f"{sheet} 尚无对抗勾稽，不能宣称 Model 通过。",
            )
        )
    else:
        findings.extend(adversarial(values))

    failures = [item for item in findings if item["severity"] == "FAIL"]
    warnings = [item for item in findings if item["severity"] == "WARN"]
    oks = [item for item in findings if item["severity"] == "OK"]
    # The orchestrated gate is always strict: WARN is also a hard stop.
    passed = not failures and not warnings
    return {
        "model": str(model),
        "sheet": sheet,
        "quarter": quarter,
        "fill": str(fill_path),
        "strict": True,
        "summary": {
            "ok": len(oks),
            "warn": len(warnings),
            "fail": len(failures),
            "passed": passed,
        },
        "findings": findings,
    }


def write_report(report: dict, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(report, ensure_ascii=False, indent=2, default=str) + "\n",
        encoding="utf-8",
        newline="\n",
    )
    return path
