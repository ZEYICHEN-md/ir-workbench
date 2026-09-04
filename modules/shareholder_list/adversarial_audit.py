"""Sheet-by-sheet adversarial audit of the generated Investor List."""
from __future__ import annotations

import json
import time
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .build import (
    ADR_VALUE_TICKERS,
    AUM_FLOOR,
    COMBINED_KEEP,
    EXCLUDE_TYPES,
    HOLDINGS_AS_OF,
    MAP_D8,
    REGION_PCT_COLS,
    STRATEGIC,
    TRAVEL_VALUE_TICKERS,
    VALID_AS_OF,
    WORKBOOK_AS_OF,
    header_row_of,
    load_mapping,
    num,
    output_filename,
    period_key,
    read_table,
    sheet_region,
    usd_mcap,
)
from .discover import MARKET_CAPS, default_combined, default_peer, default_template
from workbench.paths import Paths, find_root

DOMAIN = "shareholder-list"


def _layout():
    paths = Paths(find_root())
    out_dir = paths.outputs(DOMAIN, period_key())
    return paths, out_dir


def default_output_path() -> Path:
    _, out_dir = _layout()
    return out_dir / output_filename()

REGION_ROW5 = {
    4: "Style",
    5: "Equity Assets USD, mm",
    6: "Country/Territory",
    7: "Value TCOM USD, mm",
    8: "Value 9961_XHKG USD, mm",
    12: "Value 700_XHKG USD, mm",
    14: "Value BABA USD, mm",
    15: "Value 9988_XHKG USD, mm",
    18: "Value 3690_XHKG USD, mm",
    20: "Value NTES USD, mm",
    21: "Value 9999_XHKG USD, mm",
    24: "Value PDD USD, mm",
    26: "Value JD USD, mm",
    27: "Value 9618_XHKG USD, mm",
    30: "Value BIDU USD, mm",
    31: "Value 9888_XHKG USD, mm",
    34: "Value ABNB USD, mm",
    36: "Value BKNG USD, mm",
    38: "Value EXPE USD, mm",
    40: "Value MMYT USD, mm",
    42: "Value TRIP USD, mm",
    44: "Value TRVG USD, mm",
    46: "Value 780_XHKG USD, mm",
}
MCAP_USD_CELLS = {
    10: "TCOM",
    17: "BABA",
    23: "NTES",
    25: "PDD",
    29: "JD",
    33: "BIDU",
    35: "ABNB",
    37: "BKNG",
    39: "EXPE",
    41: "MMYT",
    43: "TRIP",
    45: "TRVG",
}
MCAP_HKD_CELLS = {13: "700_XHKG", 19: "3690_XHKG", 47: "780_XHKG"}


def names_col(ws, start: int, col: int, stop=frozenset({"TOTAL"})) -> list[str]:
    out = []
    r = start
    while True:
        v = ws.cell(r, col).value
        if not isinstance(v, str) or not v.strip() or v.strip() in stop:
            break
        out.append(v)
        r += 1
    return out


def cell_eq(a, b) -> bool:
    if a is None and b is None:
        return True
    if a is None or b is None:
        return a in (None, "") and b in (None, "")
    if isinstance(a, float) and isinstance(b, float):
        return abs(a - b) < 1e-6
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        return abs(float(a) - float(b)) < 1e-6
    return a == b


def run_audit(
    output: Path | None = None,
    peer: Path | None = None,
    combined: Path | None = None,
    template: Path | None = None,
    market_path: Path | None = None,
    report_path: Path | None = None,
) -> dict[str, Any]:
    findings: list[dict[str, Any]] = []
    sheets: dict[str, Any] = {}

    _, out_dir = _layout()
    OUT = Path(output) if output else out_dir / output_filename()
    PEER = Path(peer) if peer else default_peer()
    COMB = Path(combined) if combined else default_combined()
    TMPL = Path(template) if template else default_template()
    market_path = Path(market_path) if market_path else MARKET_CAPS
    REPORT = Path(report_path) if report_path else out_dir / "adversarial_audit.json"
    MARKET = json.loads(market_path.read_text(encoding="utf-8"))
    if PEER is None or COMB is None or TMPL is None:
        raise FileNotFoundError("need peer, combined, and template paths for audit")

    def fail(sheet: str, severity: str, where: str, msg: str) -> None:
        findings.append({"sheet": sheet, "severity": severity, "where": where, "msg": msg})

    ow = None
    last_err: Exception | None = None
    for attempt in range(5):
        try:
            ow = load_workbook(OUT, data_only=False)
            break
        except Exception as exc:
            last_err = exc
            time.sleep(0.4 * (attempt + 1))
    if ow is None:
        raise last_err
    tw = load_workbook(TMPL, data_only=False)
    pwb = load_workbook(PEER)
    cwb = load_workbook(COMB)
    ph, pr = read_table(pwb.active, header_row_of(pwb.active))
    ch, cr = read_table(cwb.active, header_row_of(cwb.active))
    pmap = {h: i for i, h in enumerate(ph)}
    cmap = {h: i for i, h in enumerate(ch)}
    pwb.close()
    cwb.close()

    expected_sheets = [
        "Top 20", "SH Summary", "SH Prior", "Data_TCOM+9961", "全球投资人地图",
        "US&CA", "EU", "APAC", "ROW", "DATA_ALL", "重要投资人", "普通投资人", "Mapping", "Market",
    ]
    if ow.sheetnames != expected_sheets:
        fail("workbook", "high", "sheet order", f"{ow.sheetnames} != {expected_sheets}")

    # ---- DATA_ALL: every header + every data cell vs peer ----
    da = ow["DATA_ALL"]
    da_headers = []
    for c in range(1, 100):
        h = da.cell(2, c).value
        da_headers.append(h)
    if da_headers[: len(ph)] != ph:
        fail("DATA_ALL", "high", "R2", "headers != peer extract")
    if da.cell(2, len(ph) + 1).value != "Fund Size" or da.cell(2, len(ph) + 2).value != "Region":
        fail("DATA_ALL", "high", "R2", "missing Fund Size/Region")
    da_names = names_col(da, 3, 1)
    if da_names != [r[0] for r in pr]:
        fail("DATA_ALL", "high", "names", f"{len(da_names)} vs peer {len(pr)}")
    mismatches = 0
    examples = []
    for i, row in enumerate(pr):
        r = 3 + i
        for c, val in enumerate(row, 1):
            got = da.cell(r, c).value
            if not cell_eq(got, val):
                mismatches += 1
                if len(examples) < 8:
                    examples.append(f"R{r}{get_column_letter(c)} peer={val!r} got={got!r}")
    if mismatches:
        fail("DATA_ALL", "high", "body", f"{mismatches} cells differ e.g. {examples}")
    sum_row = 3 + len(pr)
    if da.cell(sum_row, 1).value != "TOTAL":
        fail("DATA_ALL", "high", f"A{sum_row}", "TOTAL missing")
    if da.cell(sum_row + 1, 1).value not in (None, ""):
        fail("DATA_ALL", "high", f"A{sum_row+1}", f"leftover {da.cell(sum_row+1,1).value}")
    if da.freeze_panes != "B3":
        fail("DATA_ALL", "med", "freeze", str(da.freeze_panes))
    n = len(ph)
    fund_f = str(da.cell(3, n + 1).value or "")
    if f"{get_column_letter(ph.index('Equity Assets USD, mm')+1)}3" not in fund_f:
        fail("DATA_ALL", "high", "CT3", fund_f)
    region_f = str(da.cell(3, n + 2).value or "")
    if f"{get_column_letter(ph.index('Country/Territory')+1)}3" not in region_f:
        fail("DATA_ALL", "high", "CU3", region_f)
    last = 2 + len(pr)
    for c, h in enumerate(ph, 1):
        v = da.cell(sum_row, c).value
        letter = get_column_letter(c)
        should_sum = isinstance(h, str) and (h.startswith("Value ") or h in ("Aggregate Value USD, mm", "Equity Assets USD, mm"))
        if should_sum:
            if v != f"=SUM({letter}3:{letter}{last})":
                fail("DATA_ALL", "high", f"{letter}{sum_row}", f"bad SUM {v}")
        elif c > 1 and v not in (None, ""):
            fail("DATA_ALL", "med", f"{letter}{sum_row}", f"unexpected {v}")
    # last data row formulas exist
    if not str(da.cell(last, n + 1).value or "").startswith("=IF("):
        fail("DATA_ALL", "high", f"CT{last}", "Fund Size missing on last inst")
    sheets["DATA_ALL"] = {"rows": len(da_names), "cols": n, "sum_row": sum_row, "mismatches": mismatches}

    # ---- Combined ----
    cb = ow["Data_TCOM+9961"]
    for c, h in enumerate(COMBINED_KEEP, 1):
        if cb.cell(3, c).value != h:
            fail("Data_TCOM+9961", "high", f"R3C{c}", f"{cb.cell(3,c).value} != {h}")
            break
    cb_names = names_col(cb, 4, 1)
    comb_names = [r[0] for r in cr]
    if cb_names != comb_names:
        fail("Data_TCOM+9961", "high", "names", f"{len(cb_names)} vs {len(comb_names)}")
    id_fail = 0
    body_fail = 0
    for i, row in enumerate(cr):
        r = 4 + i
        for h in COMBINED_KEEP:
            c = COMBINED_KEEP.index(h) + 1
            src = row[cmap[h]]
            if not cell_eq(cb.cell(r, c).value, src):
                body_fail += 1
                if body_fail <= 5:
                    fail("Data_TCOM+9961", "high", f"{get_column_letter(c)}{r}", f"{h} {cb.cell(r,c).value!r} != {src!r}")
        tot = num(row[cmap["Shares (Combined)"]])
        tcom = num(row[cmap["TCOM Shares"]])
        hk = num(row[cmap["9961_XHKG Shares"]])
        if abs(tot - tcom - hk) > 1:
            id_fail += 1
    if id_fail:
        fail("Data_TCOM+9961", "high", "shares", f"{id_fail} identity fails")
    last_cb = 3 + len(cr)
    if cb.cell(last_cb + 1, 1).value not in (None, ""):
        fail("Data_TCOM+9961", "high", f"A{last_cb+1}", "leftover name")
    if "M4" not in str(cb.cell(4, 19).value):
        fail("Data_TCOM+9961", "high", "S4", str(cb.cell(4, 19).value))
    if "R4" not in str(cb.cell(4, 20).value):
        fail("Data_TCOM+9961", "high", "T4", str(cb.cell(4, 20).value))
    if "N4" not in str(cb.cell(4, 21).value):
        fail("Data_TCOM+9961", "high", "U4", str(cb.cell(4, 21).value))
    missing_peer = [n for n in comb_names if n not in {x[0] for x in pr}]
    if missing_peer:
        fail("Data_TCOM+9961", "med", "peer coverage", f"{len(missing_peer)} not in peer e.g. {missing_peer[:5]}")
    sheets["Data_TCOM+9961"] = {"rows": len(cb_names), "identity_fail": id_fail, "body_fail": body_fail}

    # ---- SH Summary ----
    sh = ow["SH Summary"]
    sh_names = names_col(sh, 6, 3)
    expected = [n for n in comb_names if n not in STRATEGIC]
    if sh_names != expected:
        fail("SH Summary", "high", "C", "order/name mismatch vs combined-strategic")
    if any(n in STRATEGIC for n in sh_names):
        fail("SH Summary", "high", "C", "Baidu leaked")
    if sh["F3"].value != MARKET["tcom_shares_outstanding"]:
        fail("SH Summary", "high", "F3", str(sh["F3"].value))
    want_g3 = f"updated as of {VALID_AS_OF}"
    if str(sh["G3"].value) != want_g3 or "TBU" in str(sh["G3"].value):
        fail("SH Summary", "high", "G3", str(sh["G3"].value))
    idx_expect = {"D": 7, "E": 10, "M": 9, "N": 12, "O": 14, "P": 16, "Q": 17, "R": 18, "S": 13}
    for col, val in idx_expect.items():
        if sh.cell(4, ord(col) - 64).value != val:
            fail("SH Summary", "high", f"{col}4", f"{sh.cell(4, ord(col)-64).value} != {val}")
    last_sh = 5 + len(sh_names)
    if sh.cell(last_sh + 1, 3).value not in (None, ""):
        fail("SH Summary", "high", f"C{last_sh+1}", "leftover name")
    # every data row formula pattern
    for i, name in enumerate(sh_names):
        r = 6 + i
        if sh.cell(r, 2).value != i + 1:
            fail("SH Summary", "high", f"B{r}", "rank")
            break
        if sh.cell(r, 3).value != name:
            fail("SH Summary", "high", f"C{r}", "name")
            break
        a = str(sh.cell(r, 1).value or "")
        if "SH Prior" not in a or "[1]" in a:
            fail("SH Summary", "high", f"A{r}", a)
            break
        if f"D{r}" not in str(sh.cell(r, 6).value):
            fail("SH Summary", "high", f"F{r}", str(sh.cell(r, 6).value))
            break
        if f"$F$3" not in str(sh.cell(r, 12).value):
            fail("SH Summary", "high", f"L{r}", str(sh.cell(r, 12).value))
            break
        if "%" not in str(sh.cell(r, 11).number_format or "") or "%" not in str(sh.cell(r, 12).number_format or ""):
            fail("SH Summary", "high", f"K/L{r}", f"K={sh.cell(r,11).number_format!r} L={sh.cell(r,12).number_format!r}")
            break
    dups = [n for n, c in __import__("collections").Counter(sh_names).items() if c > 1]
    if dups:
        fail("SH Summary", "high", "C", f"duplicates {dups[:5]}")
    sheets["SH Summary"] = {"rows": len(sh_names), "f3": sh["F3"].value}

    # ---- SH Prior ----
    prior = ow["SH Prior"]
    prior_names = names_col(prior, 2, 1)
    old_sh_names = names_col(tw["SH Summary"], 6, 3)
    if prior_names != old_sh_names:
        fail("SH Prior", "high", "A", f"{len(prior_names)} vs old SH {len(old_sh_names)}")
    old_cb = tw["Data_TCOM+9961"]
    old_h = [old_cb.cell(3, c).value for c in range(1, 22)]
    old_tcom = old_h.index("TCOM Shares") + 1
    old_hk = old_h.index("9961_XHKG Shares") + 1
    old_shares = {}
    r = 4
    while isinstance(old_cb.cell(r, 1).value, str):
        old_shares[old_cb.cell(r, 1).value] = num(old_cb.cell(r, old_tcom).value) + num(old_cb.cell(r, old_hk).value)
        r += 1
    old_f3 = num(tw["SH Summary"]["F3"].value)
    share_mismatch = 0
    for i, name in enumerate(prior_names):
        got = prior.cell(i + 2, 3).value
        exp = old_shares.get(name)
        if got in (None, "") and exp is None:
            continue
        if exp is None or abs(num(got) - exp) > 1:
            share_mismatch += 1
            if share_mismatch <= 5:
                fail("SH Prior", "high", f"C{i+2}", f"{name} {got} != {exp}")
        so = prior.cell(i + 2, 4).value
        if exp and old_f3 and so is not None and abs(num(so) - exp / old_f3) > 1e-9:
            fail("SH Prior", "med", f"D{i+2}", f"%SO {so} != {exp/old_f3}")
            break
    if prior.sheet_state != "hidden":
        fail("SH Prior", "high", "visibility", prior.sheet_state)
    sheets["SH Prior"] = {"rows": len(prior_names), "share_mismatch": share_mismatch, "hidden": True}

    # ---- Top 20 ----
    t20 = ow["Top 20"]
    if t20.cell(2, 2).value != "核心股东持仓情况":
        fail("Top 20", "high", "B2", str(t20.cell(2, 2).value))
    if "TBU" in str(t20.cell(2, 2).value or ""):
        fail("Top 20", "high", "B2", "TBU")
    prior_top = [p for p in sorted(
        [(prior.cell(i, 1).value, prior.cell(i, 2).value) for i in range(2, 2 + len(prior_names))
         if isinstance(prior.cell(i, 2).value, (int, float))],
        key=lambda x: int(x[1]),
    )[:20]]
    for i in range(20):
        r = 11 + i
        if t20.cell(r, 2).value != prior_top[i][0]:
            fail("Top 20", "high", f"B{r}", f"{t20.cell(r,2).value} != prior#{i+1} {prior_top[i][0]}")
            break
        f = str(t20.cell(r, 3).value or "")
        if "SH Summary" not in f:
            fail("Top 20", "high", f"C{r}", f)
            break
        for c in (6, 7, 8):
            if "%" not in str(t20.cell(r, c).number_format or ""):
                fail("Top 20", "high", f"{get_column_letter(c)}{r}", f"fmt {t20.cell(r,c).number_format!r}")
                break
    for i in range(20):
        r = 35 + i
        if t20.cell(r, 2).value != sh_names[i]:
            fail("Top 20", "high", f"B{r}", f"{t20.cell(r,2).value} != SH#{i+1}")
            break
        for c in (6, 7, 8):
            if "%" not in str(t20.cell(r, c).number_format or ""):
                fail("Top 20", "high", f"{get_column_letter(c)}{r}", f"fmt {t20.cell(r,c).number_format!r}")
                break
    if t20.cell(31, 2).value not in (None, ""):
        fail("Top 20", "med", "B31", "leftover in prior block")
    if t20.cell(55, 2).value not in (None, ""):
        fail("Top 20", "med", "B55", "leftover in current block")
    if "Baidu" in str(t20.cell(35, 2).value):
        fail("Top 20", "high", "B35", "starts with Baidu")
    sheets["Top 20"] = {"prior1": t20.cell(11, 2).value, "cur1": t20.cell(35, 2).value}

    # ---- Regions ----
    country_to_region = load_mapping(ow["Mapping"])
    buckets = {"US&CA": [], "EU": [], "APAC": [], "ROW": []}
    unmapped = defaultdict(int)
    for row in pr:
        aum = num(row[pmap["Equity Assets USD, mm"]])
        if aum < AUM_FLOOR:
            continue
        itype = row[pmap["Investor Type"]] if isinstance(row[pmap["Investor Type"]], str) else ""
        if itype in EXCLUDE_TYPES:
            continue
        country = row[pmap["Country/Territory"]] if isinstance(row[pmap["Country/Territory"]], str) else ""
        mapped = country_to_region.get(country)
        if not mapped:
            if country:
                unmapped[country] += 1
            continue
        sheet = sheet_region(mapped)
        if sheet:
            buckets[sheet].append((row[0], aum))
    for k in buckets:
        buckets[k].sort(key=lambda x: -x[1])
    fx = float(MARKET["fx_usd_hkd"])
    for sheet_name in ("US&CA", "EU", "APAC", "ROW"):
        ws = ow[sheet_name]
        names = names_col(ws, 8, 3)
        expect = [n for n, _ in buckets[sheet_name]]
        if names != expect:
            fail(sheet_name, "high", "C8+", f"{len(names)} vs {len(expect)} first {names[:2]} vs {expect[:2]}")
        for col, header in REGION_ROW5.items():
            want = ph.index(header) + 1
            got = ws.cell(5, col).value
            if got != want:
                fail(sheet_name, "high", f"{get_column_letter(col)}5", f"{got} != {want} ({header})")
        for col, ticker in MCAP_USD_CELLS.items():
            want = usd_mcap(MARKET, ticker)
            if abs(num(ws.cell(5, col).value) - want) > 1:
                fail(sheet_name, "high", f"{get_column_letter(col)}5", f"mcap {ws.cell(5,col).value} != {want}")
        for col, ticker in MCAP_HKD_CELLS.items():
            v = str(ws.cell(5, col).value or "")
            if not v.startswith("=") or str(int(MARKET["hkd_market_cap"][ticker])) not in v.replace(".0", ""):
                # formula stores int; allow float string
                if str(MARKET["hkd_market_cap"][ticker]) not in v and str(int(MARKET["hkd_market_cap"][ticker])) not in v:
                    fail(sheet_name, "high", f"{get_column_letter(col)}5", f"hkd formula {v}")
            if str(fx) not in v and f"{fx}" not in v:
                fail(sheet_name, "med", f"{get_column_letter(col)}5", f"fx {v}")
        leftover = ws.cell(8 + len(names), 3).value if names else ws.cell(8, 3).value
        b3 = str(ws["B3"].value or "")
        if WORKBOOK_AS_OF not in b3 or HOLDINGS_AS_OF not in b3:
            fail(sheet_name, "high", "B3", b3)
        # leftover is first empty after names - next row should be empty
        if names and ws.cell(8 + len(names), 3).value not in (None, ""):
            fail(sheet_name, "high", f"C{8+len(names)}", f"leftover {ws.cell(8+len(names),3).value}")
        if names:
            f8 = str(ws.cell(8, 7).value or "")
            if "G$5" not in f8 or "DATA_ALL" not in f8:
                fail(sheet_name, "high", "G8", f8)
            if "G5274" in f8:
                fail(sheet_name, "high", "G8", "old hardcoded sum")
            last_r = 7 + len(names)
            if "G$5" not in str(ws.cell(last_r, 7).value or ""):
                fail(sheet_name, "high", f"G{last_r}", "formula missing on last row")
            if "L$5" not in str(ws.cell(last_r, 12).value or ""):
                fail(sheet_name, "high", f"L{last_r}", str(ws.cell(last_r, 12).value))
            for rr in range(8, last_r + 1):
                bad = [c for c in REGION_PCT_COLS if "%" not in str(ws.cell(rr, c).number_format or "")]
                if bad:
                    fail(
                        sheet_name,
                        "high",
                        f"{get_column_letter(bad[0])}{rr}",
                        f"fmt {ws.cell(rr, bad[0]).number_format!r}",
                    )
                    break
        sheets[sheet_name] = {"n": len(names), "first": names[0] if names else None}

    if unmapped:
        fail("regions", "med", "Mapping", f"unmapped {dict(unmapped)}")

    # ---- Map ----
    mp = ow["全球投资人地图"]
    if mp["B3"].value != f"updated as of {WORKBOOK_AS_OF}":
        fail("全球投资人地图", "high", "B3", str(mp["B3"].value))
    if mp["D8"].value != MAP_D8:
        fail("全球投资人地图", "high", "D8", str(mp["D8"].value))
    if mp["E16"].value != "no ADR holding Equity AUM":
        fail("全球投资人地图", "high", "E16", str(mp["E16"].value))
    if mp["E17"].value != "no Travel holding Equity AUM":
        fail("全球投资人地图", "high", "E17", str(mp["E17"].value))
    if mp["E18"].value != "no TCOM holding":
        fail("全球投资人地图", "high", "E18", str(mp["E18"].value))
    c9 = str(mp["C9"].value or "")
    if "C8" not in c9 or "C7" not in c9:
        fail("全球投资人地图", "high", "C9", c9)
    if mp["C5"].value != len(pr):
        fail("全球投资人地图", "high", "C5", str(mp["C5"].value))
    tcom_l = get_column_letter(ph.index("Value TCOM USD, mm") + 1)
    hk_l = get_column_letter(ph.index("Value 9961_XHKG USD, mm") + 1)
    aum_l = get_column_letter(ph.index("Equity Assets USD, mm") + 1)
    c8 = str(mp["C8"].value or "")
    if tcom_l not in c8 or hk_l not in c8 or str(sum_row) not in c8:
        fail("全球投资人地图", "high", "C8", c8)
    if "G5274" in c8:
        fail("全球投资人地图", "high", "C8", "old G5274")
    c6 = str(mp["C6"].value or "")
    if aum_l not in c6:
        fail("全球投资人地图", "high", "C6", c6)
    c13 = str(mp["C13"].value or "")
    for t in ADR_VALUE_TICKERS:
        letter = get_column_letter(ph.index(f"Value {t} USD, mm") + 1)
        if letter not in c13:
            fail("全球投资人地图", "high", "C13", f"missing {t} col {letter}")
    c14 = str(mp["C14"].value or "")
    for t in TRAVEL_VALUE_TICKERS:
        letter = get_column_letter(ph.index(f"Value {t} USD, mm") + 1)
        if letter not in c14:
            fail("全球投资人地图", "high", "C14", f"missing {t} col {letter}")
    # independent D16-18
    def has_any(row, tickers):
        return any(num(row[pmap[f"Value {t} USD, mm"]]) != 0 for t in tickers if f"Value {t} USD, mm" in pmap)
    no_adr = no_travel = no_tcom = 0.0
    aum_i = pmap["Equity Assets USD, mm"]
    for row in pr:
        aum = num(row[aum_i])
        if not has_any(row, ADR_VALUE_TICKERS):
            no_adr += aum
        if not has_any(row, TRAVEL_VALUE_TICKERS):
            no_travel += aum
        if num(row[pmap["Value TCOM USD, mm"]]) == 0 and num(row[pmap["Value 9961_XHKG USD, mm"]]) == 0:
            no_tcom += aum
    if abs(num(mp["D16"].value) - no_adr) > 1:
        fail("全球投资人地图", "high", "D16", f"{mp['D16'].value} != {no_adr}")
    if abs(num(mp["D17"].value) - no_travel) > 1:
        fail("全球投资人地图", "high", "D17", f"{mp['D17'].value} != {no_travel}")
    if abs(num(mp["D18"].value) - no_tcom) > 1:
        fail("全球投资人地图", "high", "D18", f"{mp['D18'].value} != {no_tcom}")
    sheets["全球投资人地图"] = {"c5": mp["C5"].value, "c8": c8}

    # ---- contacts / mapping vs template ----
    for name in ("重要投资人", "普通投资人", "Mapping"):
        a, b = tw[name], ow[name]
        diffs = 0
        max_r = max(a.max_row or 1, b.max_row or 1)
        max_c = max(a.max_column or 1, b.max_column or 1)
        sample = []
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                if not cell_eq(a.cell(r, c).value, b.cell(r, c).value):
                    diffs += 1
                    if len(sample) < 5:
                        sample.append(f"{get_column_letter(c)}{r}")
        if diffs:
            fail(name, "high", "copy", f"{diffs} cells differ e.g. {sample}")
        sheets[name] = {"max_row": b.max_row, "diffs": diffs}

    # ---- Market ----
    mk = ow["Market"]
    if mk["B1"].value != VALID_AS_OF.replace("/", "-"):
        fail("Market", "high", "B1", str(mk["B1"].value))
    if mk.sheet_state != "hidden":
        fail("Market", "high", "visibility", mk.sheet_state)
    if abs(num(mk["B2"].value) - float(MARKET["fx_usd_hkd"])) > 1e-6:
        fail("Market", "high", "B2", str(mk["B2"].value))
    if mk["B3"].value != MARKET["tcom_shares_outstanding"]:
        fail("Market", "high", "B3", str(mk["B3"].value))
    listed = {}
    r = 6
    while isinstance(mk.cell(r, 1).value, str):
        listed[mk.cell(r, 1).value] = (mk.cell(r, 2).value, num(mk.cell(r, 3).value))
        r += 1
    for t, v in MARKET["usd_market_cap"].items():
        if t not in listed or listed[t][0] != "USD" or abs(listed[t][1] - v) > 1:
            fail("Market", "high", t, f"{listed.get(t)} != {v}")
    for t, v in MARKET["hkd_market_cap"].items():
        if t not in listed or listed[t][0] != "HKD" or abs(listed[t][1] - v) > 1:
            fail("Market", "high", t, f"{listed.get(t)} != {v}")
    sheets["Market"] = {"as_of": mk["B1"].value, "fx": mk["B2"].value, "so": mk["B3"].value}

    # leftover [1] formulas anywhere (sample used ranges)
    for ws in ow.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, 80), max_col=min(ws.max_column or 1, 50)):
            for cell in row:
                if isinstance(cell.value, str) and "[1]" in cell.value:
                    fail(ws.title, "high", cell.coordinate, "external link formula")

    ow.close()
    tw.close()

    report = {
        "output": str(OUT),
        "n_findings": len(findings),
        "high": sum(1 for f in findings if f["severity"] == "high"),
        "med": sum(1 for f in findings if f["severity"] == "med"),
        "findings": findings,
        "sheets": sheets,
    }
    REPORT.parent.mkdir(parents=True, exist_ok=True)
    REPORT.write_text(json.dumps(report, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
    print(json.dumps({"n": report["n_findings"], "high": report["high"], "med": report["med"], "findings": findings, "output": str(OUT)}, indent=2, ensure_ascii=False))
    return report


def main(argv: list[str] | None = None) -> int:
    import argparse

    p = argparse.ArgumentParser(description="Adversarial audit of a generated Investor List")
    p.add_argument("--output", type=Path, default=None)
    p.add_argument("--peer", type=Path, default=None)
    p.add_argument("--combined", type=Path, default=None)
    p.add_argument("--template", type=Path, default=None)
    p.add_argument("--market", type=Path, default=None)
    p.add_argument("--report", type=Path, default=None)
    args = p.parse_args(argv)
    report = run_audit(
        output=args.output,
        peer=args.peer,
        combined=args.combined,
        template=args.template,
        market_path=args.market,
        report_path=args.report,
    )
    return 1 if report["n_findings"] else 0


if __name__ == "__main__":
    raise SystemExit(main())
