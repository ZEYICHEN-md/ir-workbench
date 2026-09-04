"""Build Investor List xlsx from two Capital IQ extracts + prior workbook."""
from __future__ import annotations

import io
import json
import re
import shutil
import sys
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.formatting.rule import FormatObject, IconSet, Rule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

STRATEGIC = frozenset({"Baidu Holdings, LTD"})
AUM_FLOOR = 50_000_000_000
EXCLUDE_TYPES = frozenset({"Corporate"})
APAC_REGIONS = frozenset({"Asia", "CN", "HK", "SG", "Oceania"})
PRIOR_Q = "26Q1"
CUR_Q = "26Q2"
HOLDINGS_AS_OF = "2026/06/30"
# This refresh's validity cutoff (extract + market caps + workbook).
VALID_AS_OF = "2026/08/31"
WORKBOOK_AS_OF = "August 31, 2026"
MAP_D8 = "*小于Total Market Cap, 由于目前可以抓取到的公开披露信息，cover约53%的股东持仓数据"
INDEX_FILL = PatternFill(start_color="E2EEDA", end_color="E2EEDA", fill_type="solid")
NO_FILL = PatternFill(fill_type=None)
# Region holding / peer-%-of-mcap columns. Rewrite every name row; do not inherit General.
REGION_PCT_COLS = (10, 13, 17, 19, 23, 25, 29, 33, 35, 37, 39, 41, 43, 45, 47)


def output_filename(valid: str | None = None) -> str:
    y, m, d = (valid or VALID_AS_OF).replace("-", "/").split("/")
    return f"Investor List_{int(y):04d}{int(m):02d}{int(d):02d}.xlsx"


def period_key(valid: str | None = None) -> str:
    """Workbench period key (data_date): 2026/08/31 → 2026-08-31."""
    y, m, d = (valid or VALID_AS_OF).replace("-", "/").split("/")
    return f"{int(y):04d}-{int(m):02d}-{int(d):02d}"

COMBINED_KEEP = [
    "Institution",
    "Sustainability Considered",
    "Holder Rank (Combined)",
    "Shares (Combined)",
    "Share Change (Combined)",
    "% S/O (Combined)",
    "TCOM Shares",
    "TCOM Share Change",
    "TCOM Reported Date",
    "9961_XHKG Shares",
    "9961_XHKG Share Change",
    "9961_XHKG Reported Date",
    "Equity Assets USD, mm",
    "Style",
    "Turnover",
    "City",
    "State/Province",
    "Country/Territory",
]

ADR_VALUE_TICKERS = [
    "700_XHKG", "BABA", "9988_XHKG", "3690_XHKG", "NTES", "9999_XHKG",
    "PDD", "JD", "9618_XHKG", "BIDU", "9888_XHKG",
]
TRAVEL_VALUE_TICKERS = [
    "ABNB", "BKNG", "EXPE", "MMYT", "TRIP", "TRVG", "780_XHKG",
]


@dataclass
class Paths:
    template: Path
    peer: Path
    combined: Path
    output: Path
    market: Path


def col_map(ws: Worksheet, header_row: int) -> dict[str, int]:
    out: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(header_row, c).value
        if isinstance(h, str) and h.strip():
            out[h.strip()] = c
    return out


def header_row_of(ws: Worksheet) -> int:
    for r in range(1, 8):
        if ws.cell(r, 1).value == "Institution":
            return r
    raise ValueError(f"{ws.title}: no Institution header")


def read_table(ws: Worksheet, header_row: int) -> tuple[list[str], list[list[Any]]]:
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    while headers and headers[-1] is None:
        headers.pop()
    rows: list[list[Any]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not isinstance(name, str) or not name.strip():
            continue
        rows.append([ws.cell(r, c).value for c in range(1, len(headers) + 1)])
    return [h if h is not None else "" for h in headers], rows


def num(v: Any) -> float:
    return float(v) if isinstance(v, (int, float)) else 0.0


def fund_size_formula(aum_cell: str) -> str:
    return (
        f'IF({aum_cell}>100000000000,"100bn",'
        f'IF({aum_cell}>50000000000,"50bn - 100bn",'
        f'IF({aum_cell}>1000000000,"1bn - 50bn",'
        f'IF({aum_cell}>500000000,"500mn - 1bn","500mn"))))'
    )


def unmerge_from(ws: Worksheet, start_row: int) -> None:
    for rng in list(ws.merged_cells.ranges):
        if rng.max_row >= start_row:
            ws.unmerge_cells(str(rng))


def clear_below(ws: Worksheet, start_row: int) -> None:
    unmerge_from(ws, start_row)
    last = ws.max_row or start_row
    if last < start_row:
        return
    max_col = ws.max_column or 1
    for row in ws.iter_rows(min_row=start_row, max_row=last, max_col=max_col):
        for cell in row:
            if cell.value is not None:
                cell.value = None


def replace_sheet_keep_widths(wb: Workbook, title: str) -> Worksheet:
    old = wb[title]
    idx = wb.sheetnames.index(title)
    widths = {col: dim.width for col, dim in old.column_dimensions.items() if dim.width}
    wb.remove(old)
    ws = wb.create_sheet(title, idx)
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    return ws


def col_letter(headers: list[str], name: str) -> str:
    return get_column_letter(headers.index(name) + 1)


def index_institutions(rows: list[list[Any]], cmap: dict[str, int]) -> frozenset[str]:
    si = cmap["Style"] - 1
    out: set[str] = set()
    for row in rows:
        name = row[0]
        if isinstance(name, str) and name.strip() and row[si] == "Index":
            out.add(name.strip())
    return frozenset(out)


def paint_index_row(ws: Worksheet, row: int, cols: range, name: str, index_names: frozenset[str]) -> None:
    fill = INDEX_FILL if name in index_names else NO_FILL
    for c in cols:
        ws.cell(row, c).fill = fill


def short_holder(name: str) -> str:
    s = re.sub(r"\s*\([^)]*\)", "", name.split(",")[0]).strip()
    parts = [p for p in s.split() if p.lower() != "the"]
    return " ".join(parts[:2]) if parts else name


def write_top20_takeaways(
    ws: Worksheet,
    current_top: list[str],
    prior: list[tuple[str, Any, Any, Any]],
    shares: dict[str, float],
) -> None:
    prior_top = [p[0] for p in sorted(
        [(n, rk) for n, rk, _s, _p in prior if isinstance(rk, (int, float))],
        key=lambda x: int(x[1]),
    )[:20]]
    prior_sh = {n: num(s) for n, _rk, s, _p in prior}
    cuts: list[tuple[float, str]] = []
    adds: list[tuple[float, str]] = []
    seen: set[str] = set()
    for name in current_top + prior_top:
        if name in seen:
            continue
        seen.add(name)
        old, new = prior_sh.get(name), shares.get(name)
        if not old or old == 0 or new is None:
            continue
        pct = new / old - 1
        if pct <= -0.04:
            cuts.append((pct, name))
        elif 0.04 <= pct <= 3.0:
            adds.append((pct, name))
    cuts.sort()
    adds.sort(reverse=True)

    def bits(items: list[tuple[float, str]], verb: str, n: int) -> str:
        return "， ".join(
            f"{short_holder(nme)} {verb}{abs(round(pct * 100))}%"
            for pct, nme in items[:n]
        )

    ws["B4"] = f"需关注 {bits(cuts, '减持', 6)}" if cuts else None
    ws["B5"] = f"上季度加仓: {bits(adds, '增持', 5)}" if adds else None
    ws["B6"] = None


def fix_rank_change_arrows(ws: Worksheet) -> None:
    """3Arrows vs zero: green if rank improved (C-D>0), yellow if 0, red if fell."""
    cf = ws.conditional_formatting
    for key in list(cf._cf_rules.keys()):
        kept = [r for r in cf._cf_rules[key] if getattr(r, "type", None) != "iconSet"]
        if kept:
            cf._cf_rules[key] = kept
        else:
            del cf._cf_rules[key]
    for rng in ("E11:E30", "E35:E54"):
        rule = Rule(
            type="iconSet",
            iconSet=IconSet(
                iconSet="3Arrows",
                showValue=False,
                cfvo=[
                    FormatObject(type="percent", val=0),
                    FormatObject(type="num", val=0),
                    FormatObject(type="num", val=0, gte=False),
                ],
            ),
        )
        ws.conditional_formatting.add(rng, rule)


def write_block(ws: Worksheet, start_row: int, rows: list[list[Any]]) -> None:
    for i, row in enumerate(rows):
        r = start_row + i
        for c, v in enumerate(row, 1):
            ws.cell(r, c).value = v


def load_mapping(ws: Worksheet) -> dict[str, str]:
    out: dict[str, str] = {}
    for r in range(7, ws.max_row + 1):
        country, region = ws.cell(r, 2).value, ws.cell(r, 3).value
        if isinstance(country, str) and isinstance(region, str):
            out[country.strip()] = region.strip()
    return out


def sheet_region(mapped: str) -> str | None:
    if mapped == "US/CA":
        return "US&CA"
    if mapped == "EU":
        return "EU"
    if mapped in APAC_REGIONS:
        return "APAC"
    if mapped:
        return "ROW"
    return None


def load_market(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def usd_mcap(market: dict[str, Any], ticker: str) -> float:
    usd = market["usd_market_cap"]
    hkd = market["hkd_market_cap"]
    fx = float(market["fx_usd_hkd"])
    if ticker in usd:
        return float(usd[ticker])
    if ticker in hkd:
        return float(hkd[ticker]) / fx
    raise KeyError(ticker)


def _log(msg: str) -> None:
    print(msg, file=sys.stderr, flush=True)


def build(paths: Paths) -> dict[str, Any]:
    _log(f"copy template -> {paths.output}")
    shutil.copy2(paths.template, paths.output)
    wb = load_workbook(paths.output)
    peer_wb = load_workbook(paths.peer, data_only=False)
    comb_wb = load_workbook(paths.combined, data_only=False)
    market = load_market(paths.market)

    peer_ws = peer_wb.active
    comb_ws = comb_wb.active
    peer_hrow = header_row_of(peer_ws)
    comb_hrow = header_row_of(comb_ws)
    peer_headers, peer_rows = read_table(peer_ws, peer_hrow)
    comb_headers, comb_rows = read_table(comb_ws, comb_hrow)
    peer_wb.close()
    comb_wb.close()

    cmap = {h: i + 1 for i, h in enumerate(comb_headers)}
    pmap = {h: i + 1 for i, h in enumerate(peer_headers)}
    for name in COMBINED_KEEP:
        if name not in cmap:
            raise ValueError(f"combined extract missing {name!r}")
    for name in (
        "Institution",
        "Value TCOM USD, mm",
        "Shares TCOM",
        "Value 9961_XHKG USD, mm",
        "Equity Assets USD, mm",
        "Country/Territory",
        "Style",
        "Investor Type",
    ):
        if name not in pmap:
            raise ValueError(f"peer extract missing {name!r}")

    country_to_region = load_mapping(wb["Mapping"])
    prior = snapshot_prior(wb)
    index_names = index_institutions(comb_rows, cmap)
    shares_now = {
        row[0]: num(row[cmap["Shares (Combined)"] - 1])
        for row in comb_rows
        if isinstance(row[0], str)
    }
    _log(f"write DATA_ALL ({len(peer_rows)} institutions)")
    da_stats = write_data_all(replace_sheet_keep_widths(wb, "DATA_ALL"), peer_headers, peer_rows)
    _log(f"write Data_TCOM+9961 ({len(comb_rows)} institutions)")
    write_combined(wb["Data_TCOM+9961"], comb_rows, cmap)
    write_sh_prior(wb, prior)
    _log("write SH Summary / Top 20 / regions / map / Market")
    sh_names = write_sh_summary(wb["SH Summary"], comb_rows, market, index_names)
    write_top20(wb["Top 20"], sh_names, prior, shares_now, index_names)
    region_stats = write_regions(wb, peer_headers, peer_rows, country_to_region, market, pmap, index_names)
    write_map(wb["全球投资人地图"], peer_headers, peer_rows, da_stats, pmap, market)
    write_market_sheet(wb, market)

    stats = {
        "output": str(paths.output),
        "peer_rows": len(peer_rows),
        "combined_rows": len(comb_rows),
        "sh_rows": len(sh_names),
        "data_all_sum_row": da_stats["sum_row"],
        "tcom_value_col": pmap["Value TCOM USD, mm"],
        "hk_value_col": pmap["Value 9961_XHKG USD, mm"],
        "style_col": pmap["Style"],
        "country_col": pmap["Country/Territory"],
        "aum_col": pmap["Equity Assets USD, mm"],
        "region_counts": region_stats["counts"],
        "unmapped_countries": region_stats["unmapped"],
    }
    _log("saving workbook")
    try:
        wb._external_links.clear()
    except Exception:
        wb._external_links = []
    wb.save(paths.output)
    wb.close()
    strip_external_links(paths.output)
    return stats


def strip_external_links(path: Path) -> None:
    """Drop cached [1] workbook links copied from the prior Investor List."""
    buf = io.BytesIO()
    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            name = item.filename.replace("\\", "/")
            if name.startswith("xl/externalLinks/"):
                continue
            data = zin.read(item.filename)
            if name == "xl/workbook.xml":
                data = re.sub(rb"<externalReferences[^>]*>.*?</externalReferences>", b"", data, flags=re.S)
                data = re.sub(rb"<externalReferences[^/]*/>", b"", data)
            elif name == "xl/_rels/workbook.xml.rels":
                data = re.sub(
                    rb'<Relationship[^>]+Type="[^"]*externalLink[^"]*"[^/]*/>',
                    b"",
                    data,
                )
            elif name == "[Content_Types].xml":
                data = re.sub(
                    rb'<Override[^>]+PartName="/xl/externalLinks/[^"]*"[^/]*/>',
                    b"",
                    data,
                )
            zout.writestr(item, data)
    path.write_bytes(buf.getvalue())


def write_data_all(ws: Worksheet, headers: list[str], rows: list[list[Any]]) -> dict[str, int]:
    n = len(headers)
    for c in range(1, n + 3):
        ws.cell(1, c).value = c
    for c, h in enumerate(headers, 1):
        ws.cell(2, c).value = h
    ws.cell(2, n + 1).value = "Fund Size"
    ws.cell(2, n + 2).value = "Region"
    aum_letter = col_letter(headers, "Equity Assets USD, mm")
    country_letter = col_letter(headers, "Country/Territory")
    first, last = 3, 2 + len(rows)
    write_block(ws, first, rows)
    for r in range(first, last + 1):
        ws.cell(r, n + 1).value = "=" + fund_size_formula(f"{aum_letter}{r}")
        ws.cell(r, n + 2).value = f"=IFERROR(VLOOKUP({country_letter}{r},Mapping!B:C,2,FALSE),\"\")"
    sum_row = last + 1
    ws.cell(sum_row, 1).value = "TOTAL"
    for c in range(2, n + 3):
        ws.cell(sum_row, c).value = None
    for c, h in enumerate(headers, 1):
        if h.startswith("Value ") or h in ("Aggregate Value USD, mm", "Equity Assets USD, mm"):
            letter = get_column_letter(c)
            ws.cell(sum_row, c).value = f"=SUM({letter}{first}:{letter}{last})"
    ws.freeze_panes = "B3"
    return {"first": first, "last": last, "sum_row": sum_row, "n_inst": len(rows)}


def write_combined(ws: Worksheet, rows: list[list[Any]], cmap: dict[str, int]) -> None:
    keep = COMBINED_KEEP
    data = [[row[cmap[h] - 1] for h in keep] for row in rows]
    ws.cell(1, 1).value = "Combined Ownership"
    for c in range(1, 22):
        ws.cell(2, c).value = c
    for c, h in enumerate(keep, 1):
        ws.cell(3, c).value = h
    ws.cell(3, 19).value = "Fund Size"
    ws.cell(3, 20).value = "Region"
    ws.cell(3, 21).value = "Fund Style"
    first = 4
    write_block(ws, first, data)
    last = first + len(data) - 1
    aum_l = col_letter(keep, "Equity Assets USD, mm")
    country_l = col_letter(keep, "Country/Territory")
    style_l = col_letter(keep, "Style")
    style_i = keep.index("Style")
    for r in range(first, last + 1):
        ws.cell(r, 19).value = "=" + fund_size_formula(f"{aum_l}{r}")
        ws.cell(r, 20).value = f"=IFERROR(VLOOKUP({country_l}{r},Mapping!B:C,2,FALSE),\"\")"
        ws.cell(r, 21).value = f"=IFERROR(VLOOKUP({style_l}{r},Mapping!H:I,2,FALSE),\"\")"
        name = data[r - first][0]
        if isinstance(name, str):
            paint_index_row(
                ws, r, range(1, 19), name,
                frozenset({name}) if data[r - first][style_i] == "Index" else frozenset(),
            )
    clear_below(ws, last + 1)
    ws.freeze_panes = "B4"


def snapshot_prior(wb: Workbook) -> list[tuple[str, Any, Any, Any]]:
    sh = wb["SH Summary"]
    f3 = num(sh["F3"].value) or 648991284
    shares: dict[str, float] = {}
    cb = wb["Data_TCOM+9961"]
    headers = [cb.cell(3, c).value for c in range(1, 22)]
    tcom_c = headers.index("TCOM Shares") + 1
    hk_c = headers.index("9961_XHKG Shares") + 1
    for r in range(4, cb.max_row + 1):
        name = cb.cell(r, 1).value
        if isinstance(name, str) and name.strip():
            shares[name.strip()] = num(cb.cell(r, tcom_c).value) + num(cb.cell(r, hk_c).value)
    out: list[tuple[str, Any, Any, Any]] = []
    for r in range(6, sh.max_row + 1):
        name = sh.cell(r, 3).value
        if not isinstance(name, str) or not name.strip():
            continue
        key = name.strip()
        shs = shares.get(key)
        so = (shs / f3) if shs else None
        out.append((key, sh.cell(r, 2).value, shs, so))
    return out


def write_sh_prior(wb: Workbook, prior: list[tuple[str, Any, Any, Any]]) -> None:
    if "SH Prior" in wb.sheetnames:
        del wb["SH Prior"]
    idx = wb.sheetnames.index("SH Summary") + 1
    ws = wb.create_sheet("SH Prior", idx)
    ws.sheet_state = "hidden"
    ws["A1"] = "Institution"
    ws["B1"] = f"{PRIOR_Q} Rank"
    ws["C1"] = f"Total Shares {PRIOR_Q}"
    ws["D1"] = f"% S/O {PRIOR_Q}"
    for i, (name, rank, shares, so) in enumerate(prior, 2):
        ws.cell(i, 1).value = name
        ws.cell(i, 2).value = rank
        ws.cell(i, 3).value = shares
        ws.cell(i, 4).value = so


def write_sh_summary(
    ws: Worksheet,
    rows: list[list[Any]],
    market: dict[str, Any],
    index_names: frozenset[str],
) -> list[str]:
    names = [
        row[0]
        for row in rows
        if isinstance(row[0], str) and row[0].strip() not in STRATEGIC
    ]
    so = int(market["tcom_shares_outstanding"])
    ws["C1"] = "Trip.com Group"
    ws["C2"] = f"{CUR_Q} Institutional Ownership"
    ws["F3"] = so
    ws["G3"] = f"updated as of {VALID_AS_OF}"
    keep_index = {name: i + 1 for i, name in enumerate(COMBINED_KEEP)}
    idx = {
        "D": keep_index["TCOM Shares"],
        "E": keep_index["9961_XHKG Shares"],
        "M": keep_index["TCOM Reported Date"],
        "N": keep_index["9961_XHKG Reported Date"],
        "O": keep_index["Style"],
        "P": keep_index["City"],
        "Q": keep_index["State/Province"],
        "R": keep_index["Country/Territory"],
        "S": keep_index["Equity Assets USD, mm"],
    }
    for col, val in idx.items():
        ws.cell(4, ord(col) - 64).value = val
    ws["A5"] = f"{PRIOR_Q} Rank"
    ws["B5"] = f"{CUR_Q} Rank"
    ws["C5"] = "Institution Name"
    ws["D5"] = "TCOM"
    ws["E5"] = 9961
    ws["F5"] = f"Total Shares {CUR_Q}"
    ws["G5"] = f"Total Shares {PRIOR_Q}"
    ws["H5"] = "QoQ"
    ws["I5"] = "% S/O 25Q2"
    ws["J5"] = "% S/O 25Q3"
    ws["K5"] = f"% S/O {PRIOR_Q}"
    ws["L5"] = f"% S/O {CUR_Q}"
    ws["M5"] = "Position Date (TCOM)"
    ws["N5"] = "Position Date (9961)"
    ws["O5"] = "Dominant Style"
    ws["P5"] = "City"
    ws["Q5"] = "State"
    ws["R5"] = "Country"
    ws["S5"] = "Equity AUM $bn"
    clear_below(ws, 6)
    first = 6
    for i, name in enumerate(names):
        r = first + i
        ws.cell(r, 1).value = (
            f'=IFERROR(INDEX(\'SH Prior\'!B:B,MATCH(C{r},\'SH Prior\'!A:A,0)),"")'
        )
        ws.cell(r, 2).value = i + 1
        ws.cell(r, 3).value = name
        for col in ("D", "E", "M", "N", "O", "P", "Q", "R"):
            c = ord(col) - 64
            ws.cell(r, c).value = (
                f"=VLOOKUP($C{r},'Data_TCOM+9961'!$A:$U,'SH Summary'!{col}$4,FALSE)"
            )
        ws.cell(r, 6).value = f"=IFERROR(D{r},0)+IFERROR(E{r},0)"
        ws.cell(r, 7).value = (
            f'=IFERROR(INDEX(\'SH Prior\'!C:C,MATCH(C{r},\'SH Prior\'!A:A,0)),"")'
        )
        ws.cell(r, 8).value = f"=IFERROR(F{r}-G{r},F{r})"
        ws.cell(r, 11).value = (
            f'=IFERROR(INDEX(\'SH Prior\'!D:D,MATCH(C{r},\'SH Prior\'!A:A,0)),"")'
        )
        ws.cell(r, 12).value = f"=F{r}/$F$3"
        ws.cell(r, 19).value = (
            f"=VLOOKUP($C{r},'Data_TCOM+9961'!$A:$U,'SH Summary'!S$4,FALSE)/1000000000"
        )
        ws.cell(r, 11).number_format = "0.0%"
        ws.cell(r, 12).number_format = "0.0%"
        paint_index_row(ws, r, range(3, 20), name, index_names)
    leftover = first + len(names)
    for r in range(leftover, min((ws.max_row or leftover) + 1, leftover + 40)):
        paint_index_row(ws, r, range(3, 20), "", frozenset())
    ws.freeze_panes = "D6"
    return names


def write_top20(
    ws: Worksheet,
    sh_names: list[str],
    prior: list[tuple[str, Any, Any, Any]],
    shares: dict[str, float],
    index_names: frozenset[str],
) -> None:
    prior_top = [p[0] for p in sorted(
        [(n, rk) for n, rk, _s, _p in prior if isinstance(rk, (int, float))],
        key=lambda x: int(x[1]),
    )[:20]]
    current_top = sh_names[:20]
    ws["B2"] = "核心股东持仓情况"
    ws["B3"] = f"· {PRIOR_Q} -> {CUR_Q}, as of {WORKBOOK_AS_OF}"
    write_top20_takeaways(ws, current_top, prior, shares)
    ws["B8"] = f"Top 20  {PRIOR_Q}-> {CUR_Q}"
    headers = [
        None,
        "Institution Name",
        f"{PRIOR_Q} Rank",
        f"{CUR_Q} Rank",
        "Rank Change",
        f"% S/O {PRIOR_Q}",
        f"% S/O {CUR_Q}",
        "% S/O Change",
    ]
    for c, h in enumerate(headers, 1):
        if h:
            ws.cell(10, c).value = h

    def fill(start_row: int, names: list[str]) -> None:
        for i, name in enumerate(names):
            r = start_row + i
            ws.cell(r, 2).value = name
            ws.cell(r, 3).value = (
                f"=IFERROR(INDEX('SH Summary'!A:A,MATCH(B{r},'SH Summary'!C:C,0)),\"\")"
            )
            ws.cell(r, 4).value = (
                f"=IFERROR(INDEX('SH Summary'!B:B,MATCH(B{r},'SH Summary'!C:C,0)),\"\")"
            )
            ws.cell(r, 5).value = f"=IF(OR(C{r}=\"\",D{r}=\"\"),\"\",C{r}-D{r})"
            ws.cell(r, 6).value = (
                f"=IFERROR(VLOOKUP(B{r},'SH Summary'!C:L,9,FALSE),\"\")"
            )
            ws.cell(r, 7).value = (
                f"=IFERROR(VLOOKUP(B{r},'SH Summary'!C:L,10,FALSE),\"\")"
            )
            ws.cell(r, 8).value = (
                f"=IF(G{r}=\"\",\"Exited\","
                f"IF(OR(F{r}=0,F{r}=\"\"),\"New Entry\",G{r}/F{r}-1))"
            )
            ws.cell(r, 6).number_format = "0.0%"
            ws.cell(r, 7).number_format = "0.0%"
            ws.cell(r, 8).number_format = "0%"
            paint_index_row(ws, r, range(2, 9), name, index_names)

    fill(11, prior_top)
    ws["B32"] = f"Top 20  {CUR_Q}"
    for c, h in enumerate(headers, 1):
        if h:
            ws.cell(34, c).value = h
    fill(35, current_top)
    # clear leftover names below current 20 in both blocks
    for r in range(11 + len(prior_top), 31):
        for c in range(2, 9):
            ws.cell(r, c).value = None
            ws.cell(r, c).fill = NO_FILL
    for r in range(35 + len(current_top), 55):
        for c in range(2, 9):
            ws.cell(r, c).value = None
            ws.cell(r, c).fill = NO_FILL
    fix_rank_change_arrows(ws)


def write_regions(
    wb: Workbook,
    headers: list[str],
    rows: list[list[Any]],
    country_to_region: dict[str, str],
    market: dict[str, Any],
    pmap: dict[str, int],
    index_names: frozenset[str],
) -> dict[str, Any]:
    inst_i = pmap["Institution"] - 1
    aum_i = pmap["Equity Assets USD, mm"] - 1
    cty_i = pmap["Country/Territory"] - 1
    type_i = pmap.get("Investor Type", 0) - 1
    buckets: dict[str, list[tuple[str, float]]] = {"US&CA": [], "EU": [], "APAC": [], "ROW": []}
    unmapped: dict[str, int] = {}
    for row in rows:
        name = row[inst_i]
        aum = num(row[aum_i])
        country = row[cty_i] if isinstance(row[cty_i], str) else ""
        itype = row[type_i] if type_i >= 0 else ""
        if aum < AUM_FLOOR:
            continue
        if itype in EXCLUDE_TYPES:
            continue
        mapped = country_to_region.get(country)
        if not mapped:
            if country:
                unmapped[country] = unmapped.get(country, 0) + 1
            continue
        sheet = sheet_region(mapped)
        if sheet:
            buckets[sheet].append((name, aum))
    for sheet in buckets:
        buckets[sheet].sort(key=lambda x: -x[1])

    fx = float(market["fx_usd_hkd"])
    style_col = pmap["Style"]
    aum_col = pmap["Equity Assets USD, mm"]
    country_col = pmap["Country/Territory"]
    tcom_col = pmap["Value TCOM USD, mm"]
    hk_col = pmap["Value 9961_XHKG USD, mm"]
    da_last = get_column_letter(len(headers) + 2)
    da_range = f"DATA_ALL!$A:${da_last}"

    def peer_col(ticker: str) -> int:
        return pmap[f"Value {ticker} USD, mm"]

    for sheet_name, names in buckets.items():
        ws = wb[sheet_name]
        ws["B2"] = f"{sheet_name} Top Investors"
        ws["B3"] = f"As of {WORKBOOK_AS_OF}, AUM > $50bn, holdings as of {HOLDINGS_AS_OF}"
        ws.cell(5, 4).value = style_col
        ws.cell(5, 5).value = aum_col
        ws.cell(5, 6).value = country_col
        ws.cell(5, 7).value = tcom_col
        ws.cell(5, 8).value = hk_col
        ws.cell(5, 10).value = usd_mcap(market, "TCOM")
        ws.cell(5, 12).value = peer_col("700_XHKG")
        ws.cell(5, 13).value = f"={market['hkd_market_cap']['700_XHKG']}/{fx}"
        ws.cell(5, 14).value = peer_col("BABA")
        ws.cell(5, 15).value = peer_col("9988_XHKG")
        ws.cell(5, 17).value = usd_mcap(market, "BABA")
        ws.cell(5, 18).value = peer_col("3690_XHKG")
        ws.cell(5, 19).value = f"={market['hkd_market_cap']['3690_XHKG']}/{fx}"
        ws.cell(5, 20).value = peer_col("NTES")
        ws.cell(5, 21).value = peer_col("9999_XHKG")
        ws.cell(5, 23).value = usd_mcap(market, "NTES")
        ws.cell(5, 24).value = peer_col("PDD")
        ws.cell(5, 25).value = usd_mcap(market, "PDD")
        ws.cell(5, 26).value = peer_col("JD")
        ws.cell(5, 27).value = peer_col("9618_XHKG")
        ws.cell(5, 29).value = usd_mcap(market, "JD")
        ws.cell(5, 30).value = peer_col("BIDU")
        ws.cell(5, 31).value = peer_col("9888_XHKG")
        ws.cell(5, 33).value = usd_mcap(market, "BIDU")
        ws.cell(5, 34).value = peer_col("ABNB")
        ws.cell(5, 35).value = usd_mcap(market, "ABNB")
        ws.cell(5, 36).value = peer_col("BKNG")
        ws.cell(5, 37).value = usd_mcap(market, "BKNG")
        ws.cell(5, 38).value = peer_col("EXPE")
        ws.cell(5, 39).value = usd_mcap(market, "EXPE")
        ws.cell(5, 40).value = peer_col("MMYT")
        ws.cell(5, 41).value = usd_mcap(market, "MMYT")
        ws.cell(5, 42).value = peer_col("TRIP")
        ws.cell(5, 43).value = usd_mcap(market, "TRIP")
        ws.cell(5, 44).value = peer_col("TRVG")
        ws.cell(5, 45).value = usd_mcap(market, "TRVG")
        ws.cell(5, 46).value = peer_col("780_XHKG")
        ws.cell(5, 47).value = f"={market['hkd_market_cap']['780_XHKG']}/{fx}"

        # keep header rows 6-7 from template; rewrite names from row 8
        # find last used name row
        last_old = 8
        for r in range(8, ws.max_row + 1):
            if ws.cell(r, 3).value:
                last_old = r
        n = len(names)
        if n:
            for i in range(n):
                r = 8 + i
                ws.cell(r, 2).value = i + 1
                ws.cell(r, 3).value = names[i][0]
                paint_index_row(ws, r, range(3, 5), names[i][0], index_names)
                ws.cell(r, 4).value = f'=VLOOKUP($C{r},{da_range},D$5,FALSE)'
                ws.cell(r, 5).value = f'=VLOOKUP($C{r},{da_range},E$5,FALSE)/1000000000'
                ws.cell(r, 6).value = f'=VLOOKUP($C{r},{da_range},F$5,FALSE)'
                ws.cell(r, 7).value = (
                    f'=IF(VLOOKUP($C{r},{da_range},G$5,FALSE)="N/A",0,'
                    f'IFERROR(VLOOKUP($C{r},{da_range},G$5,FALSE),0))'
                )
                ws.cell(r, 8).value = (
                    f'=IF(VLOOKUP($C{r},{da_range},H$5,FALSE)="N/A",0,'
                    f'IFERROR(VLOOKUP($C{r},{da_range},H$5,FALSE),0))'
                )
                ws.cell(r, 9).value = f"=G{r}+H{r}"
                ws.cell(r, 10).value = f"=I{r}/$J$5"
                ws.cell(r, 11).value = (
                    f'=IFERROR(INDEX(\'SH Summary\'!B:B,MATCH(\'{sheet_name}\'!C{r},'
                    f'\'SH Summary\'!C:C,0)),"N/A")'
                )
                ws.cell(r, 12).value = f"=VLOOKUP($C{r},{da_range},L$5,FALSE)"
                ws.cell(r, 13).value = f"=L{r}/M$5"
                ws.cell(r, 14).value = f"=VLOOKUP($C{r},{da_range},N$5,FALSE)"
                ws.cell(r, 15).value = f"=VLOOKUP($C{r},{da_range},O$5,FALSE)"
                ws.cell(r, 16).value = f"=N{r}+O{r}"
                ws.cell(r, 17).value = f"=P{r}/Q$5"
                ws.cell(r, 18).value = f"=VLOOKUP($C{r},{da_range},R$5,FALSE)"
                ws.cell(r, 19).value = f"=R{r}/S$5"
                ws.cell(r, 20).value = f"=VLOOKUP($C{r},{da_range},T$5,FALSE)"
                ws.cell(r, 21).value = f"=VLOOKUP($C{r},{da_range},U$5,FALSE)"
                ws.cell(r, 22).value = f"=T{r}+U{r}"
                ws.cell(r, 23).value = f"=V{r}/W$5"
                ws.cell(r, 24).value = f"=VLOOKUP($C{r},{da_range},X$5,FALSE)"
                ws.cell(r, 25).value = f"=X{r}/Y$5"
                ws.cell(r, 26).value = f"=VLOOKUP($C{r},{da_range},Z$5,FALSE)"
                ws.cell(r, 27).value = f"=VLOOKUP($C{r},{da_range},AA$5,FALSE)"
                ws.cell(r, 28).value = f"=Z{r}+AA{r}"
                ws.cell(r, 29).value = f"=AB{r}/AC$5"
                ws.cell(r, 30).value = f"=VLOOKUP($C{r},{da_range},AD$5,FALSE)"
                ws.cell(r, 31).value = f"=VLOOKUP($C{r},{da_range},AE$5,FALSE)"
                ws.cell(r, 32).value = f"=AD{r}+AE{r}"
                ws.cell(r, 33).value = f"=AF{r}/AG$5"
                ws.cell(r, 34).value = f"=VLOOKUP($C{r},{da_range},AH$5,FALSE)"
                ws.cell(r, 35).value = f"=AH{r}/AI$5"
                ws.cell(r, 36).value = f"=VLOOKUP($C{r},{da_range},AJ$5,FALSE)"
                ws.cell(r, 37).value = f"=AJ{r}/AK$5"
                ws.cell(r, 38).value = f"=VLOOKUP($C{r},{da_range},AL$5,FALSE)"
                ws.cell(r, 39).value = f"=AL{r}/AM$5"
                ws.cell(r, 40).value = f"=VLOOKUP($C{r},{da_range},AN$5,FALSE)"
                ws.cell(r, 41).value = f"=AN{r}/AO$5"
                ws.cell(r, 42).value = f"=VLOOKUP($C{r},{da_range},AP$5,FALSE)"
                ws.cell(r, 43).value = f"=AP{r}/AQ$5"
                ws.cell(r, 44).value = f"=VLOOKUP($C{r},{da_range},AR$5,FALSE)"
                ws.cell(r, 45).value = f"=AR{r}/AS$5"
                ws.cell(r, 46).value = f"=VLOOKUP($C{r},{da_range},AT$5,FALSE)"
                ws.cell(r, 47).value = f"=AT{r}/AU$5"
                for c in REGION_PCT_COLS:
                    ws.cell(r, c).number_format = "0.00%"
        extra_start = 8 + n
        if last_old >= extra_start:
            for r in range(extra_start, last_old + 1):
                for c in range(2, 48):
                    ws.cell(r, c).value = None
                    if c in (3, 4):
                        ws.cell(r, c).fill = NO_FILL
        ws.freeze_panes = "D8"
    return {
        "counts": {k: len(v) for k, v in buckets.items()},
        "unmapped": unmapped,
    }


def write_map(
    ws: Worksheet,
    headers: list[str],
    rows: list[list[Any]],
    da: dict[str, int],
    pmap: dict[str, int],
    market: dict[str, Any],
) -> None:
    aum_i = pmap["Equity Assets USD, mm"] - 1
    tcom_i = pmap["Value TCOM USD, mm"] - 1
    hk_i = pmap["Value 9961_XHKG USD, mm"] - 1

    def has_any(row: list[Any], tickers: list[str]) -> bool:
        for t in tickers:
            key = f"Value {t} USD, mm"
            if key in pmap and num(row[pmap[key] - 1]) != 0:
                return True
        return False

    no_adr = no_travel = no_tcom = 0.0
    for row in rows:
        aum = num(row[aum_i])
        if not has_any(row, ADR_VALUE_TICKERS):
            no_adr += aum
        if not has_any(row, TRAVEL_VALUE_TICKERS):
            no_travel += aum
        if num(row[tcom_i]) == 0 and num(row[hk_i]) == 0:
            no_tcom += aum

    sum_row = da["sum_row"]
    tcom_l = get_column_letter(pmap["Value TCOM USD, mm"])
    hk_l = get_column_letter(pmap["Value 9961_XHKG USD, mm"])
    aum_l = get_column_letter(pmap["Equity Assets USD, mm"])
    adr_sum = "+".join(
        f"DATA_ALL!{get_column_letter(pmap[f'Value {t} USD, mm'])}{sum_row}"
        for t in ADR_VALUE_TICKERS
        if f"Value {t} USD, mm" in pmap
    )
    travel_sum = "+".join(
        f"DATA_ALL!{get_column_letter(pmap[f'Value {t} USD, mm'])}{sum_row}"
        for t in TRAVEL_VALUE_TICKERS
        if f"Value {t} USD, mm" in pmap
    )

    ws["B2"] = "全球投资人地图"
    ws["B3"] = f"updated as of {WORKBOOK_AS_OF}"
    ws["B5"] = "# of Institutions"
    ws["C5"] = da["n_inst"]
    ws["B6"] = "Total SH Equity AUM ($bn)"
    ws["C6"] = f"=DATA_ALL!{aum_l}{sum_row}/1000000000"
    ws["B7"] = "TCOM SH Equity AUM ($bn)"
    ws["C7"] = "=(C6*1000000000-D18)/1000000000"
    ws["B8"] = "Total TCOM Holding ($bn)"
    ws["C8"] = f"=(DATA_ALL!{tcom_l}{sum_row}+DATA_ALL!{hk_l}{sum_row})/1000000000"
    ws["D8"] = MAP_D8
    ws["B9"] = "Holding %"
    ws["C9"] = "=C8/C7"
    ws["C13"] = f"=({adr_sum})/1000000000"
    ws["C14"] = f"=({travel_sum})/1000000000"
    ws["C12"] = "=C13+C14"
    ws["D12"] = "=C6"
    ws["E12"] = "=C12/D12"
    ws["F12"] = "=E12/14"
    ws["D13"] = "=(D12*1000000000-D16)/1000000000"
    ws["E13"] = "=C13/D13"
    ws["F13"] = "=E13/7"
    ws["D14"] = "=(D12*1000000000-D17)/1000000000"
    ws["E14"] = "=C14/D14"
    ws["F14"] = "=E14/7"
    ws["D16"] = no_adr
    ws["E16"] = "no ADR holding Equity AUM"
    ws["D17"] = no_travel
    ws["E17"] = "no Travel holding Equity AUM"
    ws["D18"] = no_tcom
    ws["E18"] = "no TCOM holding"


def write_market_sheet(wb: Workbook, market: dict[str, Any]) -> None:
    if "Market" in wb.sheetnames:
        del wb["Market"]
    ws = wb.create_sheet("Market")
    ws.sheet_state = "hidden"
    ws["A1"] = "valid_as_of"
    ws["B1"] = VALID_AS_OF.replace("/", "-")
    ws["A2"] = "USD/HKD"
    ws["B2"] = market["fx_usd_hkd"]
    ws["A3"] = "TCOM shares outstanding"
    ws["B3"] = market["tcom_shares_outstanding"]
    ws["C3"] = market["source"]
    ws["A4"] = "quotes_fetched"
    ws["B4"] = market.get("as_of")
    ws["A5"] = "ticker"
    ws["B5"] = "currency"
    ws["C5"] = "native_mcap"
    ws["D5"] = "usd_mcap"
    r = 6
    for t, v in market["usd_market_cap"].items():
        ws.cell(r, 1).value = t
        ws.cell(r, 2).value = "USD"
        ws.cell(r, 3).value = v
        ws.cell(r, 4).value = v
        r += 1
    fx = float(market["fx_usd_hkd"])
    for t, v in market["hkd_market_cap"].items():
        ws.cell(r, 1).value = t
        ws.cell(r, 2).value = "HKD"
        ws.cell(r, 3).value = v
        ws.cell(r, 4).value = v / fx
        r += 1
