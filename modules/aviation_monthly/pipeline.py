"""Official monthly CAAC/Big-3 -> Airline Data -> domestic industry workbook pipeline."""
from __future__ import annotations

import hashlib
import io
import json
import os
import re
import shutil
import tempfile
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from urllib.parse import urljoin, urlparse

import pdfplumber
import requests
from modules.industry_data.layout import month_number
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
CAAC_INDEX = "https://www.caac.gov.cn/XXGK/XXGK/TJSJ/index_1215.html"
CNINFO_SEARCH = "https://www.cninfo.com.cn/new/fulltextSearch/full"
AIRLINES = {
    "南航": {"code": "600029", "anchor": "载客人数", "table": "载客人数（千人次）", "row": 2},
    "东航": {"code": "600115", "anchor": "载运旅客人次", "table": "载运旅客人次（千）", "row": 3},
    "国航": {"code": "601111", "anchor": "乘客人数", "table": "乘客人数（千）", "row": 4},
}
ERROR_TOKENS = (b"#REF!", b"#DIV/0!", b"#VALUE!", b"#N/A", b"#NAME?")


class PipelineError(RuntimeError):
    def __init__(self, kind: str, message: str):
        super().__init__(message)
        self.kind = kind


def http_session() -> requests.Session:
    retry = Retry(total=3, connect=3, read=3, backoff_factor=0.8,
                  status_forcelist=(429, 500, 502, 503, 504), allowed_methods=("GET",))
    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"})
    adapter = HTTPAdapter(max_retries=retry)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    return session


def get(session: requests.Session, url: str, **kwargs) -> requests.Response:
    try:
        timeout = kwargs.pop("timeout", (20, 120))
        response = session.get(url, timeout=timeout, **kwargs)
        response.raise_for_status()
        return response
    except requests.RequestException as exc:
        raise PipelineError("source-missing", f"GET failed: {url}: {exc}") from exc


def pdf_pages(session: requests.Session, url: str) -> list[str]:
    payload = get(session, url, timeout=(30, 240)).content
    if not payload.startswith(b"%PDF-"):
        raise PipelineError("parser-drift", f"URL did not return a PDF: {url}")
    try:
        with pdfplumber.open(io.BytesIO(payload)) as pdf:
            pages = [(page.extract_text() or "") for page in pdf.pages]
    except Exception as exc:
        raise PipelineError("parser-drift", f"Cannot parse PDF {url}: {exc}") from exc
    if not any(text.strip() for text in pages):
        raise PipelineError("parser-drift", f"PDF has no extractable text: {url}")
    return pages


def number(value: str) -> float:
    return float(value.replace(",", ""))


def first_match(pages: list[str], pattern: str, label: str) -> tuple[float, int, str]:
    regex = re.compile(pattern)
    for page_no, text in enumerate(pages, 1):
        match = regex.search(text)
        if match:
            line = next((line.strip() for line in text.splitlines() if regex.search(line)), label)
            return number(match.group(1)), page_no, line
    raise PipelineError("parser-drift", f"Missing PDF row: {label}")


def fetch_caac(session: requests.Session, year: int, month: int) -> dict:
    index = get(session, CAAC_INDEX)
    index.encoding = "utf-8"
    items = re.findall(r'href="([^"]+)"[^>]*>\s*中国民航(\d{4})年(\d{1,2})月份主要生产指标统计', index.text)
    chosen = next((href for href, y, m in items if int(y) == year and int(m) == month), None)
    if not chosen:
        raise PipelineError("source-missing", f"CAAC has not published {year}-{month:02d}")
    page_url = urljoin(CAAC_INDEX, chosen)
    page = get(session, page_url)
    page.encoding = "utf-8"
    attachment = re.search(r'href="([^"]*P\d+\.pdf)"', page.text, re.I)
    if not attachment:
        raise PipelineError("parser-drift", f"CAAC attachment PDF not found: {page_url}")
    pdf_url = urljoin(page_url, attachment.group(1))
    pages = pdf_pages(session, pdf_url)
    specs = {
        "total": (r"旅客运输量\s*万人\s+([\d,.]+)", "旅客运输量·合计"),
        "domestic_regional": (r"国内航线\s*万人\s+([\d,.]+)", "国内航线"),
        "regional": (r"港澳台航线\s*万人\s+([\d,.]+)", "其中：港澳台航线"),
        "international": (r"国际航线\s*万人\s+([\d,.]+)", "国际航线"),
    }
    values, provenance = {}, {}
    for key, (pattern, label) in specs.items():
        raw, page_no, row_text = first_match(pages, pattern, label)
        values[key] = round(raw * 10, 1)
        provenance[key] = {"page": page_no, "row": label, "extracted_row": row_text}
    return {"title": f"中国民航{year}年{month}月份主要生产指标统计", "period": f"{year}年{month}月",
            "page_url": page_url, "url": pdf_url, "published": None, "values": values, "provenance": provenance}


def find_announcement(session: requests.Session, name: str, year: int, month: int) -> dict:
    meta = AIRLINES[name]
    response = get(session, CNINFO_SEARCH, params={"searchkey": f"{meta['code']} 运营数据", "pageNum": 1, "pageSize": 30},
                   headers={"X-Requested-With": "XMLHttpRequest"})
    try:
        announcements = response.json().get("announcements") or []
    except ValueError as exc:
        raise PipelineError("parser-drift", "CNINFO search returned non-JSON content") from exc
    for item in announcements:
        title = re.sub(r"<[^>]+>", "", item.get("announcementTitle", ""))
        if str(item.get("secCode")) != meta["code"] or "运营数据" not in title:
            continue
        if f"{year}年{month}月" not in title:
            continue
        timestamp = item.get("announcementTime")
        published = datetime.fromtimestamp(timestamp / 1000).date().isoformat() if timestamp else None
        return {"title": title, "url": urljoin("https://static.cninfo.com.cn/", item["adjunctUrl"]),
                "published": published}
    raise PipelineError("source-missing", f"{name} has not published {year}-{month:02d} operating data")


def numeric_tokens(line: str) -> list[float]:
    return [number(token) for token in re.findall(r"(?<!\d)(\d[\d,]*(?:\.\d+)?)", line)]


#: 分项行的前缀写法：`其中: 国内航线`、`－国内航线`、`国内航线`
ROUTE_LINE = re.compile(r"(?:其中[:：])?\s*[－\-]?\s*(国内|国际|地区)(?:航线)?\s+([\d,]+(?:\.\d+)?)")


def total_on_anchor_line(anchor: str, line: str) -> float | None:
    """从「指标名行」本身取当月总量。

    三家公告格式不同：
      - 南航：有独立的 `合计` 行
      - 东航：`载运旅客人次（千） 14,267.34 3.98% …` —— 总量就在指标名那一行
      - 国航：`4、乘客人数(千) 15,690.8 8.5 …` —— 同上，且带 `4、` 序号前缀

    只取**锚点之后**的第一个数字，从而绕开 `4、` 这类序号；并排除分项行。
    这不是「推算总量」——该行就是公告里的官方总量行。且管道随后会用
    `Total ≈ Domestic + Intl + Regional`（容差 0.05 千人次）独立勾稽，取错会被拦下。
    """
    if anchor not in line:
        return None
    if ROUTE_LINE.match(line.strip()):
        return None
    tail = line[line.index(anchor) + len(anchor) :]
    match = re.search(r"(?<!\d)(\d[\d,]*(?:\.\d+)?)", tail)
    return number(match.group(1)) if match else None


def parse_airline_pages(name: str, pages: list[str]) -> tuple[dict, dict]:
    """按公告里显式存在的行取值；不从「附近的某个数」推算总量。"""
    anchor = AIRLINES[name]["anchor"]
    start = next(((page, line_no) for page, text in enumerate(pages)
                  for line_no, line in enumerate(text.splitlines()) if anchor in line), None)
    if start is None:
        raise PipelineError("parser-drift", f"{name}: passenger anchor not found: {anchor}")
    start_page, start_line = start
    values, provenance = {}, {}
    route_names = {"国内": "domestic", "国际": "international", "地区": "regional"}
    # 总量有两种来源：独立「合计」行（南航）与指标名行本身（东航/国航）。
    # 两者都收，最后**优先用合计行**——这样南航的既有行为一个字节都不变。
    anchor_total: tuple[float, dict] | None = None
    scanned = 0
    for page_index in range(start_page, min(len(pages), start_page + 2)):
        for line_index, raw_line in enumerate(pages[page_index].splitlines()):
            if page_index == start_page and line_index < start_line:
                continue
            line = raw_line.strip()
            scanned += 1
            route = ROUTE_LINE.match(line)
            if route:
                key = route_names[route.group(1)]
                values.setdefault(key, number(route.group(2)))
                provenance.setdefault(key, {"page": page_index + 1,
                                            "row": f"{AIRLINES[name]['table']}—{route.group(1)}航线",
                                            "extracted_row": line})
            total = re.search(r"(?:^|\s)(?:合计|总计)\s*[:：]?\s+([\d,]+(?:\.\d+)?)", line)
            if total and "total" not in values:
                values["total"] = number(total.group(1))
                provenance["total"] = {"page": page_index + 1,
                                         "row": f"{AIRLINES[name]['table']}—合计",
                                         "extracted_row": line}
            if anchor_total is None:
                candidate = total_on_anchor_line(anchor, line)
                if candidate is not None:
                    anchor_total = (
                        candidate,
                        {
                            "page": page_index + 1,
                            "row": f"{AIRLINES[name]['table']}—指标名行（总量）",
                            "extracted_row": line,
                        },
                    )
            if all(key in values for key in ("total", "domestic", "international", "regional")):
                return values, provenance
            if scanned >= 50:
                break
        if scanned >= 50:
            break

    if "total" not in values and anchor_total is not None:
        values["total"], provenance["total"] = anchor_total

    missing = [key for key in ("total", "domestic", "international", "regional") if key not in values]
    if missing:
        raise PipelineError("parser-drift", f"{name}: missing explicit passenger rows: {', '.join(missing)}")
    return values, provenance


def fetch_airline(session: requests.Session, name: str, year: int, month: int) -> dict:
    announcement = find_announcement(session, name, year, month)
    pages = pdf_pages(session, announcement["url"])
    values, provenance = parse_airline_pages(name, pages)
    return {**announcement, "period": f"{year}年{month}月", "values": values, "provenance": provenance,
            "code": AIRLINES[name]["code"], "table": AIRLINES[name]["table"]}


def check(condition: bool, name: str, detail: str, checks: list[dict], kind: str = "validation-error") -> None:
    checks.append({"name": name, "passed": bool(condition), "detail": detail})
    if not condition:
        raise PipelineError(kind, f"{name}: {detail}")


def validate_sources(caac: dict, airlines: dict, checks: list[dict]) -> None:
    cv = caac["values"]
    check(all(value > 0 for value in cv.values()), "caac-positive", str(cv), checks)
    delta = cv["total"] - cv["domestic_regional"] - cv["international"]
    check(abs(delta) <= 2.1, "caac-total-tie", f"difference={delta:.3f} thousand", checks, "source-conflict")
    for name, source in airlines.items():
        values = source["values"]
        check(all(value > 0 for value in values.values()), f"{name}-positive", str(values), checks)
        parts = values["domestic"] + values["international"] + values["regional"]
        delta = values["total"] - parts
        check(abs(delta) <= 0.05, f"{name}-total-tie", f"total-parts={delta:.3f} thousand", checks,
              "source-conflict")
        check(urlparse(source["url"]).hostname.endswith("cninfo.com.cn"), f"{name}-official-domain",
              source["url"], checks)
    check(urlparse(caac["url"]).hostname.endswith("caac.gov.cn"), "caac-official-domain", caac["url"], checks)


def formula_error_counts(path: Path) -> dict[str, int]:
    counts = {token.decode(): 0 for token in ERROR_TOKENS}
    with zipfile.ZipFile(path) as archive:
        sheets = [name for name in archive.namelist() if name.startswith("xl/worksheets/") and name.endswith(".xml")]
        for sheet in sheets:
            payload = archive.read(sheet)
            for token in ERROR_TOKENS:
                counts[token.decode()] += payload.count(token)
    return counts


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def ensure_workbook_contract(path: Path, year: int, month: int, checks: list[dict]) -> None:
    if not path.exists():
        raise PipelineError("workbook-contract-drift", f"Workbook not found: {path}")
    check(year == 2026, "supported-contract-year", f"requested={year}, supported=2026", checks,
          "workbook-contract-drift")
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        required = {"CAAC Data", "Top4 Domestic", "Top4 Intl.+Reg", "Top 4 Total", "Summary"}
        check(required.issubset(workbook.sheetnames), "airline-sheets", str(workbook.sheetnames), checks,
              "workbook-contract-drift")
        expected_col = month + 1
        expected_month = MONTHS[month - 1].lower()
        for sheet in ("CAAC Data", "Top4 Domestic", "Top4 Intl.+Reg", "Top 4 Total"):
            ws = workbook[sheet]
            matches = []
            for row in range(1, 4):
                header = str(ws.cell(row, expected_col).value or "").strip().lower()
                year_label = str(ws.cell(row, 1).value or "")
                if header.startswith(expected_month[:3]) and str(year) in year_label:
                    matches.append(row)
            check(len(matches) == 1, f"{sheet}-period-header",
                  f"year={year}, month={MONTHS[month - 1]}, matches={matches}", checks,
                  "workbook-contract-drift")

        def normalized(value) -> str:
            return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())

        def require_rows(sheet: str, expected: dict[int, str | tuple[str, ...]]) -> None:
            ws = workbook[sheet]
            drift = []
            for row, expected_value in expected.items():
                fragments = expected_value if isinstance(expected_value, tuple) else (expected_value,)
                actual = normalized(ws.cell(row, 1).value)
                if not any(normalized(fragment) in actual for fragment in fragments):
                    drift.append({"row": row, "expected": fragments, "actual": ws.cell(row, 1).value})
            check(not drift, f"{sheet}-fixed-row-labels", json.dumps(drift, ensure_ascii=False), checks,
                  "workbook-contract-drift")

        yoy = f"YoY vs {year - 1}"
        require_rows("CAAC Data", {3: "Total", 4: yoy, 7: "Domestic", 8: yoy,
                                    11: "Domestic+Regional", 12: yoy, 14: "Intl + Regional", 15: yoy,
                                    19: "Pure Intl", 20: yoy, 23: "Regional", 24: yoy,
                                    29: "Total", 33: "Domestic", 37: "Domestic+Regional",
                                    40: "Intl + Regional", 45: "Pure Intl", 49: "Regional"})
        carrier_rows = {2: "CSA", 3: "CEA", 4: "AC", 5: "Total", 8: "CSA", 9: "CEA",
                        10: "AC", 11: "Total", 14: "CSA", 15: "CEA", 16: "AC", 17: "Total"}
        require_rows("Top4 Domestic", carrier_rows)
        require_rows("Top 4 Total", carrier_rows)
        intl_rows = {}
        for start in (2, 8, 21, 27, 40, 46, 58, 77, 96):
            intl_rows.update({start: "CSA", start + 1: "CEA", start + 2: "AC", start + 3: "Total"})
        require_rows("Top4 Intl.+Reg", intl_rows)
        require_rows("Summary", {4: "Air Industry", 5: "Air Industry Domestic",
                                  6: ("Air Industry Intl", "Air Industry International"),
                                  7: "Air Industry Regional", 8: "Air Industry Pure Intl",
                                  14: "Top 3", 15: "Top 3 Domestic",
                                  16: ("Top 3 Intl", "Top 3 International"),
                                  17: "Top 3 Regional", 18: "Top 3 Pure Intl"})
        summary_col = summary_month_column(workbook, year, month)
        check(summary_col is not None, "summary-period-header", f"year={year}, month={MONTHS[month - 1]}", checks,
              "workbook-contract-drift")
    finally:
        workbook.close()


def summary_month_column(workbook, year: int, month: int) -> int | None:
    ws = workbook["Summary"]
    markers = [cell.column for cell in ws[2] if str(cell.value or "").strip() == str(year)]
    expected = MONTHS[month - 1].lower()
    for marker in markers:
        column = marker + month - 1
        if str(ws.cell(3, column).value or "").strip().lower().startswith(expected[:3]):
            return column
    return None


def clean_text(value) -> str:
    return re.sub(r"\s+", "", str(value or ""))


def industry_targets(path: Path, year: int, month: int) -> dict[str, str]:
    if not path.exists():
        raise PipelineError("workbook-contract-drift", f"Workbook not found: {path}")
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        if "国内行业数据" not in workbook.sheetnames:
            raise PipelineError("workbook-contract-drift", "Missing sheet: 国内行业数据")
        ws = workbook["国内行业数据"]
        year_rows = [row for row in range(1, ws.max_row + 1) if clean_text(ws.cell(row, 2).value) == f"{year}年"]
        if len(year_rows) != 1:
            raise PipelineError("workbook-contract-drift", f"Expected one {year}年 marker in column B, found {year_rows}")
        start = year_rows[0]
        end = next((row for row in range(start + 1, ws.max_row + 1)
                    if re.fullmatch(r"20\d{2}年", clean_text(ws.cell(row, 2).value))), ws.max_row + 1)
        header = next((row for row in range(start + 1, end) if clean_text(ws.cell(row, 2).value) == "月度"), None)
        # 月份行标签可能带后缀（底稿里 7 月写作「7月 (preliminary)」）。
        # 匹配规则由指标底稿契约统一提供——这条管道写的就是那张表，必须服从同一契约。
        month_row = next(
            (
                row
                for row in range((header or start) + 1, end)
                if month_number(ws.cell(row, 2).value) == month
            ),
            None,
        )
        if header is None or month_row is None:
            raise PipelineError("workbook-contract-drift", f"Cannot locate {year}年{month}月 monthly row")

        def group_span(label: str) -> tuple[int, int]:
            target = clean_text(label)
            for row in range(start, header + 1):
                for cell in ws[row]:
                    if clean_text(cell.value) != target:
                        continue
                    for merged in ws.merged_cells.ranges:
                        if merged.min_row <= row <= merged.max_row and merged.min_col <= cell.column <= merged.max_col:
                            return merged.min_col, merged.max_col
                    return cell.column, cell.column
            raise PipelineError("workbook-contract-drift", f"Missing industry group: {label}")

        def subcolumns(group: str) -> tuple[int, int]:
            left, right = group_span(group)
            caac_col = next((col for col in range(left, right + 1)
                             if clean_text(ws.cell(header, col).value).startswith("民航局")), None)
            big3_col = next((col for col in range(left, right + 1)
                             if clean_text(ws.cell(header, col).value).startswith("三大航")), None)
            if caac_col is None or big3_col is None:
                raise PipelineError("workbook-contract-drift", f"Missing 民航局/三大航 subcolumns under {group}")
            return caac_col, big3_col

        domestic = subcolumns("国内航空客运量")
        international = subcolumns("国际航空客运量（含港澳台）")
        return {
            "caac_domestic": f"{get_column_letter(domestic[0])}{month_row}",
            "big3_domestic": f"{get_column_letter(domestic[1])}{month_row}",
            "caac_intl_regional": f"{get_column_letter(international[0])}{month_row}",
            "big3_intl_regional": f"{get_column_letter(international[1])}{month_row}",
        }
    finally:
        workbook.close()


def ensure_industry_contract(path: Path, year: int, month: int, checks: list[dict]) -> dict[str, str]:
    targets = industry_targets(path, year, month)
    check(len(set(targets.values())) == 4, "industry-targets-unique", str(targets), checks,
          "workbook-contract-drift")
    return targets


def input_assignments(caac: dict, airlines: dict, month: int) -> list[dict]:
    col = get_column_letter(month + 1)
    entries = []

    def add(sheet: str, row: int, value: float, source: dict, key: str, metric: str, unit: str) -> None:
        provenance = source["provenance"][key]
        period = f"数据期{source['period']}"
        release = f"，公告日{source['published']}" if source.get("published") else ""
        table = "旅客运输量（万人）" if source is caac else source["table"]
        comment = (f"Source: {source['title']}, {period}{release}, PDF第{provenance['page']}页"
                   f"「{table}」表「{provenance['row']}」行, {source['url']}")
        entries.append({"sheet": sheet, "cell": f"{col}{row}", "value": value, "metric": metric,
                        "unit": unit, "title": source["title"], "period": source["period"],
                        "published": source.get("published"),
                        "url": source["url"], "page": provenance["page"], "row": provenance["row"],
                        "extracted_row": provenance["extracted_row"], "comment": comment})

    add("CAAC Data", 3, caac["values"]["total"], caac, "total", "民航局合计", "千人")
    add("CAAC Data", 11, caac["values"]["domestic_regional"], caac, "domestic_regional",
        "民航局国内含地区", "千人")
    add("CAAC Data", 19, caac["values"]["international"], caac, "international", "民航局纯国际", "千人")
    add("CAAC Data", 23, caac["values"]["regional"], caac, "regional", "民航局地区", "千人")
    for name in ("南航", "东航", "国航"):
        source = airlines[name]
        row = AIRLINES[name]["row"]
        add("Top 4 Total", row, source["values"]["total"], source, "total", f"{name}合计", "千人次")
        add("Top4 Domestic", row, source["values"]["domestic"], source, "domestic", f"{name}国内", "千人次")
        add("Top4 Intl.+Reg", row + 19, source["values"]["international"], source, "international",
            f"{name}纯国际", "千人次")
        add("Top4 Intl.+Reg", row + 38, source["values"]["regional"], source, "regional",
            f"{name}地区", "千人次")
    return entries


def formula_assignments(month: int, summary_col: int) -> dict[tuple[str, str], str]:
    col = get_column_letter(month + 1)
    formulas: dict[tuple[str, str], str] = {}

    def put(sheet: str, row: int, formula: str) -> None:
        formulas[(sheet, f"{col}{row}")] = formula

    put("CAAC Data", 7, f"={col}11-{col}23")
    put("CAAC Data", 14, f"={col}19+{col}23")
    for row, denominator in ((4, 29), (8, 33), (12, 37), (15, 40), (20, 45), (24, 49)):
        put("CAAC Data", row, f"={col}{row - 1}/{col}{denominator}-1")
    for sheet, denominator_start in (("Top4 Domestic", 14), ("Top 4 Total", 14)):
        put(sheet, 5, f"=SUM({col}2:{col}4)")
        for row in range(8, 12):
            numerator = row - 6 if row < 11 else 5
            put(sheet, row, f"={col}{numerator}/{col}{denominator_start + row - 8}-1")
    for row in range(2, 5):
        put("Top4 Intl.+Reg", row, f"={col}{row + 19}+{col}{row + 38}")
    for total_row, source_start, source_end in ((5, 2, 4), (24, 21, 23), (43, 40, 42)):
        put("Top4 Intl.+Reg", total_row, f"=SUM({col}{source_start}:{col}{source_end})")
    for output_start, numerator_start, total_row, denominator_start in ((8, 2, 5, 58), (27, 21, 24, 77),
                                                                        (46, 40, 43, 96)):
        for offset in range(4):
            numerator = numerator_start + offset if offset < 3 else total_row
            put("Top4 Intl.+Reg", output_start + offset,
                f"={col}{numerator}/{col}{denominator_start + offset}-1")
    summary_letter = get_column_letter(summary_col)
    links = {
        4: ("CAAC Data", 4), 5: ("CAAC Data", 8), 6: ("CAAC Data", 15),
        7: ("CAAC Data", 24), 8: ("CAAC Data", 20), 14: ("Top 4 Total", 11),
        15: ("Top4 Domestic", 11), 16: ("Top4 Intl.+Reg", 11),
        17: ("Top4 Intl.+Reg", 49), 18: ("Top4 Intl.+Reg", 30),
    }
    for row, (sheet, source_row) in links.items():
        formulas[("Summary", f"{summary_letter}{row}")] = f"='{sheet}'!{col}{source_row}"
    return formulas


def numeric_cell(ws, row: int, column: int, label: str) -> float:
    value = ws.cell(row, column).value
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise PipelineError("workbook-contract-drift", f"{label} is not a cached number: {value!r}")
    return float(value)


def independent_results(path: Path, caac: dict, airlines: dict, month: int) -> dict[str, float]:
    col = month + 1
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        caac_domestic_base = numeric_cell(workbook["CAAC Data"], 33, col, "CAAC domestic prior-year base")
        caac_intl_base = numeric_cell(workbook["CAAC Data"], 40, col, "CAAC intl+regional prior-year base")
        big3_domestic_base = numeric_cell(workbook["Top4 Domestic"], 17, col, "Big3 domestic prior-year base")
        big3_intl_base = numeric_cell(workbook["Top4 Intl.+Reg"], 61, col, "Big3 intl+regional prior-year base")
    finally:
        workbook.close()
    current_caac_domestic = caac["values"]["domestic_regional"] - caac["values"]["regional"]
    current_caac_intl = caac["values"]["international"] + caac["values"]["regional"]
    current_big3_domestic = sum(source["values"]["domestic"] for source in airlines.values())
    current_big3_intl = sum(source["values"]["international"] + source["values"]["regional"]
                            for source in airlines.values())
    bases = (caac_domestic_base, caac_intl_base, big3_domestic_base, big3_intl_base)
    if any(base <= 0 for base in bases):
        raise PipelineError("workbook-contract-drift", f"Non-positive prior-year base: {bases}")
    return {
        "caac_domestic": current_caac_domestic / caac_domestic_base - 1,
        "caac_intl_regional": current_caac_intl / caac_intl_base - 1,
        "big3_domestic": current_big3_domestic / big3_domestic_base - 1,
        "big3_intl_regional": current_big3_intl / big3_intl_base - 1,
    }


def values_equal(left, right, tolerance: float = 1e-9) -> bool:
    return isinstance(left, (int, float)) and not isinstance(left, bool) and abs(float(left) - float(right)) <= tolerance


def ensure_no_value_conflicts(path: Path, sheet: str, assignments: list[tuple[str, float]], checks: list[dict],
                              label: str) -> None:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        ws = workbook[sheet]
        conflicts = []
        for address, new_value in assignments:
            old_value = ws[address].value
            if old_value is None or old_value == "" or values_equal(old_value, new_value):
                continue
            conflicts.append({"cell": f"{sheet}!{address}", "existing": old_value, "official": new_value})
        check(not conflicts, f"{label}-existing-values", json.dumps(conflicts, ensure_ascii=False), checks,
              "source-conflict")
    finally:
        workbook.close()


def preflight_conflicts(airline_path: Path, industry_path: Path, inputs: list[dict], targets: dict[str, str],
                        results: dict[str, float], checks: list[dict]) -> None:
    by_sheet: dict[str, list[tuple[str, float]]] = {}
    for entry in inputs:
        by_sheet.setdefault(entry["sheet"], []).append((entry["cell"], entry["value"]))
    for sheet, assignments in by_sheet.items():
        ensure_no_value_conflicts(airline_path, sheet, assignments, checks, f"airline-{sheet}")
    ensure_no_value_conflicts(industry_path, "国内行业数据",
                              [(targets[key], results[key]) for key in targets], checks, "industry")


def write_airline(path: Path, inputs: list[dict], formulas: dict[tuple[str, str], str]) -> None:
    workbook = load_workbook(path, data_only=False, keep_links=True)
    try:
        for entry in inputs:
            cell = workbook[entry["sheet"]][entry["cell"]]
            cell.value = entry["value"]
            cell.comment = Comment(entry["comment"], "Travel Pulse")
        for (sheet, address), formula in formulas.items():
            workbook[sheet][address] = formula
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
        workbook.save(path)
    finally:
        workbook.close()


def recalculate_excel(path: Path, writes: dict[str, dict[str, float]] | None = None) -> None:
    try:
        import pythoncom
        import win32com.client
    except ImportError as exc:
        raise PipelineError("recalc-unavailable", f"pywin32 is required: {exc}") from exc
    app = None
    book = None
    pythoncom.CoInitialize()
    try:
        app = win32com.client.DispatchEx("Excel.Application")
        app.Visible = False
        app.DisplayAlerts = False
        app.AskToUpdateLinks = False
        book = app.Workbooks.Open(str(path.resolve()), UpdateLinks=0, ReadOnly=False,
                                 IgnoreReadOnlyRecommended=True, AddToMru=False)
        for sheet, cells in (writes or {}).items():
            ws = book.Worksheets(sheet)
            for address, value in cells.items():
                ws.Range(address).Value = value
        app.CalculateFullRebuild()
        book.Save()
    except Exception as exc:
        raise PipelineError("recalc-unavailable", f"Excel full rebuild failed for {path}: {exc}") from exc
    finally:
        if book is not None:
            try:
                book.Close(SaveChanges=False)
            except Exception:
                pass
        if app is not None:
            try:
                app.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()


def cached_number(workbook, sheet: str, address: str) -> float:
    value = workbook[sheet][address].value
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise PipelineError("recalc-unavailable", f"Missing numeric formula cache at {sheet}!{address}: {value!r}")
    return float(value)


def validate_airline_output(path: Path, baseline_errors: dict[str, int], inputs: list[dict],
                            formulas: dict[tuple[str, str], str], expected: dict[str, float],
                            summary_col: int, month: int, checks: list[dict]) -> dict:
    col = get_column_letter(month + 1)
    formula_book = load_workbook(path, read_only=False, data_only=False)
    value_book = load_workbook(path, read_only=True, data_only=True)
    try:
        input_ok = all(values_equal(value_book[e["sheet"]][e["cell"]].value, e["value"]) for e in inputs)
        check(input_ok, "airline-input-roundtrip", "16 official inputs", checks)
        comment_ok = all((formula_book[e["sheet"]][e["cell"]].comment is not None and
                          formula_book[e["sheet"]][e["cell"]].comment.text == e["comment"]) for e in inputs)
        check(comment_ok, "airline-source-comments", "16 comments", checks)
        formula_ok = all(formula_book[sheet][address].value == formula
                         for (sheet, address), formula in formulas.items())
        check(formula_ok, "airline-formulas-preserved", f"{len(formulas)} formulas", checks)
        result_cells = {
            "caac_domestic": ("CAAC Data", f"{col}8"),
            "caac_intl_regional": ("CAAC Data", f"{col}15"),
            "big3_domestic": ("Top4 Domestic", f"{col}11"),
            "big3_intl_regional": ("Top4 Intl.+Reg", f"{col}11"),
        }
        cached = {key: cached_number(value_book, *location) for key, location in result_cells.items()}
        formula_match = all(abs(cached[key] - expected[key]) <= 1e-10 for key in expected)
        check(formula_match, "formula-vs-independent", json.dumps({
            key: {"formula": cached[key], "independent": expected[key], "difference": cached[key] - expected[key]}
            for key in expected}, ensure_ascii=False), checks)
        summary_letter = get_column_letter(summary_col)
        source_rows = {
            4: ("CAAC Data", 4), 5: ("CAAC Data", 8), 6: ("CAAC Data", 15),
            7: ("CAAC Data", 24), 8: ("CAAC Data", 20), 14: ("Top 4 Total", 11),
            15: ("Top4 Domestic", 11), 16: ("Top4 Intl.+Reg", 11),
            17: ("Top4 Intl.+Reg", 49), 18: ("Top4 Intl.+Reg", 30),
        }
        summary_values = {}
        summary_ok = True
        for row, (source_sheet, source_row) in source_rows.items():
            summary_value = cached_number(value_book, "Summary", f"{summary_letter}{row}")
            source_value = cached_number(value_book, source_sheet, f"{col}{source_row}")
            summary_values[f"Summary!{summary_letter}{row}"] = summary_value
            summary_ok = summary_ok and abs(summary_value - source_value) <= 1e-10
        check(summary_ok, "summary-month-ok", json.dumps(summary_values, ensure_ascii=False), checks)
    finally:
        formula_book.close()
        value_book.close()
    output_errors = formula_error_counts(path)
    errors_ok = all(output_errors[token] <= baseline_errors[token] for token in baseline_errors)
    check(errors_ok, "airline-error-count-not-increased",
          json.dumps({"before": baseline_errors, "after": output_errors}, ensure_ascii=False), checks)
    return {"formula_results": cached, "summary": summary_values, "errors_before": baseline_errors,
            "errors_after": output_errors, "formula_vs_independent": formula_match,
            "summary_month_ok": summary_ok, "error_count_not_increased": errors_ok}


def comment_count(path: Path) -> int:
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        return sum(1 for ws in workbook.worksheets for row in ws.iter_rows() for cell in row if cell.comment is not None)
    finally:
        workbook.close()


def validate_industry_output(path: Path, targets: dict[str, str], expected: dict[str, float],
                             baseline_errors: dict[str, int], baseline_comments: int,
                             checks: list[dict]) -> dict:
    workbook = load_workbook(path, read_only=False, data_only=True)
    try:
        ws = workbook["国内行业数据"]
        values = {key: ws[address].value for key, address in targets.items()}
        roundtrip = all(values_equal(values[key], expected[key], 1e-12) for key in expected)
        no_source_comments = all(not (ws[address].comment and ws[address].comment.text.startswith("Source:"))
                                 for address in targets.values())
    finally:
        workbook.close()
    check(roundtrip, "industry-roundtrip", json.dumps(values, ensure_ascii=False), checks)
    check(no_source_comments, "industry-no-source-comments", str(targets), checks)
    output_comments = comment_count(path)
    check(output_comments == baseline_comments, "industry-comment-count-unchanged",
          f"before={baseline_comments}, after={output_comments}", checks)
    output_errors = formula_error_counts(path)
    errors_ok = all(output_errors[token] <= baseline_errors[token] for token in baseline_errors)
    check(errors_ok, "industry-error-count-not-increased",
          json.dumps({"before": baseline_errors, "after": output_errors}, ensure_ascii=False), checks)
    return {"values": values, "targets": targets, "industry_roundtrip": roundtrip,
            "no_source_comments": no_source_comments, "comment_count_before": baseline_comments,
            "comment_count_after": output_comments, "errors_before": baseline_errors,
            "errors_after": output_errors, "error_count_not_increased": errors_ok}


def output_paths(args) -> tuple[Path, Path, Path]:
    output_dir = Path(args.output_dir).resolve() if args.output_dir else Path(args.airline_input).resolve().parent
    stamp = date.today().strftime("%Y%m%d")
    airline = Path(args.airline_output).resolve() if args.airline_output else output_dir / f"Airline Data_{stamp}.xlsx"
    industry = Path(args.industry_output).resolve() if args.industry_output else output_dir / f"国内行业数据_{stamp}.xlsx"
    manifest = Path(args.manifest_output).resolve() if args.manifest_output else output_dir / f"aviation_monthly_{args.year}{args.month:02d}_{stamp}.json"
    return airline, industry, manifest


def ensure_outputs_available(paths: tuple[Path, ...], overwrite: bool, checks: list[dict]) -> None:
    for path in paths:
        path.parent.mkdir(parents=True, exist_ok=True)
        check(overwrite or not path.exists(), f"output-available-{path.name}", str(path), checks,
              "workbook-contract-drift")


def install_outputs(staged: list[tuple[Path, Path]], overwrite: bool) -> None:
    installed: list[Path] = []
    backups: list[tuple[Path, Path]] = []
    try:
        for _, final in staged:
            if final.exists():
                if not overwrite:
                    raise PipelineError("workbook-contract-drift", f"Output already exists: {final}")
                backup = final.with_name(f".{final.name}.pipeline-backup")
                if backup.exists():
                    backup.unlink()
                os.replace(final, backup)
                backups.append((backup, final))
        for source, final in staged:
            os.replace(source, final)
            installed.append(final)
        for backup, _ in backups:
            backup.unlink(missing_ok=True)
    except Exception as exc:
        for final in reversed(installed):
            final.unlink(missing_ok=True)
        for backup, final in reversed(backups):
            if backup.exists():
                os.replace(backup, final)
        if isinstance(exc, PipelineError):
            raise
        raise PipelineError("workbook-contract-drift", f"Atomic output install failed: {exc}") from exc


def provenance_manifest(inputs: list[dict]) -> list[dict]:
    fields = ("metric", "value", "unit", "sheet", "cell", "title", "period", "published", "url", "page", "row",
              "extracted_row")
    return [{field: entry[field] for field in fields} for entry in inputs]


def run(args, checks: list[dict] | None = None) -> dict:
    checks = checks if checks is not None else []
    airline_input = Path(args.airline_input).resolve()
    industry_input = Path(args.industry_input).resolve()
    final_airline, final_industry, final_manifest = output_paths(args)
    ensure_outputs_available((final_airline, final_industry, final_manifest), args.overwrite, checks)
    ensure_workbook_contract(airline_input, args.year, args.month, checks)
    targets = ensure_industry_contract(industry_input, args.year, args.month, checks)
    session = http_session()
    caac = fetch_caac(session, args.year, args.month)
    airlines = {name: fetch_airline(session, name, args.year, args.month) for name in AIRLINES}
    validate_sources(caac, airlines, checks)
    inputs = input_assignments(caac, airlines, args.month)
    check(len(inputs) == 16, "official-input-count", f"count={len(inputs)}", checks, "parser-drift")
    expected = independent_results(airline_input, caac, airlines, args.month)
    preflight_conflicts(airline_input, industry_input, inputs, targets, expected, checks)
    source_hashes = {"airline": sha256(airline_input), "industry": sha256(industry_input)}
    baseline = {
        "airline_errors": formula_error_counts(airline_input),
        "industry_errors": formula_error_counts(industry_input),
        "industry_comments": comment_count(industry_input),
    }
    summary_book = load_workbook(airline_input, read_only=True, data_only=False)
    try:
        summary_col = summary_month_column(summary_book, args.year, args.month)
    finally:
        summary_book.close()
    if summary_col is None:
        raise PipelineError("workbook-contract-drift", "Summary period column disappeared after preflight")
    formulas = formula_assignments(args.month, summary_col)
    common = {
        "period": f"{args.year}年{args.month}月",
        "inputs": {"airline": str(airline_input), "industry": str(industry_input)},
        "input_sha256": source_hashes,
        "outputs": {"airline": str(final_airline), "industry": str(final_industry),
                    "manifest": str(final_manifest)},
        "official_values": provenance_manifest(inputs),
        "independent_results": expected,
        "checks": checks,
    }
    if not args.commit:
        return {"status": "dry-run-ok", **common}

    work_dir = Path(tempfile.mkdtemp(prefix="aviation-pipeline-", dir=str(final_airline.parent)))
    staged_airline = work_dir / "airline.xlsx"
    staged_industry = work_dir / "industry.xlsx"
    staged_manifest = work_dir / "manifest.json"
    try:
        shutil.copy2(airline_input, staged_airline)
        shutil.copy2(industry_input, staged_industry)
        write_airline(staged_airline, inputs, formulas)
        recalculate_excel(staged_airline)
        airline_validation = validate_airline_output(staged_airline, baseline["airline_errors"], inputs, formulas,
                                                     expected, summary_col, args.month, checks)
        industry_writes = {"国内行业数据": {address: expected[key] for key, address in targets.items()}}
        recalculate_excel(staged_industry, industry_writes)
        industry_validation = validate_industry_output(staged_industry, targets, expected,
                                                       baseline["industry_errors"], baseline["industry_comments"], checks)
        check({"airline": sha256(airline_input), "industry": sha256(industry_input)} == source_hashes,
              "source-workbooks-unchanged", json.dumps(source_hashes, ensure_ascii=False), checks,
              "workbook-contract-drift")
        manifest = {
            "status": "success", **common,
            "validation": {
                "formula_vs_independent": airline_validation["formula_vs_independent"],
                "summary_month_ok": airline_validation["summary_month_ok"],
                "industry_roundtrip": industry_validation["industry_roundtrip"],
                "error_count_not_increased": (airline_validation["error_count_not_increased"] and
                                               industry_validation["error_count_not_increased"]),
                "airline": airline_validation,
                "industry": industry_validation,
            },
            "sha256": {"airline": sha256(staged_airline), "industry": sha256(staged_industry)},
        }
        staged_manifest.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
        install_outputs([(staged_airline, final_airline), (staged_industry, final_industry),
                         (staged_manifest, final_manifest)], args.overwrite)
        return manifest
    finally:
        shutil.rmtree(work_dir, ignore_errors=True)


@dataclass
class Request:
    """`run()` 需要的参数。

    迁入工作台后**不再有独立的命令行入口**——原来的 `parser()` / `main()` 已删除，
    唯一入口是 `ir aviation ...`（见 cli.py）。工作簿路径也不再由调用方传，
    而是从 `ir config` 锁定的那两份解析（ADR 0001：不按文件名猜）。
    """

    year: int
    month: int
    airline_input: str
    industry_input: str
    commit: bool = False
    overwrite: bool = False
    output_dir: str | None = None
    airline_output: str | None = None
    industry_output: str | None = None
    manifest_output: str | None = None
