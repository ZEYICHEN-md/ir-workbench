"""Model 只读规划、Excel COM 副本写入与回读审计。"""
from __future__ import annotations

import re
import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .charts import audit_chart_sheets, update_chart_sheets
from .contracts import Contract
from .pdf_source import sha256
from .periods import Period, scan_periods, source_period

_SIMPLE_ARITH = re.compile(
    r"^=?-?\(?-?\d+(?:\.\d+)?(?:\s*[+\-*/]\s*-?\d+(?:\.\d+)?)+\)?$"
)


class ModelError(ValueError):
    pass


def _label(ws, row: int) -> str:
    parts = []
    for col in range(1, min(4, ws.max_column + 1)):
        value = ws.cell(row, col).value
        if value not in (None, ""):
            parts.append(str(value).strip())
    return " | ".join(parts)


def _is_formula(value) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _coerce_value(value):
    if isinstance(value, bool) or value is None:
        return value
    if isinstance(value, str) and value.startswith("="):
        body = value.strip()
        if _SIMPLE_ARITH.fullmatch(body.replace(" ", "")) or _SIMPLE_ARITH.fullmatch(body):
            try:
                return float(eval(body.lstrip("="), {"__builtins__": {}}, {}))  # noqa: S307
            except Exception:
                return value
    return value


def inspect_workbook(model: Path, contract: Contract) -> dict:
    wb = load_workbook(model, read_only=True, data_only=False, keep_links=True)
    try:
        missing = [spec["name"] for spec in contract.sheets if spec["name"] not in wb.sheetnames]
        missing += [sheet for sheet in contract.chart_sheets if sheet not in wb.sheetnames]
        sheets = []
        for spec in contract.sheets:
            if spec["name"] not in wb.sheetnames:
                continue
            ws = wb[spec["name"]]
            periods = scan_periods(ws, int(spec["header_row"]))
            last = max(periods, default=None)
            sheets.append({
                "name": spec["name"], "header_row": spec["header_row"],
                "kinds": spec.get("kinds", []),
                "last_period": None if last is None else last.key,
                "periods": {period.key: col for period, col in periods.items()},
            })
        return {"file": str(model), "sheets": sheets, "missing": missing}
    finally:
        wb.close()


def _column_has_content(ws, header_row: int, col: int) -> bool:
    for row in range(header_row + 1, (ws.max_row or 0) + 1):
        if ws.cell(row, col).value not in (None, ""):
            return True
    return False


def _sheet_plan(ws, spec: dict, target: Period) -> dict:
    periods = scan_periods(ws, int(spec["header_row"]))
    by_kind = spec.get("require_previous_by_kind", {})
    require_previous = bool(by_kind.get(target.kind, spec.get("require_previous", True)))
    if target.kind not in spec.get("kinds", []):
        return {"sheet": spec["name"], "state": "skipped", "reason": f"该 sheet 不更新 {target.kind}"}
    remaining = dict(periods)
    src = source_period(target, remaining, require_previous=require_previous)
    header_row = int(spec["header_row"])
    while src is not None and not _column_has_content(ws, header_row, remaining[src]):
        remaining.pop(src, None)
        src = source_period(target, remaining, require_previous=require_previous)
    if src is None:
        return {"sheet": spec["name"], "state": "skipped",
                "reason": "没有可比期间列", "optional": bool(spec.get("optional"))}
    src_col = periods[src]
    target_col = periods.get(target)
    insert = target_col is None
    if target_col is None:
        same_cols = [col for period, col in periods.items() if period.kind == target.kind]
        if not same_cols:
            return {"sheet": spec["name"], "state": "skipped", "reason": "模板没有该期间类型"}
        target_col = max(same_cols) + 1
    return {
        "sheet": spec["name"], "header_row": header_row,
        "source_period": src.key, "source_col": src_col, "target_period": target.key,
        "target_col": target_col, "insert": insert, "optional": bool(spec.get("optional")),
        "already_exists": not insert,
    }


def _echo_role(value, source: Period) -> str | None:
    period = Period.parse(value)
    if period == source:
        return "echo"
    if isinstance(value, str) and value.strip() and Period.parse(value) is None:
        text = value.strip()
        if text.isalpha() and 2 <= len(text) <= 24:
            return "echo"
    return None


def build_template(model: Path, contract: Contract, target: Period, pdfs: list[Path]) -> dict:
    wb = load_workbook(model, read_only=True, data_only=False, keep_links=True)
    sheet_plans, facts = [], []
    try:
        for spec in contract.sheets:
            if spec["name"] not in wb.sheetnames:
                if spec.get("optional"):
                    sheet_plans.append({"sheet": spec["name"], "state": "skipped", "reason": "可选 sheet 不存在"})
                    continue
                raise ModelError(f"缺少工作表：{spec['name']}")
            ws = wb[spec["name"]]
            sheet_plan = _sheet_plan(ws, spec, target)
            sheet_plans.append(sheet_plan)
            if sheet_plan.get("state") == "skipped":
                continue
            src_col = sheet_plan["source_col"]
            src = Period.parse(sheet_plan["source_period"])
            default_pdf = pdfs[0] if len(pdfs) == 1 else None
            for row in range(int(spec["header_row"]) + 1, (ws.max_row or 0) + 1):
                cell = ws.cell(row, src_col)
                value = cell.value
                if value in (None, "") or _is_formula(value):
                    continue
                role = _echo_role(value, src) or "disclosed"
                facts.append({
                    "sheet": spec["name"], "row": row, "label": _label(ws, row),
                    "source_period": sheet_plan["source_period"], "previous_value": value,
                    "value": value if role == "echo" and _echo_role(value, src) == "echo"
                    and Period.parse(value) is None else None,
                    "unit": None, "role": role,
                    "source": {
                        "file": str(default_pdf) if default_pdf else None,
                        "sha256": sha256(default_pdf) if default_pdf else None,
                        "page": None, "table": None, "row": None,
                        "quote": None, "value_text": None,
                        "numeric_value": None, "unit": None, "conversion_factor": 1,
                    },
                })
    finally:
        wb.close()
    return {
        "company": contract.company, "period": target.key, "model": str(model),
        "pdfs": [str(path) for path in pdfs], "sheet_plans": sheet_plans, "facts": facts,
        "instructions": (
            "只保留本期披露且需要硬编码写入的行。公式行由 Excel 从上一可比列平移，不要手填。"
            "value 必须是数字，不要写单元格引用。echo 行可由系统按上一列自动改期间标签。"
            "每个 disclosed 数必须带 PDF 页码、表/行、原话、value_text 和单位换算。"
        ),
    }


def harvest_facts(model: Path, contract: Contract, target: Period) -> dict:
    """从已有目标列收割硬编码值，供 holdout 自测。role=replay，不走 PDF 门。"""
    wb = load_workbook(model, read_only=True, data_only=False, keep_links=True)
    facts, sheet_plans = [], []
    try:
        for spec in contract.sheets:
            if spec["name"] not in wb.sheetnames:
                continue
            ws = wb[spec["name"]]
            plan = _sheet_plan(ws, spec, target)
            sheet_plans.append(plan)
            if plan.get("state") == "skipped" or plan.get("insert"):
                continue
            col = plan["target_col"]
            src_col = plan["source_col"]
            src = Period.parse(plan["source_period"])
            for row in range(int(spec["header_row"]) + 1, (ws.max_row or 0) + 1):
                if _is_formula(ws.cell(row, src_col).value):
                    continue
                value = ws.cell(row, col).value
                if value in (None, "") or _is_formula(value):
                    continue
                role = _echo_role(value, target) or "replay"
                facts.append({
                    "sheet": spec["name"], "row": row, "label": _label(ws, row),
                    "source_period": plan["source_period"], "previous_value": None,
                    "value": value, "unit": "model", "role": "replay" if role != "echo" else "echo",
                    "source": {},
                })
    finally:
        wb.close()
    if not facts:
        raise ModelError(f"{contract.company} {target.key} 没有可收割的硬编码单元格")
    return {"company": contract.company, "period": target.key, "model": str(model),
            "pdfs": [], "sheet_plans": sheet_plans, "facts": facts}


def build_plan(model: Path, contract: Contract, target: Period, facts: dict) -> dict:
    template = build_template(model, contract, target, [Path(p) for p in facts.get("pdfs", [])])
    fact_rows = {(str(item["sheet"]), int(item["row"])): item
                 for item in facts.get("facts", []) if item.get("value") not in (None, "")}
    allowed = contract.writable_sheets
    unknown = [f"{sheet}!{row}" for sheet, row in fact_rows if sheet not in allowed]
    if unknown:
        raise ModelError("facts 包含未授权工作表：" + ", ".join(unknown))
    if not fact_rows:
        raise ModelError("facts 没有任何待写入的硬编码数值")
    operations = []
    skipped_formula_writes = []
    allowed_input_rows = {
        (str(item["sheet"]), int(item["row"])) for item in template["facts"]
    }
    for sheet_plan in template["sheet_plans"]:
        if sheet_plan.get("state") == "skipped":
            continue
        writes, dropped = [], []
        for (sheet, row), fact in fact_rows.items():
            if sheet != sheet_plan["sheet"]:
                continue
            if (sheet, row) not in allowed_input_rows:
                dropped.append({"sheet": sheet, "row": row, "reason": "上一可比列是公式，不能覆盖"})
                continue
            writes.append(fact)
        skipped_formula_writes.extend(dropped)
        operations.append({**sheet_plan, "writes": writes})
    if not operations:
        raise ModelError("没有可执行的工作表计划")
    if not any(item["writes"] for item in operations):
        raise ModelError("facts 没有对应到上一列硬编码输入行")
    return {
        "company": contract.company, "period": target.key, "model": str(model),
        "operations": operations,
        "skipped": [item for item in template["sheet_plans"] if item.get("state") == "skipped"],
        "skipped_formula_writes": skipped_formula_writes,
        "chart_sheets": list(contract.charts_for(target.kind)),
        "chart_policy": "2019 同期起；排除 2020-2022；接 2023 至目标期间；只标同期点",
    }


def _style_signature(cell) -> dict:
    font = cell.Font
    def _color(value):
        try:
            return int(value)
        except (TypeError, ValueError):
            return value
    return {
        "number_format": str(cell.NumberFormat),
        "font": [font.Name, font.Size, bool(font.Bold), bool(font.Italic), _color(font.Color)],
        "interior": _color(cell.Interior.Color),
        "align": [cell.HorizontalAlignment, cell.VerticalAlignment],
        "locked": cell.Locked,
    }


def _error_count(wb) -> int:
    count = 0
    for ws in wb.Worksheets:
        used = ws.UsedRange
        values = used.Formula
        stack = [values]
        while stack:
            item = stack.pop()
            if isinstance(item, tuple):
                stack.extend(item)
            elif isinstance(item, str) and any(
                token in item for token in ("#REF!", "#VALUE!", "#DIV/0!", "#NAME?")
            ):
                count += 1
    return count


def _period_map_com(ws, header_row: int) -> dict[Period, int]:
    used_end = int(ws.UsedRange.Column) + int(ws.UsedRange.Columns.Count) - 1
    found = {}
    for col in range(1, used_end + 1):
        period = Period.parse(ws.Cells(header_row, col).Value)
        if period is not None:
            found[period] = col
    return found


def _equal(expected, actual) -> bool:
    expected, actual = _coerce_value(expected), actual
    if isinstance(expected, (int, float)) and isinstance(actual, (int, float)):
        return abs(float(expected) - float(actual)) <= max(1e-6, abs(float(expected)) * 1e-6)
    if expected is None and actual in (None, ""):
        return True
    return expected == actual


def _excel():
    import pythoncom
    import win32com.client as win32
    pythoncom.CoInitialize()
    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False
    return pythoncom, excel


def delete_period_columns(path: Path, contract: Contract, target: Period) -> dict:
    """从副本删除目标期间列，用于 holdout。不碰权威文件。"""
    pythoncom, excel = _excel()
    wb = None
    deleted = []
    try:
        wb = excel.Workbooks.Open(str(path.resolve()), UpdateLinks=0, ReadOnly=False)
        for spec in contract.sheets:
            if spec["name"] not in contract.writable_sheets:
                continue
            try:
                ws = wb.Worksheets(spec["name"])
            except Exception:
                continue
            mapping = _period_map_com(ws, int(spec["header_row"]))
            col = mapping.get(target)
            if col is None:
                continue
            ws.Columns(col).Delete()
            deleted.append({"sheet": spec["name"], "col": col})
        wb.Save()
        wb.Close(SaveChanges=True)
        wb = None
        excel.Quit()
        excel = None
        return {"deleted": deleted}
    finally:
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass
        pythoncom.CoUninitialize()


def apply_plan(source: Path, destination: Path, contract: Contract, target: Period, plan: dict) -> dict:
    """只修改新 destination；保存后关闭并重新打开源文件和副本做最终审计。"""
    if destination.resolve() == source.resolve():
        raise ModelError("输出路径不能与权威 Model 相同")
    destination.parent.mkdir(parents=True, exist_ok=True)
    if destination.exists():
        raise ModelError(f"输出已存在，拒绝覆盖：{destination}")
    shutil.copy2(source, destination)

    pythoncom, excel = _excel()
    wb = source_wb = verify_wb = None
    errors, checks, column_audits = [], [], []
    baseline_errors = final_errors = 0
    chart_result: dict = {}
    chart_audit: dict = {}
    try:
        source_wb = excel.Workbooks.Open(str(source.resolve()), UpdateLinks=0, ReadOnly=True)
        baseline_errors = _error_count(source_wb)
        source_wb.Close(False)
        source_wb = None

        wb = excel.Workbooks.Open(str(destination.resolve()), UpdateLinks=0, ReadOnly=False)
        operation_audits = []
        for operation in plan["operations"]:
            sheet_name = operation["sheet"]
            if sheet_name not in contract.writable_sheets:
                raise ModelError(f"未授权写入工作表：{sheet_name}")
            ws = wb.Worksheets(sheet_name)
            header_row = int(operation["header_row"])
            source_col = int(operation["source_col"])
            target_col = int(operation["target_col"])
            source_header = ws.Cells(header_row, source_col).Value
            used_end = int(ws.UsedRange.Row) + int(ws.UsedRange.Rows.Count) - 1
            source_period_obj = Period.parse(operation["source_period"])
            if operation["insert"]:
                ws.Columns(target_col).Insert(Shift=-4161)
                ws.Columns(source_col).Copy(Destination=ws.Columns(target_col))
                for row in range(header_row + 1, used_end + 1):
                    cell = ws.Cells(row, target_col)
                    try:
                        has_formula = bool(cell.HasFormula)
                    except Exception:
                        has_formula = False
                    src_value = ws.Cells(row, source_col).Value
                    if has_formula:
                        continue
                    if Period.parse(src_value) == source_period_obj:
                        cell.Value = target.label_like(src_value)
                    elif cell.Value not in (None, ""):
                        cell.Value = None
            ws.Cells(header_row, target_col).Value = target.label_like(source_header)
            for fact in operation["writes"]:
                row = int(fact["row"])
                src_cell = ws.Cells(row, source_col)
                target_cell = ws.Cells(row, target_col)
                try:
                    source_has_formula = bool(src_cell.HasFormula)
                except Exception:
                    source_has_formula = False
                if source_has_formula:
                    continue
                src_cell.Copy()
                target_cell.PasteSpecial(Paste=-4122)
                target_cell.Value = _coerce_value(fact["value"])
                checks.append({
                    "sheet": sheet_name, "row": row, "col": target_col,
                    "expected": _coerce_value(fact["value"]),
                    "style_source_col": source_col, "source": fact.get("source"),
                })
            operation_audits.append({
                "sheet": sheet_name, "header_row": header_row,
                "source_col": source_col, "target_col": target_col,
                "used_end": used_end, "inserted": bool(operation["insert"]),
            })

        try:
            excel.CutCopyMode = False
        except Exception:
            pass
        period_maps = {}
        for spec in contract.sheets:
            try:
                ws = wb.Worksheets(spec["name"])
            except Exception:
                continue
            period_maps[spec["name"]] = _period_map_com(ws, int(spec["header_row"]))
        excel.CalculateFullRebuild()
        chart_result = update_chart_sheets(wb, contract, target, period_maps)
        errors.extend(chart_result.get("errors", []))
        excel.CalculateFullRebuild()
        wb.Save()
        wb.Close(SaveChanges=True)
        wb = None

        source_wb = excel.Workbooks.Open(str(source.resolve()), UpdateLinks=0, ReadOnly=True)
        verify_wb = excel.Workbooks.Open(str(destination.resolve()), UpdateLinks=0, ReadOnly=True)
        final_errors = _error_count(verify_wb)
        if final_errors > baseline_errors:
            errors.append(f"工作簿错误值由 {baseline_errors} 增至 {final_errors}")

        for operation in operation_audits:
            sheet_name = operation["sheet"]
            source_ws = source_wb.Worksheets(sheet_name)
            output_ws = verify_wb.Worksheets(sheet_name)
            header_row = operation["header_row"]
            source_col = operation["source_col"]
            target_col = operation["target_col"]
            actual_period = Period.parse(output_ws.Cells(header_row, target_col).Value)
            if actual_period != target:
                errors.append(f"{sheet_name}!{get_column_letter(target_col)}{header_row} 期间表头不正确")

            style_mismatches, formula_mismatches = [], []
            if operation["inserted"]:
                for row in range(header_row + 1, operation["used_end"] + 1):
                    source_cell = source_ws.Cells(row, source_col)
                    target_cell = output_ws.Cells(row, target_col)
                    if _style_signature(source_cell) != _style_signature(target_cell):
                        style_mismatches.append(row)
                    try:
                        source_has_formula = bool(source_cell.HasFormula)
                        target_has_formula = bool(target_cell.HasFormula)
                    except Exception:
                        source_has_formula = target_has_formula = False
                    if source_has_formula:
                        same_formula = target_has_formula and (
                            str(source_cell.FormulaR1C1) == str(target_cell.FormulaR1C1)
                        )
                        if not same_formula:
                            formula_mismatches.append(row)
                if style_mismatches:
                    errors.append(
                        f"{sheet_name} 新列有 {len(style_mismatches)} 行格式未继承；"
                        f"首批行：{style_mismatches[:10]}"
                    )
                if formula_mismatches:
                    errors.append(
                        f"{sheet_name} 新列有 {len(formula_mismatches)} 行公式未按 R1C1 平移；"
                        f"首批行：{formula_mismatches[:10]}"
                    )
            column_audits.append({
                "sheet": sheet_name, "source_col": source_col, "target_col": target_col,
                "rows_checked": max(0, operation["used_end"] - header_row),
                "style_mismatches": style_mismatches,
                "formula_mismatches": formula_mismatches,
            })

        for check in checks:
            source_cell = source_wb.Worksheets(check["sheet"]).Cells(
                check["row"], check["style_source_col"])
            target_cell = verify_wb.Worksheets(check["sheet"]).Cells(check["row"], check["col"])
            check["actual"] = target_cell.Value
            check["value_ok"] = _equal(check["expected"], check["actual"])
            check["style_ok"] = _style_signature(source_cell) == _style_signature(target_cell)
            check["ok"] = check["value_ok"] and check["style_ok"]
            if not check["value_ok"]:
                errors.append(f"{check['sheet']}!{get_column_letter(check['col'])}{check['row']} 回读不一致")
            if not check["style_ok"]:
                errors.append(f"{check['sheet']}!{get_column_letter(check['col'])}{check['row']} 格式未继承")

        period_maps = {}
        for spec in contract.sheets:
            try:
                ws = verify_wb.Worksheets(spec["name"])
            except Exception:
                continue
            period_maps[spec["name"]] = _period_map_com(ws, int(spec["header_row"]))
        chart_audit = audit_chart_sheets(verify_wb, contract, target, period_maps)
        errors.extend(chart_audit.get("errors", []))
        verify_wb.Close(False)
        verify_wb = None
        source_wb.Close(False)
        source_wb = None
        excel.Quit()
        excel = None
        return {
            "passed": not errors, "source": str(source), "output": str(destination),
            "checks": checks, "column_audits": column_audits,
            "charts": {"update": chart_result, "readback": chart_audit},
            "baseline_errors": baseline_errors, "final_errors": final_errors, "errors": errors,
        }
    except Exception as error:
        errors.append(f"{type(error).__name__}: {error}")
        raise ModelError(errors[-1]) from error
    finally:
        for book in (verify_wb, wb, source_wb):
            try:
                if book is not None:
                    book.Close(SaveChanges=False)
            except Exception:
                pass
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass
        pythoncom.CoUninitialize()
