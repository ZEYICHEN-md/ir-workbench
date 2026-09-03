"""用往期 Model 做 holdout：删掉目标列再按收割值重写，核对硬编码与图表政策。"""
from __future__ import annotations

import json
import shutil
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from workbench.config import Config
from workbench.result import Result

from . import excel_model
from .contracts import ContractError, load, workbook
from .periods import Period, scan_periods, source_period

DOMAIN = "peers-model"
COMPANIES = ("BKNG", "EXPE", "ABNB", "MEITUAN", "TCEL")


def _hardcoded(path: Path, contract, target: Period) -> dict[tuple[str, int], object]:
    wb = load_workbook(path, data_only=False, keep_links=True)
    found = {}
    try:
        for spec in contract.sheets:
            if spec["name"] not in wb.sheetnames:
                continue
            ws = wb[spec["name"]]
            periods = scan_periods(ws, int(spec["header_row"]))
            col = periods.get(target)
            if col is None:
                continue
            src = source_period(target, periods, require_previous=False)
            src_col = periods.get(src) if src is not None else None
            for row in range(int(spec["header_row"]) + 1, (ws.max_row or 0) + 1):
                if src_col is not None and excel_model._is_formula(ws.cell(row, src_col).value):
                    continue
                value = ws.cell(row, col).value
                if value in (None, "") or excel_model._is_formula(value):
                    continue
                found[(spec["name"], row)] = value
    finally:
        wb.close()
    return found


def run_one(paths, company: str, period_text: str) -> dict:
    contract = load(company)
    period = Period.parse(period_text)
    if period is None:
        raise ValueError("期间格式应为 26Q3、26H1 或 FY2026")
    gold = workbook(contract, Config(paths))
    inspection = excel_model.inspect_workbook(gold, contract)
    if inspection["missing"]:
        raise excel_model.ModelError("缺少工作表：" + "、".join(inspection["missing"]))
    harvest = excel_model.harvest_facts(gold, contract, period)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S")
    folder = paths.outputs(DOMAIN, f"{contract.company}-{period.key}") / "selftest" / stamp
    folder.mkdir(parents=True, exist_ok=True)
    prior = folder / f"{gold.stem}_prior{gold.suffix}"
    output = folder / f"{gold.stem}_{period.key}_replay{gold.suffix}"
    shutil.copy2(gold, prior)
    deleted = excel_model.delete_period_columns(prior, contract, period)
    gold_values = _hardcoded(gold, contract, period)
    if _hardcoded(prior, contract, period):
        raise excel_model.ModelError("删除目标列后仍能读到该期间，holdout 无效")
    plan = excel_model.build_plan(prior, contract, period, harvest)
    (folder / "harvest.json").write_text(json.dumps(harvest, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    (folder / "plan.json").write_text(json.dumps(plan, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    audit = excel_model.apply_plan(prior, output, contract, period, plan)
    replay_values = _hardcoded(output, contract, period)
    missing = []
    mismatched = []
    for key, expected in gold_values.items():
        actual = replay_values.get(key)
        if actual is None:
            missing.append({"sheet": key[0], "row": key[1], "expected": expected})
        elif not excel_model._equal(expected, actual):
            mismatched.append({"sheet": key[0], "row": key[1], "expected": expected, "actual": actual})
    extra = [
        {"sheet": sheet, "row": row, "actual": value}
        for (sheet, row), value in replay_values.items()
        if (sheet, row) not in gold_values
    ]
    errors = list(audit.get("errors") or [])
    if missing:
        errors.append(f"{len(missing)} 个金标准硬编码格缺失")
    if mismatched:
        errors.append(f"{len(mismatched)} 个金标准硬编码格数值不符")
    passed = not errors
    report = {
        "company": contract.company, "period": period.key, "passed": passed,
        "gold": str(gold), "prior": str(prior), "output": str(output),
        "deleted": deleted, "harvested": len(harvest["facts"]),
        "gold_hardcoded": len(gold_values), "replay_hardcoded": len(replay_values),
        "missing": missing[:20], "mismatched": mismatched[:20], "extra": extra[:20],
        "audit": audit, "errors": errors,
    }
    (folder / "report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8"
    )
    report["report"] = str(folder / "report.json")
    return report


def run(paths, company: str | None, period_text: str) -> Result:
    names = COMPANIES if not company or company.upper() == "ALL" else (company,)
    reports = []
    try:
        for name in names:
            reports.append(run_one(paths, name, period_text))
    except (ContractError, ValueError, excel_model.ModelError) as error:
        return Result(
            status="failed", summary=str(error), domain=DOMAIN,
            data={"reports": reports},
        )
    failed = [item["company"] for item in reports if not item["passed"]]
    status = "success" if not failed else "failed"
    summary = (
        f"holdout 全部通过：{', '.join(item['company'] for item in reports)}。"
        if not failed else
        f"holdout 未通过：{', '.join(failed)}。原始 Model 未改。"
    )
    return Result(
        status=status, summary=summary, domain=DOMAIN,
        checks=[{
            "name": item["company"],
            "level": "ok" if item["passed"] else "fail",
            "detail": item.get("report", ""),
        } for item in reports],
        warnings=[] if not failed else ["; ".join(item["errors"][:5]) for item in reports if not item["passed"]],
        data={"reports": reports},
    )
