"""Adversarial checks on the generated Investor List (formulas, not Excel cache)."""
from __future__ import annotations

import json
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from shareholder_list.build import (
    COMBINED_KEEP,
    HOLDINGS_AS_OF,
    MAP_D8,
    REGION_PCT_COLS,
    STRATEGIC,
    VALID_AS_OF,
    WORKBOOK_AS_OF,
    header_row_of,
    num,
    read_table,
)


def _is_pct(cell) -> bool:
    return "%" in str(cell.number_format or "")


def _fill_rgb(cell) -> str:
    fill = cell.fill
    if fill is None or fill.fill_type in (None, "none"):
        return ""
    fg = fill.fgColor
    return str(getattr(fg, "rgb", "") or "")


def _names_from_col(ws, start_row: int, col: int, stop: frozenset[str] = frozenset({"TOTAL"})) -> list[str]:
    out: list[str] = []
    r = start_row
    while True:
        v = ws.cell(r, col).value
        if not isinstance(v, str) or not v.strip() or v.strip() in stop:
            break
        out.append(v)
        r += 1
    return out


def validate(output: Path, peer: Path, combined: Path, market_path: Path) -> dict[str, Any]:
    failures: list[str] = []
    market = json.loads(market_path.read_text(encoding="utf-8"))
    out = load_workbook(output, data_only=False)
    pwb = load_workbook(peer, data_only=False)
    cwb = load_workbook(combined, data_only=False)
    peer_ws = pwb.active
    comb_ws = cwb.active
    ph, pr = read_table(peer_ws, header_row_of(peer_ws))
    ch, cr = read_table(comb_ws, header_row_of(comb_ws))
    pwb.close()
    cwb.close()
    pmap = {h: i for i, h in enumerate(ph)}
    cmap = {h: i for i, h in enumerate(ch)}

    da = out["DATA_ALL"]
    da_headers = [da.cell(2, c).value for c in range(1, 100)]
    if da.cell(2, 1).value != "Institution":
        failures.append("DATA_ALL R2A is not Institution")
    if da.cell(2, 3).value != "Value TCOM USD, mm":
        failures.append(f"DATA_ALL col3 is {da.cell(2,3).value}, expected Value TCOM")
    if da.cell(2, 6).value != "Shares TCOM":
        failures.append(f"DATA_ALL col6 is {da.cell(2,6).value}")
    peer_names = [r[0] for r in pr]
    da_names = _names_from_col(da, 3, 1)
    if da_names != peer_names:
        failures.append(f"DATA_ALL names {len(da_names)} vs peer {len(peer_names)}")
    sum_row = 3 + len(da_names)
    if da.cell(sum_row, 1).value != "TOTAL":
        failures.append(f"DATA_ALL sum row missing at {sum_row}, got {da.cell(sum_row,1).value}")
    if da.cell(sum_row + 1, 1).value not in (None, ""):
        failures.append(f"DATA_ALL leftover under TOTAL: {da.cell(sum_row + 1, 1).value}")
    tcom_col = da_headers.index("Value TCOM USD, mm") + 1
    if not str(da.cell(sum_row, tcom_col).value or "").startswith("=SUM("):
        failures.append("DATA_ALL TCOM value SUM missing")
    fund_col = da_headers.index("Fund Size") + 1 if "Fund Size" in da_headers else 98
    if not str(da.cell(3, fund_col).value or "").startswith("=IF("):
        failures.append("DATA_ALL Fund Size formula missing")

    cb = out["Data_TCOM+9961"]
    comb_names = [row[0] for row in cr]
    cb_names = _names_from_col(cb, 4, 1)
    if cb_names != comb_names:
        failures.append(f"combined names {len(cb_names)} vs extract {len(comb_names)}")
    if cb_names and cb_names[0] != "Baidu Holdings, LTD":
        failures.append(f"combined #1 is {cb_names[0]}")
    for name in COMBINED_KEEP:
        if name not in ch:
            failures.append(f"combined extract missing {name}")
    tcom_i, hk_i, tot_i = cmap["TCOM Shares"], cmap["9961_XHKG Shares"], cmap["Shares (Combined)"]
    for row in cr:
        if abs(num(row[tot_i]) - num(row[tcom_i]) - num(row[hk_i])) > 1:
            failures.append(f"combined share identity fail {row[0]}")
            break
    if cb.cell(4, 19).value in (None, "") or not str(cb.cell(4, 19).value).startswith("="):
        failures.append("combined Fund Size formula missing")
    for c, h in enumerate(COMBINED_KEEP, 1):
        if cb.cell(3, c).value != h:
            failures.append(f"combined header {c} {cb.cell(3,c).value} != {h}")
            break

    sh = out["SH Summary"]
    sh_names = _names_from_col(sh, 6, 3)
    expected = [n for n in comb_names if n not in STRATEGIC]
    if sh_names != expected:
        failures.append(f"SH names {len(sh_names)} vs combined-strategic {len(expected)}")
        missing = list(set(expected) - set(sh_names))[:5]
        if missing:
            failures.append("SH missing " + ", ".join(missing))
    if any(n in STRATEGIC for n in sh_names):
        failures.append("strategic holder leaked into SH Summary")
    if sh.cell(6, 2).value != 1:
        failures.append("SH B6 rank is not 1")
    if sh["F3"].value != market["tcom_shares_outstanding"]:
        failures.append(f"F3 {sh['F3'].value} != market SO {market['tcom_shares_outstanding']}")
    if sh["G3"].value != f"updated as of {VALID_AS_OF}":
        failures.append(f"SH G3 {sh['G3'].value!r} is not validity as-of {VALID_AS_OF}")
    if "TBU" in str(sh["G3"].value or ""):
        failures.append("SH G3 still has TBU")
    if sh.cell(4, 4).value != 7 or sh.cell(4, 5).value != 10:
        failures.append("SH VLOOKUP indices D/E changed")
    f6 = str(sh.cell(6, 6).value or "")
    if "D6" not in f6:
        failures.append("SH F6 is not D+E")
    a6 = str(sh.cell(6, 1).value or "")
    if "[1]SH Summary" in a6:
        failures.append("SH still uses external workbook link")
    if "SH Prior" not in a6:
        failures.append("SH A6 does not point at SH Prior")

    if "SH Prior" not in out.sheetnames:
        failures.append("missing SH Prior")
    elif out["SH Prior"].cell(2, 1).value is None:
        failures.append("SH Prior empty")
    elif out["SH Prior"].sheet_state != "hidden":
        failures.append("SH Prior is not hidden")
    if sh.freeze_panes != "D6":
        failures.append(f"SH freeze_panes {sh.freeze_panes} != D6 (June leftover F265)")

    t20 = out["Top 20"]
    if sh_names and t20.cell(35, 2).value != sh_names[0]:
        failures.append(f"Top20 current #1 {t20.cell(35,2).value} != SH {sh_names[0]}")
    if "Baidu" in str(t20.cell(35, 2).value):
        failures.append("Top20 starts with Baidu")
    if "IFERROR" not in str(t20.cell(11, 3).value or ""):
        failures.append("Top 20 prior-block INDEX is not IFERROR-wrapped")
    if "TBU" in str(t20.cell(2, 2).value or ""):
        failures.append("Top 20 title still has TBU")
    if t20.cell(2, 2).value != "核心股东持仓情况":
        failures.append(f"Top 20 B2 {t20.cell(2,2).value!r} != May")
    b4 = str(t20["B4"].value or "")
    if not b4.startswith("需关注"):
        failures.append(f"Top 20 B4 takeaways stale/missing: {b4[:80]!r}")
    if t20["B6"].value not in (None, ""):
        failures.append("Top 20 B6 still has leftover note")
    if sh_names and "E2EEDA" in _fill_rgb(sh.cell(6, 3)):
        failures.append("SH #1 should not be Index green")
    idx_probe = "BlackRock Fund Advisors"
    if idx_name := next((n for n in sh_names if n == idx_probe), None):
        r = 6 + sh_names.index(idx_name)
        if "E2EEDA" not in _fill_rgb(sh.cell(r, 3)):
            failures.append("Index holder missing green fill on SH")
    for r in range(35, 55):
        name = t20.cell(r, 2).value
        if name == idx_probe and "E2EEDA" not in _fill_rgb(t20.cell(r, 2)):
            failures.append("Index holder missing green fill on Top 20")
            break
        if name == "Capital World Investors (U.S.)" and "E2EEDA" in _fill_rgb(t20.cell(r, 2)):
            failures.append("Growth holder CWI should not be Index green on Top 20")
            break
    t20_pct_fail = 0
    for r in list(range(11, 31)) + list(range(35, 55)):
        if not isinstance(t20.cell(r, 2).value, str):
            continue
        for c, label in ((6, "F"), (7, "G"), (8, "H")):
            if not _is_pct(t20.cell(r, c)):
                t20_pct_fail += 1
                if t20_pct_fail <= 5:
                    failures.append(
                        f"Top 20 {label}{r} number_format {t20.cell(r, c).number_format!r} is not percent"
                    )
    if t20_pct_fail > 5:
        failures.append(f"Top 20 {t20_pct_fail} percent-format misses (showing first 5)")
    sh_pct_fail = 0
    for i, _name in enumerate(sh_names):
        r = 6 + i
        for c, label in ((11, "K"), (12, "L")):
            if not _is_pct(sh.cell(r, c)):
                sh_pct_fail += 1
                if sh_pct_fail <= 5:
                    failures.append(
                        f"SH {label}{r} number_format {sh.cell(r, c).number_format!r} is not percent"
                    )
    if sh_pct_fail > 5:
        failures.append(f"SH K/L {sh_pct_fail} percent-format misses (showing first 5)")
    for region in ("US&CA", "EU", "APAC", "ROW"):
        ws = out[region]
        r = 8
        region_miss = 0
        while isinstance(ws.cell(r, 3).value, str) and ws.cell(r, 3).value.strip():
            for c in REGION_PCT_COLS:
                if not _is_pct(ws.cell(r, c)):
                    region_miss += 1
                    if region_miss <= 3:
                        failures.append(
                            f"{region} {get_column_letter(c)}{r} number_format "
                            f"{ws.cell(r, c).number_format!r} is not percent"
                        )
            r += 1
        if region_miss > 3:
            failures.append(f"{region} {region_miss} percent-format misses (showing first 3)")

    us = out["US&CA"]
    if us.cell(5, 7).value != tcom_col:
        failures.append(f"US&CA G5 {us.cell(5,7).value} != TCOM value col {tcom_col}")
    hk_col = da_headers.index("Value 9961_XHKG USD, mm") + 1
    if us.cell(5, 8).value != hk_col:
        failures.append(f"US&CA H5 {us.cell(5,8).value} != 9961 value col {hk_col}")
    if us.cell(5, 4).value != da_headers.index("Style") + 1:
        failures.append("US&CA Style index wrong")
    if us.cell(5, 5).value != da_headers.index("Equity Assets USD, mm") + 1:
        failures.append("US&CA AUM index wrong")
    if not isinstance(us.cell(8, 3).value, str):
        failures.append("US&CA has no names")
    if us.cell(5, 12).value != da_headers.index("Value 700_XHKG USD, mm") + 1:
        failures.append("US&CA L5 Tencent col wrong")
    if abs(num(us.cell(5, 10).value) - num(market["usd_market_cap"]["TCOM"])) > 1:
        failures.append("US&CA J5 TCOM mcap mismatch")
    for region in ("US&CA", "EU", "APAC", "ROW"):
        ws = out[region]
        if ws.cell(5, 7).value != tcom_col or ws.cell(5, 8).value != hk_col:
            failures.append(f"{region} G5/H5 column indices drifted")
        if "G5274" in str(ws.cell(8, 7).value or ""):
            failures.append(f"{region} still hardcodes G5274")
        b3 = str(ws["B3"].value or "")
        if WORKBOOK_AS_OF not in b3 or HOLDINGS_AS_OF not in b3:
            failures.append(f"{region} B3 missing validity/holdings dates: {b3}")
        if ws.freeze_panes != "D8":
            failures.append(f"{region} freeze_panes {ws.freeze_panes} != D8 (June leftover)")

    mp = out["全球投资人地图"]
    if mp["C5"].value != len(peer_names):
        failures.append(f"map C5 {mp['C5'].value} != {len(peer_names)}")
    c8 = str(mp["C8"].value or "")
    if get_column_letter(tcom_col) not in c8:
        failures.append(f"map C8 does not use TCOM col {get_column_letter(tcom_col)}: {c8}")
    if "G5274" in c8 or "I5274" in c8:
        failures.append("map still hardcodes old sum cells")
    if get_column_letter(hk_col) not in c8:
        failures.append(f"map C8 does not use 9961 col {get_column_letter(hk_col)}")
    if mp["D8"].value != MAP_D8:
        failures.append(f"map D8 footnote drifted: {mp['D8'].value!r}")
    if mp["E16"].value != "no ADR holding Equity AUM":
        failures.append("map E16 label drifted")
    if mp["E17"].value != "no Travel holding Equity AUM":
        failures.append("map E17 label drifted")
    if mp["E18"].value != "no TCOM holding":
        failures.append("map E18 label drifted")
    if mp["B3"].value != f"updated as of {WORKBOOK_AS_OF}":
        failures.append(f"map B3 {mp['B3'].value!r}")

    # CWI identity across both extracts and DATA_ALL / combined
    cwi = "Capital World Investors (U.S.)"
    peer_by = {row[0]: row for row in pr}
    comb_by = {row[0]: row for row in cr}
    if cwi not in peer_by or cwi not in comb_by:
        failures.append("CWI missing from an extract")
    else:
        pt, ct = peer_by[cwi], comb_by[cwi]
        if abs(num(pt[pmap["Shares TCOM"]]) - num(ct[cmap["TCOM Shares"]])) > 1:
            failures.append("CWI TCOM shares peer != combined")
        if abs(num(pt[pmap["Shares 9961_XHKG"]]) - num(ct[cmap["9961_XHKG Shares"]])) > 1:
            failures.append("CWI 9961 shares peer != combined")
        da_row = 3 + peer_names.index(cwi)
        if abs(num(da.cell(da_row, 6).value) - num(pt[pmap["Shares TCOM"]])) > 1:
            failures.append("CWI DATA_ALL Shares TCOM mismatch")
        so = num(market["tcom_shares_outstanding"])
        expected_so = (num(ct[cmap["TCOM Shares"]]) + num(ct[cmap["9961_XHKG Shares"]])) / so
        if expected_so <= 0.04 or expected_so >= 0.07:
            failures.append(f"CWI implied %S/O {expected_so:.4f} outside 4-7% band")

    if "Baidu Holdings, LTD" in sh_names:
        failures.append("Baidu in SH names")
    if sh_names and sh_names[0] != expected[0]:
        failures.append(f"SH #1 {sh_names[0]} != {expected[0]}")

    missing = [n for n in comb_names if n not in set(peer_names)]
    if missing:
        failures.append(f"{len(missing)} combined names missing from peer e.g. {missing[:3]}")

    if out["重要投资人"].cell(1, 1).value != "分类":
        failures.append("重要投资人 header damaged")

    if "Market" not in out.sheetnames:
        failures.append("missing Market sheet")
    else:
        if abs(num(out["Market"]["B2"].value) - num(market["fx_usd_hkd"])) > 1e-6:
            failures.append("Market USD/HKD mismatch")
        if out["Market"]["B1"].value != VALID_AS_OF.replace("/", "-"):
            failures.append(f"Market B1 {out['Market']['B1'].value} != valid {VALID_AS_OF}")
        if out["Market"].sheet_state != "hidden":
            failures.append("Market is not hidden")

    if da.freeze_panes != "B3":
        failures.append(f"DATA_ALL freeze_panes {da.freeze_panes}")

    cwi_shares = None
    if cwi in comb_by:
        cwi_shares = num(comb_by[cwi][cmap["Shares (Combined)"]])
    spot_check = {
        "g3": sh["G3"].value,
        "market_b1": out["Market"]["B1"].value if "Market" in out.sheetnames else None,
        "map_b3": mp["B3"].value,
        "combined_1": cb_names[0] if cb_names else None,
        "sh_1": sh_names[0] if sh_names else None,
        "cwi": cwi,
        "cwi_combined_shares": cwi_shares,
    }

    for sheet in out.worksheets:
        for row in sheet.iter_rows(max_row=min(sheet.max_row or 1, 40), max_col=min(sheet.max_column or 1, 25)):
            for cell in row:
                if isinstance(cell.value, str) and "[1]" in cell.value:
                    failures.append(f"external link leftover {sheet.title}!{cell.coordinate}")
                    break
            else:
                continue
            break

    out.close()
    with zipfile.ZipFile(output) as zf:
        ext = [n for n in zf.namelist() if n.replace("\\", "/").startswith("xl/externalLinks/")]
        if ext:
            failures.append(f"xlsx still contains {ext}")

    return {
        "ok": len(failures) == 0,
        "failures": failures,
        "counts": {
            "peer": len(peer_names),
            "combined": len(comb_names),
            "sh": len(sh_names),
        },
        "spot_check": spot_check,
    }
